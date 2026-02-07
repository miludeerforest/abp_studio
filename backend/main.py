import base64
import asyncio
import httpx
import os
import json
import logging
import subprocess
import uuid
import random
import time
import zipfile
import tempfile
import re
from pathlib import Path
# Fix for Starlette/python-multipart strict limits
try:
    # Patch python-multipart (if applicable)
    import multipart
    # Try multiple paths for safety
    try:
        multipart.multipart.MAX_MEMORY_FILE_SIZE = 100 * 1024 * 1024
        multipart.multipart.MAX_FILE_SIZE = 100 * 1024 * 1024
    except:
        pass

    # Patch Starlette (Critical for "Part exceeded" error)
    from starlette.formparsers import MultiPartParser
    MultiPartParser.max_file_size = 100 * 1024 * 1024 # 100MB
    MultiPartParser.max_part_size = 100 * 1024 * 1024 # 100MB
except ImportError:
    pass

import logging
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException, Depends, status, BackgroundTasks, WebSocket, WebSocketDisconnect, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from sqlalchemy import create_engine, Column, String, Integer, DateTime, JSON, Text, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from datetime import datetime, timedelta, timezone

# Timezone: UTC+8 for China
CHINA_TZ = timezone(timedelta(hours=8))

def get_china_now():
    """Get current time in China timezone (UTC+8)."""
    return datetime.now(CHINA_TZ).replace(tzinfo=None)  # Remove tzinfo for DB compatibility

# Import WebSocket and Queue managers
from websocket_manager import connection_manager, init_websocket_manager, shutdown_websocket_manager
from queue_manager import get_task_queue, get_concurrency_limiter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Global Request Throttler (Anti-CF Rate Limit) ---
_request_timestamps = []
_request_lock = asyncio.Lock()
THROTTLE_WINDOW_SECONDS = 10  # Time window for rate limiting
THROTTLE_MAX_REQUESTS = 15     # Max requests per window

async def throttle_request():
    """
    Global request throttler to prevent Cloudflare 429 rate limits.
    Limits requests to THROTTLE_MAX_REQUESTS per THROTTLE_WINDOW_SECONDS.
    """
    async with _request_lock:
        current_time = time.time()
        # Clean up old timestamps outside the window
        _request_timestamps[:] = [t for t in _request_timestamps if current_time - t < THROTTLE_WINDOW_SECONDS]
        
        # If we've hit the limit, wait until the oldest request expires
        if len(_request_timestamps) >= THROTTLE_MAX_REQUESTS:
            oldest = _request_timestamps[0]
            wait = THROTTLE_WINDOW_SECONDS - (current_time - oldest) + random.uniform(0.5, 2.0)
            if wait > 0:
                logger.info(f"Rate limiting: waiting {wait:.1f}s before next request (anti-CF)")
                await asyncio.sleep(wait)
        
        # Record this request timestamp
        _request_timestamps.append(time.time())

from fastapi.staticfiles import StaticFiles
from starlette.staticfiles import StaticFiles as StarletteStaticFiles
from starlette.responses import Response, FileResponse
from PIL import Image
import io

# Custom StaticFiles with 7-day cache for videos and images
class CachedStaticFiles(StarletteStaticFiles):
    """StaticFiles with Cache-Control header for media files."""
    
    async def get_response(self, path: str, scope) -> Response:
        response = await super().get_response(path, scope)
        # Set 7-day cache for media files
        if path.endswith(('.mp4', '.webm', '.mov', '.avi', '.jpg', '.jpeg', '.png', '.gif', '.webp')):
            response.headers["Cache-Control"] = "public, max-age=604800"  # 7 days
        return response

app = FastAPI(title="Product Scene Generator API")

# Mount uploads directory with 7-day cache
os.makedirs("/app/uploads", exist_ok=True)
app.mount("/uploads", CachedStaticFiles(directory="/app/uploads"), name="uploads")

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Database Setup ---
# Read from environment variable (set in .env or docker-compose)
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://user:password@localhost:5432/db")

# Increase connection pool size to handle concurrent requests
engine = create_engine(
    DATABASE_URL,
    pool_size=30,           # Increased from 20 to handle more concurrent tasks
    max_overflow=50,        # Increased from 30 to allow more burst capacity
    pool_timeout=45,        # Reduced timeout to fail faster on exhaustion
    pool_recycle=900,       # Recycle connections every 15 minutes (was 30)
    pool_pre_ping=True,     # Test connection health before use
    pool_use_lifo=True      # LIFO helps reuse recently-active connections (better efficiency)
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

from passlib.context import CryptContext
from jose import JWTError, jwt

# --- Auth Config ---
# JWT secret key - MUST be set in .env for production
SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret-key-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 # 1 day

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# --- Data Models ---
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)  # Login name (cannot change)
    nickname = Column(String, nullable=True)  # Display name (can change)
    avatar = Column(String, nullable=True)  # Avatar URL path
    default_share = Column(Boolean, default=True)  # 默认开启分享，创作内容自动同步到公开画廊
    hashed_password = Column(String)
    role = Column(String, default="user") # 'admin', 'user'
    created_at = Column(DateTime, default=get_china_now)
    # 等级与经验值系统
    experience = Column(Integer, default=0)           # 经验值
    level = Column(Integer, default=1)                # 等级 1-5
    exp_updated_at = Column(DateTime, nullable=True)  # 最后经验值更新时间

class ImageGenerationLog(Base):
    __tablename__ = "image_logs"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer)
    count = Column(Integer, default=1)
    created_at = Column(DateTime, default=get_china_now)

class SystemConfig(Base):
    __tablename__ = "system_config"
    key = Column(String, primary_key=True, index=True)
    value = Column(String)

class VideoQueueItem(Base):
    __tablename__ = "video_queue"
    id = Column(String, primary_key=True, index=True)
    filename = Column(String)
    file_path = Column(String)
    prompt = Column(String)
    status = Column(String, default="pending") # pending, processing, done, error
    result_url = Column(String, nullable=True)
    error_msg = Column(String, nullable=True)
    user_id = Column(Integer, nullable=True) # Linked to User
    category = Column(String, nullable=True, default="other")  # Product category
    is_merged = Column(Boolean, nullable=True, default=False)  # Flag for merged/composite videos
    is_shared = Column(Boolean, nullable=False, default=True)  # 默认分享到公开画廊，用户可取消
    created_at = Column(DateTime, default=get_china_now)  # Use China timezone
    _preview_url = Column("preview_url", String, nullable=True)  # Custom preview URL
    retry_count = Column(Integer, default=0)  # 累计重试次数
    last_retry_at = Column(DateTime, nullable=True)  # 上次重试/处理时间
    # Video Quality Review Fields
    review_score = Column(Integer, nullable=True)      # AI感评分 1-10 (越低越自然)
    review_result = Column(Text, nullable=True)        # 详细审查结果 JSON
    review_status = Column(String, nullable=True)      # pending/done/error
    reviewed_at = Column(DateTime, nullable=True)      # 审查时间

    @property
    def preview_url(self):
        # Return custom preview_url if set, otherwise derive from file_path
        if self._preview_url:
            return self._preview_url
        if self.file_path:
            # /app/uploads/queue/xxx -> /uploads/queue/xxx
            if self.file_path.startswith("/app/uploads"):
                return self.file_path.replace("/app/uploads", "/uploads")
        return None
    
    @preview_url.setter
    def preview_url(self, value):
        self._preview_url = value

# --- NEW: Gallery Model ---
class SavedImage(Base):
    __tablename__ = "saved_images"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True)
    filename = Column(String)
    file_path = Column(String) # Local path /app/uploads/gallery/...
    url = Column(String) # Web URL /uploads/gallery/...
    prompt = Column(String)
    width = Column(Integer, nullable=True)  # Image width in pixels
    height = Column(Integer, nullable=True)  # Image height in pixels
    category = Column(String, nullable=True, default="other")  # Product category
    is_shared = Column(Boolean, nullable=False, default=True)  # 默认分享到公开画廊，用户可取消
    created_at = Column(DateTime, default=get_china_now)

# Create tables
try:
    Base.metadata.create_all(bind=engine)
except Exception as e:
    logger.error(f"Database connection failed: {e}")

# ... (dependency omitted) ...
# ... (dependency omitted) ...
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- Data Models ---
class ImageResult(BaseModel):
    angle_name: str
    image_base64: Optional[str] = None
    image_url: Optional[str] = None
    video_prompt: Optional[str] = None
    error: Optional[str] = None

class ConfigItem(BaseModel):
    api_url: str
    api_key: str
    model_name: str
    video_api_url: Optional[str] = None
    video_api_key: Optional[str] = None
    video_model_name: Optional[str] = None
    app_url: Optional[str] = None
    analysis_model_name: Optional[str] = None
    site_title: Optional[str] = None
    site_subtitle: Optional[str] = None
    cache_retention_days: Optional[int] = 7  # 0 = permanent, otherwise days to keep
    # Concurrency settings
    max_concurrent_image: Optional[int] = 5    # 图片生成全局并发数
    max_concurrent_video: Optional[int] = 3    # 视频生成全局并发数
    max_concurrent_story: Optional[int] = 2    # Story Chain 全局并发数
    max_concurrent_per_user: Optional[int] = 2 # 每用户并发上限
    # Thai Dubbing - Gemini Flash (Video Analysis)
    gemini_flash_api_url: Optional[str] = None
    gemini_flash_api_key: Optional[str] = None
    gemini_flash_model_name: Optional[str] = "gemini-2.0-flash"
    # Thai Dubbing - Gemini TTS (Speech Synthesis)
    gemini_tts_api_url: Optional[str] = None
    gemini_tts_api_key: Optional[str] = None
    gemini_tts_model_name: Optional[str] = "gemini-2.5-flash-preview-tts"
    # Video Quality Review (OAI Compatible API)
    review_api_url: Optional[str] = None           # 审查服务地址
    review_api_key: Optional[str] = None           # API 密钥
    review_model_name: Optional[str] = "gpt-4o"    # 模型名称
    review_enabled: Optional[bool] = False         # 是否启用自动审查
    # Feishu/Lark Bitable Integration
    feishu_app_id: Optional[str] = None            # 飞书应用 App ID
    feishu_app_secret: Optional[str] = None        # 飞书应用 App Secret
    feishu_app_token: Optional[str] = None         # 核心词提取 - 多维表格 App Token
    feishu_table_id: Optional[str] = None          # 核心词提取 - 数据表 Table ID
    feishu_description_app_token: Optional[str] = None  # 产品描述 - 多维表格 App Token
    feishu_description_table_id: Optional[str] = None   # 产品描述 - 数据表 Table ID
    # Content Review (内容审核)
    content_review_enabled: Optional[bool] = False           # 是否启用内容审核
    content_review_api_url: Optional[str] = None
    content_review_api_key: Optional[str] = None
    content_review_model: Optional[str] = None
    thai_dubbing_url: Optional[str] = None
    voice_clone_api_url: Optional[str] = None
    voice_clone_api_key: Optional[str] = None
    voice_clone_analysis_model: Optional[str] = None
    voice_clone_tts_model: Optional[str] = None



# --- Background Task for Video Generation ---
async def process_video_background(item_id: str, video_api_url: str, video_api_key: str, video_model_name: str):
    logger.info(f"Background Task: Starting Video Generation for {item_id}")
    
    # Create a fresh DB session for the background task
    db = SessionLocal()
# --- Prompt Matrix ---
ANGLES_PROMPTS = {
    "01_Front_View": "Generate a photorealistic front-view product shot. The product should be centered, facing the camera directly. Lighting should highlight the main features.",
    "02_Side_View_45": "Generate a 3/4 angle product shot (45-degree turn). Show the depth and side profile of the product clearly.",
    "03_Top_Down_View": "Generate a top-down flat lay view of the product. The camera is looking straight down.",
    "04_Low_Angle_Hero": "Generate a low-angle 'hero' shot, looking up at the product to make it look imposing and premium.",
    "05_Close_Up_Detail": "Generate a macro close-up shot focusing on the texture and material quality of the product. Shallow depth of field.",
    "06_Lifestyle_Usage": "Generate a lifestyle scene showing the product being used or placed naturally in the environment.",
    "07_Back_View": "Generate a rear view of the product to show the back design details.",
    "08_Floating_Composition": "Generate a creative shot where the product is slightly floating or tilted dynamically to add energy.",
    "09_Atmospheric_Wide": "Generate a wider shot showing the product fully integrated into the background environment with dramatic lighting."
}

# --- Auth Logic ---
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/v1/login")

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise credentials_exception
    return user

async def get_current_admin(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    return current_user

# --- User Activity Model ---
class UserActivity(Base):
    __tablename__ = "user_activities"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True)
    action = Column(String)  # "image_gen_start", "video_gen_start", "login", etc.
    details = Column(Text, nullable=True)
    created_at = Column(DateTime, default=get_china_now)

# --- Experience Log Model (经验值变更记录) ---
class ExperienceLog(Base):
    __tablename__ = "experience_logs"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True)
    video_id = Column(String, nullable=True)      # 关联的视频ID
    score = Column(Integer)                        # 审查评分
    exp_change = Column(Integer)                   # 经验值变化（可正可负）
    exp_before = Column(Integer)                   # 变化前经验值
    exp_after = Column(Integer)                    # 变化后经验值
    level_before = Column(Integer)                 # 变化前等级
    level_after = Column(Integer)                  # 变化后等级
    created_at = Column(DateTime, default=get_china_now)

# --- Level & Experience Constants and Helpers ---
# 修仙等级体系 (每个境界分前中后期，经验值梯度递增)
# 共 9 大境界 × 3 阶段 = 27 个小等级
# 梯度设计：前期境界较易，后期境界大幅提升难度
LEVEL_THRESHOLDS = [
    # 凡人境 (总需 500)
    (1, 0, 99, "凡人前期"),            # 需 100 经验
    (2, 100, 249, "凡人中期"),         # 需 150 经验
    (3, 250, 499, "凡人后期"),         # 需 250 经验
    
    # 练气境 (总需 1500)
    (4, 500, 899, "练气前期"),         # 需 400 经验
    (5, 900, 1399, "练气中期"),        # 需 500 经验
    (6, 1400, 1999, "练气后期"),       # 需 600 经验
    
    # 筑基境 (总需 3000)
    (7, 2000, 2699, "筑基前期"),       # 需 700 经验
    (8, 2700, 3599, "筑基中期"),       # 需 900 经验
    (9, 3600, 4999, "筑基后期"),       # 需 1400 经验
    
    # 结丹境 (总需 5000)
    (10, 5000, 6499, "结丹前期"),      # 需 1500 经验
    (11, 6500, 8199, "结丹中期"),      # 需 1700 经验
    (12, 8200, 9999, "结丹后期"),      # 需 1800 经验
    
    # 元婴境 (总需 8000)
    (13, 10000, 12499, "元婴前期"),    # 需 2500 经验
    (14, 12500, 15199, "元婴中期"),    # 需 2700 经验
    (15, 15200, 17999, "元婴后期"),    # 需 2800 经验
    
    # 化神境 (总需 12000)
    (16, 18000, 21999, "化神前期"),    # 需 4000 经验
    (17, 22000, 25999, "化神中期"),    # 需 4000 经验
    (18, 26000, 29999, "化神后期"),    # 需 4000 经验
    
    # 炼虚境 (总需 20000)
    (19, 30000, 36999, "炼虚前期"),    # 需 7000 经验
    (20, 37000, 43499, "炼虚中期"),    # 需 6500 经验
    (21, 43500, 49999, "炼虚后期"),    # 需 6500 经验
    
    # 合体境 (总需 30000)
    (22, 50000, 59999, "合体前期"),    # 需 10000 经验
    (23, 60000, 69999, "合体中期"),    # 需 10000 经验
    (24, 70000, 79999, "合体后期"),    # 需 10000 经验
    
    # 大乘境 (最高境界)
    (25, 80000, 94999, "大乘前期"),    # 需 15000 经验
    (26, 95000, 109999, "大乘中期"),   # 需 15000 经验
    (27, 110000, float('inf'), "大乘后期")  # 圆满境界
]

def calculate_level(experience: int) -> tuple:
    """根据经验值计算等级，返回 (level, name)"""
    for level, min_exp, max_exp, name in LEVEL_THRESHOLDS:
        if min_exp <= experience <= max_exp:
            return level, name
    return 27, "大乘后期"

def calculate_exp_change(review_score: int) -> int:
    """根据审查评分计算经验值变化"""
    if review_score >= 8:
        return 20  # 优秀
    elif review_score >= 5:
        return 10  # 一般
    else:
        return -5  # 较差，扣分

def get_level_progress(experience: int) -> float:
    """计算当前等级进度百分比"""
    level, _ = calculate_level(experience)
    for lvl, min_exp, max_exp, name in LEVEL_THRESHOLDS:
        if lvl == level:
            if max_exp == float('inf'):
                return 100.0
            return ((experience - min_exp) / (max_exp - min_exp + 1)) * 100
    return 0.0

# Startup: Create Admin and Init Services
@app.on_event("startup")
async def startup_event():
    # Initialize database tables
    try:
        Base.metadata.create_all(bind=engine)
    except Exception as e:
        logger.warning(f"Table creation warning: {e}")
    
    # Initialize WebSocket manager with Redis (will resolve hostname automatically)
    try:
        await init_websocket_manager()
        logger.info("WebSocket manager initialized")
    except Exception as e:
        logger.warning(f"WebSocket manager init failed (Redis may not be ready): {e}")
    
    # Create admin user
    db = SessionLocal()
    try:

        admin_user = os.getenv("ADMIN_USER", "admin")
        admin_pass = os.getenv("ADMIN_PASSWORD", "change_this_password")
        
        # Check if admin user exists
        user = db.query(User).filter(User.username == admin_user).first()
        if not user:
            print(f"Creating default admin user: {admin_user}...")
            admin = User(
                username=admin_user,
                hashed_password=get_password_hash(admin_pass),
                role="admin"
            )
            db.add(admin)
            db.commit()
        else:
            # Always update password to match ENV (so user can reset via ENV)
            print(f"Updating admin user password for: {admin_user}")
            user.hashed_password = get_password_hash(admin_pass)
            # Ensure role is admin
            user.role = "admin"
            db.commit()
        
        # --- 启动时恢复阻滞的视频任务 ---
        MAX_AUTO_RETRIES = 3
        stale_tasks = db.query(VideoQueueItem).filter(
            VideoQueueItem.status == "processing"
        ).all()
        
        if stale_tasks:
            recovered_count = 0
            failed_count = 0
            for task in stale_tasks:
                if task.retry_count >= MAX_AUTO_RETRIES:
                    task.status = "error"
                    task.error_msg = f"启动恢复: 已达最大重试次数 ({task.retry_count}/{MAX_AUTO_RETRIES})"
                    failed_count += 1
                else:
                    task.status = "pending"
                    # 不增加 retry_count，因为任务被中断而非失败
                    # 不设置 last_retry_at，避免触发 cooldown
                    recovered_count += 1
            db.commit()
            logger.info(f"[Startup] 恢复 {recovered_count} 个阻滞任务, {failed_count} 个标记为错误")
        else:
            logger.info("[Startup] 无阻滞任务需要恢复")
    finally:
        db.close()

@app.on_event("shutdown")
async def shutdown_event():
    await shutdown_websocket_manager()
    logger.info("Shutdown complete")

# --- Public APIs (No Auth Required) ---

@app.get("/api/v1/public/videos")
def get_public_videos(
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Get shared videos for public display (no auth required)."""
    query = db.query(VideoQueueItem).filter(
        VideoQueueItem.status.in_(["done", "archived"]),
        VideoQueueItem.is_shared == True
    )
    
    total = query.count()
    videos = query.order_by(VideoQueueItem.created_at.desc()).limit(limit).offset(offset).all()
    
    result_items = []
    for vid in videos:
        creator = db.query(User).filter(User.id == vid.user_id).first()
        result_items.append({
            "id": vid.id,
            "prompt": vid.prompt,
            "result_url": vid.result_url,
            "preview_url": vid.preview_url,
            "username": (creator.nickname or creator.username) if creator else "Creator",
            "avatar": creator.avatar if creator else None,
            "category": vid.category or "other",
            "is_merged": vid.is_merged or False,
            "created_at": vid.created_at
        })
    
    return {"total": total, "items": result_items}

@app.get("/api/v1/public/config")
def get_public_config(db: Session = Depends(get_db)):
    """Get public site configuration (no auth required)."""
    site_title = db.query(SystemConfig).filter(SystemConfig.key == "site_title").first()
    site_subtitle = db.query(SystemConfig).filter(SystemConfig.key == "site_subtitle").first()
    return {
        "site_title": site_title.value if site_title else os.getenv("SITE_TITLE", "BNP Studio"),
        "site_subtitle": site_subtitle.value if site_subtitle else os.getenv("SITE_SUBTITLE", "AI Video Gallery")
    }

# Login Endpoint
class Token(BaseModel):
    access_token: str
    token_type: str
    username: str
    role: str
    user_id: int

# Cloudflare Turnstile Secret Key (from .env)
TURNSTILE_SECRET_KEY = os.getenv("TURNSTILE_SECRET_KEY", "")

async def verify_turnstile(token: str, ip: str = None) -> bool:
    """Verify Cloudflare Turnstile token."""
    if not token:
        return False
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://challenges.cloudflare.com/turnstile/v0/siteverify",
                data={
                    "secret": TURNSTILE_SECRET_KEY,
                    "response": token,
                    "remoteip": ip
                },
                timeout=10.0
            )
            result = response.json()
            return result.get("success", False)
    except Exception as e:
        logger.error(f"Turnstile verification failed: {e}")
        return False

@app.post("/api/v1/login", response_model=Token)
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    turnstile_token: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    # Verify Turnstile if token provided and secret key is configured
    if turnstile_token and TURNSTILE_SECRET_KEY:
        client_ip = request.client.host if request.client else None
        is_valid = await verify_turnstile(turnstile_token, client_ip)
        if not is_valid:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="人机验证失败，请重试"
            )
    
    user = db.query(User).filter(User.username == username).first()
    if not user or not verify_password(password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="用户名或密码错误",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role, "uid": user.id},
        expires_delta=access_token_expires
    )
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "username": user.username,
        "role": user.role,
        "user_id": user.id
    }

# verify_token: Legacy function for endpoints that only need authentication without user object
# For endpoints that need user data, use get_current_user instead

def verify_token(token: str = Depends(oauth2_scheme)):
    try:
        jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return token
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid token",
             headers={"WWW-Authenticate": "Bearer"},
        )

# --- User Profile APIs ---

class UserProfileResponse(BaseModel):
    id: int
    username: str
    nickname: Optional[str] = None
    avatar: Optional[str] = None
    role: str
    created_at: datetime

class UpdateProfileRequest(BaseModel):
    nickname: Optional[str] = None
    password: Optional[str] = None
    default_share: Optional[bool] = None

@app.get("/api/v1/user/profile")
def get_user_profile(user: User = Depends(get_current_user)):
    """Get current user's profile with level and experience info."""
    level, level_name = calculate_level(user.experience or 0)
    level_progress = get_level_progress(user.experience or 0)
    
    return {
        "id": user.id,
        "username": user.username,
        "nickname": user.nickname or user.username,
        "avatar": user.avatar,
        "role": user.role,
        "default_share": user.default_share or False,
        "created_at": user.created_at,
        # 等级与经验值信息
        "experience": user.experience or 0,
        "level": level,
        "level_name": level_name,
        "level_progress": round(level_progress, 1)
    }

@app.put("/api/v1/user/profile")
def update_user_profile(
    request: UpdateProfileRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Update current user's profile (nickname, password, default_share)."""
    if request.nickname is not None:
        user.nickname = request.nickname
    if request.password is not None and request.password.strip():
        user.hashed_password = get_password_hash(request.password)
    if request.default_share is not None:
        user.default_share = request.default_share
    db.commit()
    return {"message": "Profile updated successfully", "nickname": user.nickname, "default_share": user.default_share}

@app.post("/api/v1/user/avatar")
async def upload_avatar(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Upload user avatar image."""
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, GIF, WebP allowed.")
    
    # Create avatars directory if not exists
    avatar_dir = "/app/uploads/avatars"
    os.makedirs(avatar_dir, exist_ok=True)
    
    # Generate unique filename
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    filename = f"avatar_{user.id}_{int(datetime.now().timestamp())}.{ext}"
    file_path = f"{avatar_dir}/{filename}"
    
    # Save file
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Update user avatar URL
    avatar_url = f"/uploads/avatars/{filename}"
    user.avatar = avatar_url
    db.commit()
    
    return {"message": "Avatar uploaded successfully", "avatar": avatar_url}

@app.get("/api/v1/user/experience/history")
def get_experience_history(
    limit: int = 20,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取用户经验值变更历史记录"""
    logs = db.query(ExperienceLog).filter(
        ExperienceLog.user_id == user.id
    ).order_by(ExperienceLog.created_at.desc()).limit(limit).all()
    
    return [{
        "video_id": log.video_id,
        "score": log.score,
        "exp_change": log.exp_change,
        "exp_before": log.exp_before,
        "exp_after": log.exp_after,
        "level_before": log.level_before,
        "level_after": log.level_after,
        "created_at": log.created_at
    } for log in logs]

# --- User Manage Models ---
class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "user"

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None

class UserOut(BaseModel):
    id: int
    username: str
    role: str
    created_at: datetime
    experience: Optional[int] = 0
    level: Optional[int] = 1
    level_name: Optional[str] = None
    class Config:
        from_attributes = True

class StatsResponse(BaseModel):
    by_day: List[dict]
    by_week: List[dict]
    by_month: List[dict]
    user_stats: List[dict]

# --- Admin Endpoints ---

@app.get("/api/v1/users")
def get_users(db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    import random
    # 负分时显示的随机动物列表
    NEGATIVE_ANIMALS = ["🐸 蛤蟆", "🐛 毛虫", "🪱 蚯蚓", "🐌 蜗牛", "🦎 蜥蜴", "🐁 老鼠", "🪳 蟑螂", "🦠 变形虫"]
    
    users = db.query(User).all()
    result = []
    for u in users:
        exp = u.experience or 0
        if exp < 0:
            # 负分用随机动物
            level_name = random.choice(NEGATIVE_ANIMALS)
            level = 0
        else:
            level, level_name = calculate_level(exp)
        
        result.append({
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "created_at": u.created_at,
            "experience": exp,
            "level": level,
            "level_name": level_name
        })
    return result

@app.post("/api/v1/users", response_model=UserOut)
def create_user(user: UserCreate, db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    db_user = db.query(User).filter(User.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    hashed_password = get_password_hash(user.password)
    new_user = User(username=user.username, hashed_password=hashed_password, role=user.role)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.put("/api/v1/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, user: UserUpdate, db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.username:
        db_user.username = user.username
    if user.password:
        db_user.hashed_password = get_password_hash(user.password)
    # Only allow role change if not self or careful logic? Allow for now.
    if user.role:
        db_user.role = user.role
        
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/api/v1/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    """Delete a user (admin only). Cannot delete self."""
    if admin.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if db_user.role == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete admin users")
    
    # Delete user's data (optional: could keep data or cascade)
    db.delete(db_user)
    db.commit()
    return {"message": "User deleted successfully"}

from sqlalchemy import func

@app.get("/api/v1/stats")
def get_stats(db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    # 1. User Summary Stats (Total Counts)
    users = db.query(User).all()
    user_stats = []
    
    # Get today's start in China timezone (00:00), then convert to UTC for comparison
    # This handles both old UTC data and new China timezone data
    now = get_china_now()
    today_start_china = now.replace(hour=0, minute=0, second=0, microsecond=0)
    # Convert China time 00:00 to UTC (subtract 8 hours): yesterday 16:00 UTC
    today_start_utc = today_start_china - timedelta(hours=8)
    
    for u in users:
        # Count actual saved images in gallery (not just generation attempts) - Total
        img_count = db.query(SavedImage).filter(SavedImage.user_id == u.id).count()
        # Count only completed videos - Total
        vid_count = db.query(VideoQueueItem).filter(
            VideoQueueItem.user_id == u.id,
            VideoQueueItem.status.in_(["done", "archived"])
        ).count()
        
        # Today's counts - use UTC start time for comparison
        # This counts items where created_at >= yesterday 16:00 UTC (= today 00:00 Beijing)
        today_img = db.query(SavedImage).filter(
            SavedImage.user_id == u.id,
            SavedImage.created_at >= today_start_utc
        ).count()
        today_vid = db.query(VideoQueueItem).filter(
            VideoQueueItem.user_id == u.id,
            VideoQueueItem.status.in_(["done", "archived"]),
            VideoQueueItem.created_at >= today_start_utc
        ).count()
        
        user_stats.append({
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "image_count": img_count,
            "video_count": vid_count,
            "today_images": today_img,
            "today_videos": today_vid
        })

    # 2. Daily Activity (Last 30 Days)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    
    # SQLite uses strftime or date function. func.date works in most.
    # Group by Date only (All users combined) for the main chart
    img_daily = db.query(
        func.date(ImageGenerationLog.created_at).label('date'),
        func.count(ImageGenerationLog.id).label('count')
    ).filter(ImageGenerationLog.created_at >= thirty_days_ago).group_by('date').all()
    
    vid_daily = db.query(
        func.date(VideoQueueItem.created_at).label('date'),
        func.count(VideoQueueItem.id).label('count')
    ).filter(VideoQueueItem.created_at >= thirty_days_ago).group_by('date').all()

    # Format for JSON
    by_day_img = [{"date": str(r.date), "count": r.count, "type": "image"} for r in img_daily]
    by_day_vid = [{"date": str(r.date), "count": r.count, "type": "video"} for r in vid_daily]

    return {
        "user_stats": user_stats,
        "by_day": by_day_img + by_day_vid,
        "by_week": [], # Placeholder
        "by_month": [] # Placeholder
    }
    
    # For Detailed Charts (Day/Week/Month), maybe return raw list of recent activity?
    # Or implement date_trunc logic.
    # Let's implement dynamic date range stats if requested.
    # User asked: "Statistics of each user and all users... (Day, Week, Month)"
    # I'll add logic to fetch last 30 days grouped by day.


@app.get("/api/v1/config", response_model=ConfigItem)
def get_config(db: Session = Depends(get_db), token: str = Depends(verify_token)):
    # Defaults
    defaults = {
        "api_url": os.getenv("DEFAULT_API_URL", "https://generativelanguage.googleapis.com"),
        "api_key": os.getenv("DEFAULT_API_KEY", ""),
        "model_name": os.getenv("DEFAULT_MODEL_NAME", "gemini-3-pro-image-preview"),
        "video_api_url": os.getenv("VIDEO_API_URL", ""),
        "video_api_key": os.getenv("VIDEO_API_KEY", ""),
        "video_model_name": os.getenv("VIDEO_MODEL_NAME", "sora-video-portrait"),
        "app_url": os.getenv("APP_URL", "http://localhost:33012"),
        "analysis_model_name": os.getenv("DEFAULT_ANALYSIS_MODEL_NAME", "gemini-3-pro-preview"),
        "site_title": os.getenv("SITE_TITLE", "Banana Product"),
        "site_subtitle": os.getenv("SITE_SUBTITLE", ""),
        "cache_retention_days": int(os.getenv("CACHE_RETENTION_DAYS", "7")),  # 0 = permanent
        # Concurrency settings
        "max_concurrent_image": int(os.getenv("MAX_CONCURRENT_IMAGE", "5")),
        "max_concurrent_video": int(os.getenv("MAX_CONCURRENT_VIDEO", "3")),
        "max_concurrent_story": int(os.getenv("MAX_CONCURRENT_STORY", "2")),
        "max_concurrent_per_user": int(os.getenv("MAX_CONCURRENT_PER_USER", "2")),
        # Thai Dubbing - Gemini Flash (Video Analysis)
        "gemini_flash_api_url": os.getenv("GEMINI_FLASH_API_URL", "https://generativelanguage.googleapis.com"),
        "gemini_flash_api_key": os.getenv("GEMINI_FLASH_API_KEY", ""),
        "gemini_flash_model_name": os.getenv("GEMINI_FLASH_MODEL_NAME", "gemini-2.0-flash"),
        # Thai Dubbing - Gemini TTS (Speech Synthesis)
        "gemini_tts_api_url": os.getenv("GEMINI_TTS_API_URL", "https://generativelanguage.googleapis.com"),
        "gemini_tts_api_key": os.getenv("GEMINI_TTS_API_KEY", ""),
        "gemini_tts_model_name": os.getenv("GEMINI_TTS_MODEL_NAME", "gemini-2.5-flash-preview-tts"),
        # Video Quality Review (OAI Compatible API)
        "review_api_url": os.getenv("REVIEW_API_URL", ""),
        "review_api_key": os.getenv("REVIEW_API_KEY", ""),
        "review_model_name": os.getenv("REVIEW_MODEL_NAME", "gpt-4o"),
        "review_enabled": os.getenv("REVIEW_ENABLED", "false").lower() == "true",
        "feishu_app_id": os.getenv("FEISHU_APP_ID", ""),
        "feishu_app_secret": os.getenv("FEISHU_APP_SECRET", ""),
        "feishu_app_token": os.getenv("FEISHU_APP_TOKEN", ""),
        "feishu_table_id": os.getenv("FEISHU_TABLE_ID", ""),
        "feishu_description_app_token": os.getenv("FEISHU_DESCRIPTION_APP_TOKEN", ""),
        "feishu_description_table_id": os.getenv("FEISHU_DESCRIPTION_TABLE_ID", ""),
        "content_review_enabled": os.getenv("CONTENT_REVIEW_ENABLED", "false").lower() == "true",
        "content_review_api_url": os.getenv("CONTENT_REVIEW_API_URL", ""),
        "content_review_api_key": os.getenv("CONTENT_REVIEW_API_KEY", ""),
        "content_review_model": os.getenv("CONTENT_REVIEW_MODEL", ""),
        "voice_clone_api_url": os.getenv("VOICE_CLONE_API_URL", ""),
        "voice_clone_api_key": os.getenv("VOICE_CLONE_API_KEY", ""),
        "voice_clone_analysis_model": os.getenv("VOICE_CLONE_ANALYSIS_MODEL", ""),
        "voice_clone_tts_model": os.getenv("VOICE_CLONE_TTS_MODEL", ""),
    }
    
    # Check if DB is empty for these keys, if so, seed them
    for key, val in defaults.items():
        item = db.query(SystemConfig).filter(SystemConfig.key == key).first()
        if not item:
            # Seed default
            new_item = SystemConfig(key=key, value=val)
            db.add(new_item)
            db.commit()
            db.refresh(new_item)
        else:
            defaults[key] = item.value
            
    return ConfigItem(**defaults)

@app.post("/api/v1/config", response_model=ConfigItem)
def update_config(config: ConfigItem, db: Session = Depends(get_db), token: str = Depends(verify_token)):
    for key, val in config.dict().items():
        if val is not None:
            str_val = str(val).lower() if isinstance(val, bool) else str(val)
            item = db.query(SystemConfig).filter(SystemConfig.key == key).first()
            if item:
                item.value = str_val
            else:
                item = SystemConfig(key=key, value=str_val)
                db.add(item)
    
    db.commit()
    return config


# --- Models Proxy Endpoint ---
class ModelsRequest(BaseModel):
    api_url: str
    api_key: str

@app.post("/api/v1/models")
async def get_models(request: ModelsRequest, token: str = Depends(verify_token)):
    """
    Proxy endpoint to fetch models list from an OpenAI-compatible API.
    This avoids CORS issues when frontend tries to call external APIs directly.
    """
    import httpx
    
    api_url = request.api_url.rstrip('/')
    if not api_url.endswith('/models'):
        api_url = api_url.replace('/chat/completions', '').rstrip('/')
        if api_url.endswith('/v1'):
            api_url = api_url + '/models'
        else:
            api_url = api_url + '/v1/models'
    
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(
                api_url,
                headers={'Authorization': f'Bearer {request.api_key}'}
            )
            
            if response.status_code == 200:
                data = response.json()
                models = data.get('data') or data.get('models') or []
                # Extract model IDs
                model_ids = []
                for m in models:
                    if isinstance(m, dict):
                        model_ids.append(m.get('id') or m.get('name') or str(m))
                    else:
                        model_ids.append(str(m))
                return {"models": model_ids}
            else:
                return {"models": [], "error": f"API returned {response.status_code}"}
    except Exception as e:
        return {"models": [], "error": str(e)}


# --- Concurrency Config Getter for ConcurrencyLimiter ---
async def get_concurrency_config() -> dict:
    """
    Get concurrency-related config from database.
    Used by ConcurrencyLimiter for dynamic limits.
    """
    db = SessionLocal()
    try:
        config = {}
        keys = ["max_concurrent_image", "max_concurrent_video", "max_concurrent_story", "max_concurrent_per_user"]
        for key in keys:
            item = db.query(SystemConfig).filter(SystemConfig.key == key).first()
            if item:
                config[key] = item.value
        return config
    finally:
        db.close()

# ... (lines 170-432 omitted) ...

# --- Helper: File to Base64 ---
async def file_to_base64(file: UploadFile) -> str:
    content = await file.read()
    return base64.b64encode(content).decode('utf-8')

async def file_to_base64_compressed(file: UploadFile, max_size: int = 800, quality: int = 75) -> str:
    """Convert uploaded file to base64 with compression to reduce payload size."""
    from PIL import Image
    from io import BytesIO
    
    content = await file.read()
    
    try:
        img = Image.open(BytesIO(content))
        
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        width, height = img.size
        if width > max_size or height > max_size:
            ratio = min(max_size / width, max_size / height)
            new_size = (int(width * ratio), int(height * ratio))
            img = img.resize(new_size, Image.LANCZOS)
            logger.info(f"Compressed image from {width}x{height} to {new_size[0]}x{new_size[1]}")
        
        buffer = BytesIO()
        img.save(buffer, format='JPEG', quality=quality, optimize=True)
        compressed_content = buffer.getvalue()
        
        logger.info(f"Image size: {len(content)} -> {len(compressed_content)} bytes ({len(compressed_content)*100//len(content)}%)")
        
        return base64.b64encode(compressed_content).decode('utf-8')
    except Exception as e:
        logger.warning(f"Image compression failed, using original: {e}")
        return base64.b64encode(content).decode('utf-8')

# --- Analysis Models ---
class ScriptItem(BaseModel):
    angle_name: str
    script: str

class AnalyzeResponse(BaseModel):
    product_description: str
    environment_analysis: str
    placement_mode: str
    scripts: List[ScriptItem]

# --- Helper: Image Generation ---
async def call_openai_compatible_api(
    client: httpx.AsyncClient, 
    api_url: str, 
    api_key: str, 
    product_b64: str, 
    bg_b64: str, 
    angle_name: str, 
    angle_prompt: str,
    model: str = "gemini-3-pro-image-preview" 
) -> ImageResult:
    print(f"DEBUG: Entering call_openai_compatible_api for {angle_name}. Model: {model}", flush=True)
    logger.info(f"Image Generation Prompt for {angle_name}: {angle_prompt[:200]}...")  # Log first 200 chars
    
    system_instruction = (
        "You are a professional product photographer and image compositing expert. "
        "=== ABSOLUTE PRODUCT FIDELITY RULES (HIGHEST PRIORITY) ===\n"
        "🔴 The product in the FIRST image is the HERO SUBJECT. Your task is PHOTO COMPOSITING, not reimagining.\n"
        "🔴 PRESERVE 100%: Exact shape, silhouette, proportions, colors, textures, materials, logos, labels, text, reflections, and every detail.\n"
        "🔴 FORBIDDEN: Restyling, recoloring, adding/removing features, changing angles, morphing shapes, or ANY visual alteration to the product.\n"
        "🔴 Think of this as: 'Cut the product from the first image and paste it into a new scene' - NOT 'Draw the product again in a new style'.\n"
        "\n=== SCENE GENERATION RULES (STRICT REFERENCE REPLICATION) ===\n"
        "✅ DO: Create new backgrounds, lighting, shadows, reflections consistent with the new environment.\n"
        "✅ DO: Match the lighting direction on the product to the new scene's light sources.\n"
        "✅ DO: STRICTLY REPLICATE the scene type and environment from the SECOND (reference) image.\n"
        "✅ DO: If the reference image contains human figures, include similar poses, gestures, and clothing styles.\n"
        "✅ DO: Copy the overall composition, props, and atmospheric elements from the reference.\n"
        "✅ DO: Apply the reference image's color grading, mood, and lighting atmosphere.\n"
        "✅ Return ONLY the final composited image."
    )
    
    # Structured user prompt for clarity
    user_prompt_text = (
        f"=== PRODUCT (Image 1) ===\n"
        f"This is the EXACT product to composite. DO NOT modify it.\n\n"
        f"=== SCENE REFERENCE (Image 2) - STRICT REPLICATION REQUIRED ===\n"
        f"REPLICATE this scene's environment, props, composition, and atmosphere.\n"
        f"If human figures are present, include similar poses and styling.\n"
        f"Apply this image's color grading and lighting mood.\n\n"
        f"=== SCENE INSTRUCTION ===\n"
        f"{angle_prompt}"
    )
    
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": system_instruction
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text", 
                        "text": user_prompt_text
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{product_b64}"
                        }
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{bg_b64}"
                        }
                    }
                ]
            }
        ],
        "temperature": 0.2,
        "max_tokens": 4096
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    # Enhanced retry logic with Jitter for Cloudflare rate limit prevention
    max_retries = 4  # Increased from 3
    base_delay = 3.0  # Increased from 2.0
    
    for attempt in range(max_retries + 1):
        try:
            # Apply global request throttling (anti-CF)
            await throttle_request()
            
            target_url = api_url
            if not target_url.endswith("/chat/completions") and not target_url.endswith(":generateContent"):
                 target_url = f"{target_url.rstrip('/')}/chat/completions"

            logger.info(f"Sending request for {angle_name} to {target_url} (Attempt {attempt+1}/{max_retries+1})")
            
            # Enable streaming
            payload["stream"] = True
            
            # Staged timeout: quick connect, longer read
            timeout = httpx.Timeout(connect=15.0, read=300.0, write=30.0, pool=30.0)
            
            async with client.stream("POST", target_url, json=payload, headers=headers, timeout=timeout) as response:
                if response.status_code == 429:
                    if attempt < max_retries:
                        jitter = random.uniform(0.8, 1.5)
                        wait_time = base_delay * (2 ** attempt) * jitter
                        logger.warning(f"Rate limited (429). Retrying in {wait_time:.1f}s with jitter...")
                        await asyncio.sleep(wait_time)
                        continue
                    else:
                        error_text = await response.aread()
                        logger.error(f"API Error {response.status_code}: {error_text}")
                        return ImageResult(angle_name=angle_name, error=f"Rate Limit Exceeded (429): {error_text.decode('utf-8')}")

                if response.status_code == 524:
                    logger.warning(f"Cloudflare 524 timeout for {angle_name} (attempt {attempt+1}/{max_retries+1})")
                    if attempt < max_retries:
                        await asyncio.sleep(2)
                        continue
                    return ImageResult(angle_name=angle_name, error="服务器处理超时，请稍后重试")

                if response.status_code != 200:
                    error_text = await response.aread()
                    logger.error(f"API Error {response.status_code}: {error_text}")
                    return ImageResult(angle_name=angle_name, error=f"API Error {response.status_code}: {error_text.decode('utf-8')}")

                full_content = ""
                async for line in response.aiter_lines():
                    if line.strip():
                        logger.info(f"Stream line: {line[:100]}")
                    if line.startswith("data: "):
                        data_str = line[6:]
                        if data_str.strip() == "[DONE]":
                            break
                        try:
                            import json
                            chunk = json.loads(data_str)
                            # Log the first chunk structure to debug
                            if not full_content:
                                logger.info(f"First chunk: {chunk}")
                            
                            choices = chunk.get("choices", [])
                            if choices and len(choices) > 0:
                                delta = choices[0].get("delta", {}).get("content", "")
                                if delta:
                                    full_content += delta
                            else:
                                # Some chunks might be usage info or empty
                                pass
                        except Exception as e:
                            logger.error(f"Chunk parse error: {e}")
                            pass
                
                content = full_content
                logger.info(f"API Response Content for {angle_name}: {content[:200]}...")
                
                # Check for HTML error response (e.g. Cloudflare 504/502)
                content_sample = content.strip().lower()
                # Check if response is HTML (Cloudflare 429 page or other gateway errors)
                if content_sample.startswith("html>") or content_sample.startswith("<!doctype") or "<html" in content_sample:
                    logger.warning(f"Received HTML content (likely Cloudflare rate limit) for {angle_name}: {content[:300]}")
                    
                    # Check if it's a Cloudflare rate limit page
                    if "429" in content or "Too many requests" in content.lower() or "Just a moment" in content:
                        if attempt < max_retries:
                            # Randomized wait for Cloudflare with extra buffer
                            cloudflare_extra = random.uniform(8, 15)
                            jitter = random.uniform(0.8, 1.5)
                            wait_time = base_delay * (2 ** attempt) * jitter + cloudflare_extra
                            logger.warning(f"Detected Cloudflare rate limit. Retrying in {wait_time:.1f}s (attempt {attempt+1}/{max_retries+1})...")
                            await asyncio.sleep(wait_time)
                            continue
                        else:
                            logger.error(f"Rate limit persists after {max_retries+1} attempts for {angle_name}")
                            return ImageResult(angle_name=angle_name, error="Rate Limit Exceeded (Cloudflare Protection)")
                    
                    # Other HTML errors (gateway timeout, etc.)
                    logger.error(f"API Gateway Error for {angle_name}")
                    return ImageResult(angle_name=angle_name, error="API Gateway Timeout (Upstream Error)")

                import re
                # Try to find markdown image
                img_match = re.search(r'!\[.*?\]\((.*?)\)', content)
                if img_match:
                    return ImageResult(angle_name=angle_name, image_base64=img_match.group(1))
                    
                # If no markdown image, return content as is (might be base64 or text description)
                return ImageResult(angle_name=angle_name, image_base64=content)
            
        except Exception as e:
            logger.error(f"Request failed: {e}")
            if attempt < max_retries:
                jitter = random.uniform(0.8, 1.5)
                wait_time = base_delay * (2 ** attempt) * jitter
                logger.warning(f"Request failed with {e}. Retrying in {wait_time:.1f}s...")
                await asyncio.sleep(wait_time)
                continue
            return ImageResult(angle_name=angle_name, error=str(e))

# --- Helper: Analysis Logic ---
async def analyze_product_scene(
    client: httpx.AsyncClient,
    api_url: str,
    api_key: str,
    product_b64: str,
    ref_b64: str,
    category: str,
    model: str = "gemini-3-pro-preview",
    custom_product_name: Optional[str] = None,
    gen_count: int = 9  # User-selected number of scripts
) -> AnalyzeResponse:
    
    # Enhanced system instruction with structured prompt engineering framework
    system_instruction = """Role: You are an elite commercial photographer, art director, and prompt engineer with expertise in:
- Product photography & styling across all categories
- Lighting design (natural, studio, dramatic, ambient, cinematic)
- Composition principles (rule of thirds, golden ratio, leading lines, negative space)
- Visual storytelling through angles, environments, and mood

=== ANALYSIS METHODOLOGY ===
When analyzing product and reference images, apply this framework:

1. PRODUCT IDENTITY EXTRACTION
   - Core visual DNA: shape, silhouette, primary materials
   - Surface properties: texture, reflectivity, transparency, color palette
   - Brand positioning: luxury/mass-market, modern/classic, minimal/expressive

2. STYLE REFERENCE ANALYSIS (CRITICAL - STRICT REPLICATION REQUIRED)
   - Scene Type: Identify the EXACT scene category (indoor/outdoor, room type, location type)
   - Scene Elements: List ALL visible objects, furniture, props, architectural features
   - Human Figures: If present, describe pose, gesture, clothing style, positioning relative to camera
   - Mood transfer elements: color temperature, contrast ratio, saturation
   - Lighting signature: direction, quality (soft/hard), color cast
   - Environmental textures and atmospheric elements
   - NOTE: The generated scripts MUST replicate this scene's environment and human poses

3. PLACEMENT LOGIC ENGINE
   - Tech/Electronics → Minimalist surfaces, floating, sleek environments
   - Beauty/Cosmetics → Soft textures, organic materials, lifestyle contexts
   - Food/Beverage → Natural settings, appetite-appeal lighting
   - Fashion → Editorial styling, dramatic lighting, dynamic poses
   - Home/Furniture → Contextual interior scenes, warm lighting
   - Industrial/Security → Professional environments, technical precision

4. DIVERSITY MATRIX
   Generate scripts that maximize variation across these axes:
   - Lighting: Natural daylight | Golden hour | Studio softbox | Neon/Colored | Dramatic rim light | Ambient glow
   - Environment: Indoor minimal | Outdoor nature | Urban texture | Abstract gradient | Lifestyle context | Studio infinity
   - Composition: Hero center | Dynamic diagonal | Extreme close-up | Environmental wide | Floating/levitation | Pattern/repetition
   - Mood: Clean professional | Warm inviting | Cool modern | Dramatic cinematic | Playful vibrant | Serene meditative

=== SCRIPT QUALITY REQUIREMENTS ===
Each script must be a production-ready photography brief containing:
- Specific lighting setup with direction and quality descriptors
- Detailed environment description with textures and props
- Mood and atmosphere keywords for consistent generation
- Composition guidance with framing and depth of field
- Avoid generic descriptions; be specific and evocative"""

    product_identity_prompt = "1. Identify the product in the first image."
    if custom_product_name:
        product_identity_prompt = f"1. The user has identified this product as: '{custom_product_name}'. Verify this and describe its key visual features (shape, materials, colors, unique characteristics)."

    prompt = (
        f"Product Category: {category}\n\n"
        f"{product_identity_prompt}\n\n"
        "2. STYLE REFERENCE ANALYSIS (🔴 CRITICAL - STRICT SCENE REPLICATION):\\n"
        "   - Scene Type: What EXACT type of scene is this? (e.g., modern kitchen, outdoor cafe, studio backdrop, urban street)\\n"
        "   - Scene Elements: List ALL visible objects, furniture, props, decorations in the reference image\\n"
        "   - Human Figures: If people are present, describe their EXACT pose, gesture, clothing style, and position\\n"
        "   - Lighting: Identify direction, quality, color temperature, and any special lighting effects\\n"
        "   - Color Palette: Dominant colors, mood, and atmospheric elements\\n"
        "   🔴 IMPORTANT: ALL generated scripts MUST recreate this EXACT scene type and include similar human figures if present\\n\\n"
        "3. PLACEMENT MODE DETERMINATION:\n"
        "   Analyze product type + reference environment to determine optimal placement:\n"
        "   - Wall-mounted: cameras, sensors, displays (when reference shows vertical surfaces)\n"
        "   - Tabletop: bottles, cosmetics, electronics, small items\n"
        "   - Floor-standing: large appliances, furniture, equipment\n"
        "   - Floating/Levitation: tech products, premium items for dramatic effect\n"
        "   - Embedded/Contextual: lifestyle products shown in use-case environments\n\n"
        f"4. Generate exactly {gen_count} DISTINCT photography scripts. (CRITICAL: You MUST generate exactly {gen_count} items. Do not generate fewer.)\n\n"
        "   DIVERSITY REQUIREMENTS:\n"
        "   - Each script MUST explore a unique combination from the Diversity Matrix\n"
        "   - NO two scripts should share the same lighting + environment combination\n"
        "   - Include at least: 2 natural light, 2 studio light, 2 dramatic/creative light scenarios\n"
        "   - Vary composition: mix hero shots, close-ups, environmental shots, and creative angles\n\n"
        "   SCRIPT CONTENT (🔴 MUST REPLICATE REFERENCE SCENE):\\n"
        "   - angle_name: A short, descriptive title (e.g., 'Golden Hour Warmth', 'Studio Minimal', 'Urban Texture')\\n"
        "   - script: A detailed 40-80 word prompt that MUST include:\\n"
        "     * The EXACT scene type from the reference image (same environment category)\\n"
        "     * Similar props, furniture, and scene elements as the reference\\n"
        "     * If reference has human figures: include similar pose, gesture, and clothing style\\n"
        "     * Specific lighting, mood, and composition guidance\\n\\n"
        "Return JSON format ONLY:\n"
        "{ \"product_description\": \"...\", \"environment_analysis\": \"...\", \"placement_mode\": \"...\", \"scripts\": [ {\"angle_name\": \"...\", \"script\": \"...\"}, ... ] }"
    )

    payload = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": system_instruction
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{product_b64}"}
                    },
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{ref_b64}"}
                    }
                ]
            }
        ],
        "temperature": 0.4,
        "max_tokens": 4096,
        "response_format": {"type": "json_object"}
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    target_url = api_url
    if not target_url.endswith("/chat/completions") and not target_url.endswith(":generateContent"):
         target_url = f"{target_url.rstrip('/')}/chat/completions"
         
    response = await client.post(target_url, json=payload, headers=headers, timeout=300.0)
    
    if response.status_code != 200:
        raise HTTPException(status_code=response.status_code, detail=f"Analysis API Error: {response.text}")
        
    data = response.json()
    content = data.get("choices", [])[0].get("message", {}).get("content", "")
    
    import json
    try:
        # Clean markdown code blocks if present
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[0].strip()
            
        parsed = json.loads(content)
        
        # --- FIX: Ensure exact number of scripts ---
        scripts = parsed.get("scripts", [])
        if len(scripts) < gen_count:
            logger.warning(f"AI generated {len(scripts)} scripts, but user requested {gen_count}. Padding with variants...")
            
            original_count = len(scripts)
            if original_count > 0:
                import copy
                needed = gen_count - original_count
                for i in range(needed):
                    # Round-robin selection from original scripts
                    source_script = scripts[i % original_count]
                    new_script = copy.deepcopy(source_script)
                    new_script["angle_name"] = f"{source_script['angle_name']} (Var {i+1})"
                    # Optionally slight tweak to script text could be done here, but simple dup is safer than broken JSON
                    scripts.append(new_script)
            else:
                 # Fallback if 0 scripts returned (rare)
                 for i in range(gen_count):
                     scripts.append({
                         "angle_name": f"Auto Generated {i+1}",
                         "script": f"Product shot in professional lighting, angle {i+1}, clean composition, high quality."
                     })
        
        # Truncate if too many (rare but possible)
        if len(scripts) > gen_count:
            scripts = scripts[:gen_count]
            
        parsed["scripts"] = scripts
        # ---------------------------------------------
            
        return AnalyzeResponse(**parsed)
    except Exception as e:
        logger.error(f"Failed to parse analysis JSON: {content}")
        raise HTTPException(status_code=500, detail=f"Failed to parse analysis result: {e}")

@app.post("/api/v1/analyze", response_model=AnalyzeResponse)
async def analyze_endpoint(
    product_img: UploadFile = File(...),
    ref_img: UploadFile = File(...),
    category: str = Form(...),
    custom_product_name: Optional[str] = Form(None),
    api_url: str = Form(None),
    gemini_api_key: str = Form(None),
    model_name: str = Form(None),
    gen_count: int = Form(9),  # User-selected number of scripts to generate
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    # Resolve Config
    if not api_url or not gemini_api_key:
        db_config = get_config(db)
        if not api_url: api_url = db_config.api_url
        if not gemini_api_key: gemini_api_key = db_config.api_key
        # Use provided model or default from config
        if not model_name: model_name = db_config.analysis_model_name or "gemini-3-pro-preview"

    if not api_url or not gemini_api_key:
        raise HTTPException(status_code=400, detail="Missing API Configuration")

    product_b64 = await file_to_base64(product_img)
    ref_b64 = await file_to_base64(ref_img)
    
    async with httpx.AsyncClient() as client:
        return await analyze_product_scene(
            client, api_url, gemini_api_key, product_b64, ref_b64, category, model_name, custom_product_name, gen_count
        )

# --- Main Endpoint ---
@app.post("/api/v1/batch-generate", status_code=202)
async def batch_generate_workflow(
    product_img: UploadFile = File(...),
    ref_img: UploadFile = File(...),
    scripts: str = Form(...),
    api_url: str = Form(None),
    gemini_api_key: str = Form(None),
    model_name: str = Form(None),
    aspect_ratio: str = Form("1:1"),
    category: str = Form("other"),  # Product category
    scene_style_prompt: str = Form(""),  # Visual style prompt for scene generation
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    # Log Usage
    try:
        # Scripts is JSON string list of objects.
        # Estimate count?
        import json
        scripts_list = json.loads(scripts)
        count = len(scripts_list)
        
        log = ImageGenerationLog(user_id=user.id, count=count)
        db.add(log)
        
        # Log activity for monitoring
        activity = UserActivity(
            user_id=user.id,
            action="image_gen_start",
            details=f"开始生成 {count} 张图片 | 类目: {category} | 比例: {aspect_ratio}"
        )
        db.add(activity)
        db.commit()
        
        # Update user's real-time status and broadcast to admins
        try:
            await connection_manager.update_user_activity(user.id, f"正在生成 {count} 张图片")
        except Exception:
            pass  # WebSocket not required
    except Exception as e:
        logger.error(f"Failed to log usage: {e}")

    print(f"DEBUG: Received batch-generate request. Scripts length: {len(scripts)}", flush=True)
    # Resolve Config
    db_config = None
    if not api_url or not gemini_api_key or not model_name:
        # Fetch from DB if not provided in form (e.g. from frontend state bugs)
        db_config_list = db.query(SystemConfig).all()
        config_dict = {item.key: item.value for item in db_config_list}
        if not api_url: api_url = config_dict.get("api_url")
        if not gemini_api_key: gemini_api_key = config_dict.get("api_key")
        if not model_name: model_name = config_dict.get("model_name")
    
    # Analysis Model
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    analysis_model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")

    # 1. Read Images
    product_bytes = await product_img.read()
    ref_bytes = await ref_img.read()
    product_b64 = base64.b64encode(product_bytes).decode('utf-8')
    ref_b64 = base64.b64encode(ref_bytes).decode('utf-8')

    # 3. Determine Prompts
    prompts_map = ANGLES_PROMPTS
    if scripts:
        try:
            import json
            script_list = json.loads(scripts)
            # script_list should be [{'angle_name': '...', 'script': '...'}, ...]
            prompts_map = { item['angle_name']: item['script'] for item in script_list }
        except Exception as e:
            logger.error(f"Failed to parse scripts: {e}")
            pass
    print(f"DEBUG: Prompts Map size: {len(prompts_map)} Keys: {list(prompts_map.keys())}", flush=True)

    # 4. Concurrency
    sem = asyncio.Semaphore(3) 

    async def clean_prompt_for_video(client, api_url, api_key, original_prompt, model_name):
        system_instruction = (
            "You are an expert prompt engineer. Your task is to rewrite the given image generation prompt "
            "into a fluent, descriptive, natural language paragraph suitable for a text-to-video model (like Sora or Veo). "
            "Remove any 'Generate a...' commands or system/technical instructions. "
            "Focus on describing the visual subject, lighting, and composition as a continuous scene. "
            "Keep it under 60 words."
        )
        
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": system_instruction},
                {"role": "user", "content": f"Original Prompt: {original_prompt}"}
            ],
            "max_tokens": 100
        }
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        logger.info(f"Cleaning prompt using model: {model_name}")

        try:
            target_url = api_url
            if not target_url.endswith("/chat/completions"):
                 target_url = f"{target_url.rstrip('/')}/chat/completions"
            
            resp = await client.post(target_url, json=payload, headers=headers, timeout=60.0)
            if resp.status_code == 200:
                data = resp.json()
                return data.get("choices", [])[0].get("message", {}).get("content", "").strip()
            else:
                 logger.error(f"Prompt cleaning API error: {resp.text}")
        except Exception as e:
            logger.error(f"Prompt cleaning failed: {e}")
        return original_prompt # Fallback

    async def safe_call(name, prompt):
        print(f"DEBUG: Starting safe_call for {name}", flush=True)
        async with sem:
            async with httpx.AsyncClient() as client:
                # Add Aspect Ratio to prompt if needed
                final_prompt = prompt
                if aspect_ratio:
                     final_prompt = f"{prompt} --ar {aspect_ratio}"

                result = await call_openai_compatible_api(
                    client, api_url, gemini_api_key, product_b64, ref_b64, name, final_prompt, model_name
                )
                
                # Use original Step 2 Prompt directly (User Request)
                # cleaned_prompt = await clean_prompt_for_video(client, api_url, gemini_api_key, prompt, analysis_model_name)
                # logger.info(f"Video Prompt generated for {name}: {cleaned_prompt[:50]}...")
                
                result.video_prompt = prompt
                
                # Verify persistence
                logger.info(f"SafeCall Result for {name}: video_prompt len={len(prompt) if prompt else 0}")
                
                return result

    tasks = []
    for name, prompt in prompts_map.items():
        tasks.append(safe_call(name, prompt))
    
    results = await asyncio.gather(*tasks)
    
    # --- Persistence Logic for Gallery ---
    gallery_dir = "/app/uploads/gallery"
    os.makedirs(gallery_dir, exist_ok=True)
    
    saved_count = 0
    async with httpx.AsyncClient() as download_client:
        for idx, r in enumerate(results):
            logger.info(f"Gallery save check [{idx}]: has_base64={bool(r.image_base64)}, error={r.error}, base64_len={len(r.image_base64) if r.image_base64 else 0}")
            if r.image_base64 and not r.error:
                try:
                    img_data = None
                    b64_data = r.image_base64
                    
                    # Check if it's a URL (API sometimes returns image URL instead of base64)
                    if b64_data.startswith("http://") or b64_data.startswith("https://"):
                        logger.info(f"Downloading image from URL: {b64_data[:100]}...")
                        try:
                            resp = await download_client.get(b64_data, timeout=60.0)
                            if resp.status_code == 200:
                                img_data = resp.content
                                logger.info(f"Downloaded image: {len(img_data)} bytes")
                            else:
                                logger.error(f"Failed to download image: HTTP {resp.status_code}")
                                continue
                        except Exception as dl_err:
                            logger.error(f"Image download error: {dl_err}")
                            continue
                    else:
                        # Handle base64 data
                        if "," in b64_data:
                            b64_data = b64_data.split(",")[1]
                        
                        img_data = base64.b64decode(b64_data)
                    
                    if not img_data:
                        logger.error("No image data to save")
                        continue
                    
                    # Validate that we have valid image data (JPEG or PNG magic bytes)
                    if not (img_data[:2] == b'\xff\xd8' or img_data[:8] == b'\x89PNG\r\n\x1a\n'):
                        logger.error(f"Invalid image data (first bytes: {img_data[:10]})")
                        continue
                    
                    # Determine extension based on magic bytes
                    ext = ".jpg" if img_data[:2] == b'\xff\xd8' else ".png"
                    
                    # 2. Save to Disk
                    filename = f"gen_{user.id}_{uuid.uuid4().hex}{ext}"
                    file_path = os.path.join(gallery_dir, filename)
                    
                    with open(file_path, "wb") as f:
                        f.write(img_data)
                    
                    logger.info(f"Saved gallery image: {filename} ({len(img_data)} bytes)")
                    
                    # Get image dimensions
                    img_width, img_height = None, None
                    try:
                        from PIL import Image
                        from io import BytesIO
                        img_pil = Image.open(BytesIO(img_data))
                        img_width, img_height = img_pil.size
                    except Exception as dim_err:
                        logger.warning(f"Could not get image dimensions: {dim_err}")
                        
                    # 3. Save to DB with category and dimensions
                    # Note: r.video_prompt holds the prompt used for this image
                    new_image = SavedImage(
                        user_id=user.id,
                        filename=filename,
                        file_path=file_path,
                        url=f"/uploads/gallery/{filename}",
                        prompt=r.video_prompt or r.angle_name,  # Fallback
                        width=img_width,
                        height=img_height,
                        category=category,  # Use category from request
                        is_shared=user.default_share if user.default_share is not None else True
                    )
                    db.add(new_image)
                    saved_count += 1
                except Exception as e:
                    logger.error(f"Failed to save gallery image: {e}")
    
    if saved_count > 0:
        db.commit()
    # -------------------------------------

    # Debug results
    valid_results = []
    for r in results:
        r_dict = r.dict()
        # Explicitly ensure video_prompt is copied if missing (though .dict() should work)
        if r.video_prompt and not r_dict.get("video_prompt"):
             r_dict["video_prompt"] = r.video_prompt
        valid_results.append(r_dict)
        
    logger.info(f"Final Batch Results Sample: {valid_results[0].keys()} has_prompt={'video_prompt' in valid_results[0]}")
    
    # Log completion activity and reset user status
    try:
        activity = UserActivity(
            user_id=user.id,
            action="image_gen_complete",
            details=f"图片生成完成 | 生成 {len(valid_results)} 张 | 类目: {category}"
        )
        db.add(activity)
        db.commit()
        
        # Reset user status to idle
        await connection_manager.update_user_activity(user.id, "空闲")
    except Exception as act_err:
        logger.warning(f"Failed to log image completion: {act_err}")
    
    return {
        "status": "completed",
        "total_generated": len(valid_results),
        "results": valid_results
    }


# --- Simple Batch Generate (Multiple Reference Images) ---
async def call_multi_image_gen(
    client: httpx.AsyncClient,
    api_url: str,
    api_key: str,
    image_b64_list: List[str],
    scene_prompt: str,
    variation_index: int,
    model: str = "gemini-3-pro-image-preview",
    aspect_ratio: str = "1:1"
) -> ImageResult:
    """Generate image using multiple input images as reference + text prompt"""
    angle_name = f"Result_{variation_index + 1}"
    logger.info(f"Multi-Image Gen for {angle_name} with {len(image_b64_list)} input images, model: {model}")
    
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    timeout = httpx.Timeout(connect=15.0, read=300.0, write=30.0, pool=30.0)
    
    if "imagen" in model.lower():
        logger.info(f"Using /images/generations endpoint for imagen model: {model}")
        target_url = api_url.rstrip("/")
        if target_url.endswith("/v1"):
            target_url = f"{target_url}/images/generations"
        elif not target_url.endswith("/images/generations"):
            target_url = f"{target_url}/images/generations"
        
        size_map = {
            "1:1": "1024x1024",
            "9:16": "768x1344",
            "16:9": "1344x768",
            "4:5": "896x1120",
            "3:4": "896x1152"
        }
        image_size = size_map.get(aspect_ratio, "1024x1024")
        
        payload = {
            "model": model,
            "prompt": scene_prompt,
            "n": 1,
            "size": image_size
        }
        
        max_retries = 3
        for attempt in range(max_retries + 1):
            try:
                await throttle_request()
                logger.info(f"Imagen request for {angle_name} (Attempt {attempt+1}/{max_retries+1}) to {target_url}")
                response = await client.post(target_url, json=payload, headers=headers, timeout=timeout)
                
                if response.status_code == 429:
                    if attempt < max_retries:
                        wait_time = 2.0 * (2 ** attempt)
                        logger.warning(f"Rate limited (429). Retrying in {wait_time:.1f}s...")
                        await asyncio.sleep(wait_time)
                        continue
                    return ImageResult(angle_name=angle_name, error="Rate Limit Exceeded (429)")
                
                if response.status_code != 200:
                    error_msg = response.text[:300]
                    logger.error(f"Imagen API error: {response.status_code} - {error_msg}")
                    if attempt < max_retries:
                        await asyncio.sleep(2)
                        continue
                    return ImageResult(angle_name=angle_name, error=f"API Error {response.status_code}")
                
                data = response.json()
                if data.get("data") and len(data["data"]) > 0:
                    item = data["data"][0]
                    if "b64_json" in item:
                        logger.info(f"Got base64 image for {angle_name}, length: {len(item['b64_json'])}")
                        return ImageResult(angle_name=angle_name, image_base64=item["b64_json"], video_prompt=scene_prompt)
                    elif "url" in item:
                        logger.info(f"Got image URL for {angle_name}: {item['url'][:80]}...")
                        return ImageResult(angle_name=angle_name, image_url=item["url"], video_prompt=scene_prompt)
                
                logger.warning(f"Empty data array from imagen API for {angle_name}")
                if attempt < max_retries:
                    await asyncio.sleep(2)
                    continue
                return ImageResult(angle_name=angle_name, error="No image data returned from API")
                
            except Exception as e:
                logger.error(f"Exception in Imagen Gen for {angle_name}: {type(e).__name__}: {e}")
                if attempt < max_retries:
                    await asyncio.sleep(2)
                    continue
                return ImageResult(angle_name=angle_name, error=str(e))
        
        return ImageResult(angle_name=angle_name, error="Max retries exceeded")
    
    system_instruction = (
        "You are a professional product photographer and image compositing expert. "
        "=== YOUR TASK ===\n"
        "You will receive multiple product/reference images. Analyze ALL of them together and create a NEW composite image.\n"
        "🔴 PRESERVE the products exactly as shown - shapes, colors, textures, logos, details.\n"
        "🔴 Combine elements from the reference images as instructed in the prompt.\n"
        "🔴 Create a cohesive scene that incorporates the products naturally.\n"
        "\n=== SCENE GENERATION RULES ===\n"
        "✅ DO: Create backgrounds, lighting, shadows consistent with the prompt.\n"
        "✅ DO: Position and arrange products naturally in the scene.\n"
        "✅ DO: Follow the scene description in the prompt precisely.\n"
        "✅ Return ONLY the final composited image."
    )
    
    image_count = len(image_b64_list)
    user_prompt_text = (
        f"=== INPUT IMAGES ({image_count} images) ===\n"
        f"Reference these {image_count} image(s) to create the final composition.\n\n"
        f"=== SCENE INSTRUCTION ===\n"
        f"{scene_prompt}"
    )
    
    content_parts = [{"type": "text", "text": user_prompt_text}]
    for img_b64 in image_b64_list:
        content_parts.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}})
    
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_instruction},
            {"role": "user", "content": content_parts}
        ],
        "temperature": 0.3,
        "max_tokens": 4096,
        "stream": True
    }
    
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    max_retries = 3
    base_delay = 2.0
    
    # Build target URL
    target_url = api_url
    if not target_url.endswith("/chat/completions") and not target_url.endswith(":generateContent"):
        target_url = f"{target_url.rstrip('/')}/chat/completions"
    
    for attempt in range(max_retries + 1):
        try:
            await throttle_request()
            logger.info(f"Multi-Image Gen request for {angle_name} (Attempt {attempt+1}/{max_retries+1}) to {target_url}")
            timeout = httpx.Timeout(connect=15.0, read=300.0, write=30.0, pool=30.0)
            async with client.stream("POST", target_url, json=payload, headers=headers, timeout=timeout) as response:
                if response.status_code == 429:
                    if attempt < max_retries:
                        wait_time = base_delay * (2 ** attempt) * random.uniform(0.8, 1.5)
                        logger.warning(f"Rate limited (429). Retrying in {wait_time:.1f}s...")
                        await asyncio.sleep(wait_time)
                        continue
                    return ImageResult(angle_name=angle_name, error="Rate Limit Exceeded (429)")
                
                if response.status_code == 524:
                    if attempt < max_retries:
                        await asyncio.sleep(2)
                        continue
                    return ImageResult(angle_name=angle_name, error="服务器处理超时，请稍后重试")
                
                if response.status_code != 200:
                    error_text = await response.aread()
                    return ImageResult(angle_name=angle_name, error=f"API Error {response.status_code}: {error_text.decode('utf-8')[:200]}")
                
                full_content = ""
                raw_lines = []
                chunk_count = 0
                async for line in response.aiter_lines():
                    if line.strip():
                        raw_lines.append(line)
                    if line.startswith("data: "):
                        data_str = line[6:]
                        if data_str.strip() == "[DONE]":
                            break
                        try:
                            chunk = json.loads(data_str)
                            chunk_count += 1
                            if chunk_count <= 3:
                                logger.info(f"Chunk {chunk_count} for {angle_name}: {json.dumps(chunk)[:500]}")
                            
                            choices = chunk.get("choices", [])
                            if choices:
                                choice = choices[0]
                                delta = choice.get("delta", {})
                                
                                # Check for images array (gemini-3-pro-image-preview format)
                                images = delta.get("images", [])
                                if images:
                                    for img in images:
                                        if img.get("type") == "image_url":
                                            img_url = img.get("image_url", {}).get("url", "")
                                            if img_url:
                                                full_content += img_url
                                
                                delta_content = delta.get("content", "")
                                if delta_content:
                                    full_content += delta_content
                                message = choice.get("message", {})
                                
                                # Also check message.images
                                msg_images = message.get("images", [])
                                if msg_images:
                                    for img in msg_images:
                                        if img.get("type") == "image_url":
                                            img_url = img.get("image_url", {}).get("url", "")
                                            if img_url:
                                                full_content += img_url
                                
                                msg_content = message.get("content", "")
                                if msg_content:
                                    full_content += msg_content
                        except Exception as parse_err:
                            logger.warning(f"Chunk parse error: {parse_err}, line: {data_str[:200]}")
                
                logger.info(f"Total chunks for {angle_name}: {chunk_count}, content length: {len(full_content)}")
                
                if not full_content:
                    logger.warning(f"Empty content for {angle_name}. Raw lines count: {len(raw_lines)}, first 3: {raw_lines[:3]}")
                    if raw_lines and raw_lines[0].lower().startswith("<!doctype"):
                        return ImageResult(angle_name=angle_name, error="API返回了HTML错误页面，请稍后重试")
                    return ImageResult(angle_name=angle_name, error="No content in response")
                
                content_sample = full_content[:50].lower()
                if content_sample.startswith("<!doctype") or "<html" in content_sample:
                    if attempt < max_retries:
                        await asyncio.sleep(2)
                        continue
                    return ImageResult(angle_name=angle_name, error="Received HTML error page")
                
                img_match = re.search(r'!\[.*?\]\((https?://[^\)]+)\)', full_content)
                if img_match:
                    logger.info(f"Found image URL for {angle_name}: {img_match.group(1)}")
                    return ImageResult(angle_name=angle_name, image_url=img_match.group(1), video_prompt=scene_prompt)
                
                b64_match = re.search(r'data:image/[^;]+;base64,([A-Za-z0-9+/=]+)', full_content)
                if b64_match:
                    return ImageResult(angle_name=angle_name, image_base64=b64_match.group(1), video_prompt=scene_prompt)
                
                if len(full_content) > 1000 and re.match(r'^[A-Za-z0-9+/=]+$', full_content.strip()):
                    return ImageResult(angle_name=angle_name, image_base64=full_content.strip(), video_prompt=scene_prompt)
                
                return ImageResult(angle_name=angle_name, error=f"Could not extract image from response: {full_content[:100]}")
                
        except Exception as e:
            logger.error(f"Exception in Multi-Image Gen for {angle_name}: {type(e).__name__}: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            if attempt < max_retries:
                await asyncio.sleep(base_delay)
                continue
            return ImageResult(angle_name=angle_name, error=str(e))
    
    return ImageResult(angle_name=angle_name, error="Max retries exceeded")


@app.post("/api/v1/simple-batch-generate", status_code=202)
async def simple_batch_generate(
    product_imgs: List[UploadFile] = File(...),
    prompt: str = Form(...),
    category: str = Form("other"),
    aspect_ratio: str = Form("1:1"),
    scene_style_prompt: str = Form(""),
    gen_count: int = Form(3),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Simple batch generate: Upload 1-4 product images, AI references ALL images together to generate N result images.
    All uploaded images are sent together to the AI as reference material.
    """
    if len(product_imgs) > 1:
        raise HTTPException(status_code=400, detail="当前仅支持单张图片生成，多图融合功能暂不可用")
    if len(product_imgs) == 0:
        raise HTTPException(status_code=400, detail="请上传1张产品图")
    if gen_count < 1 or gen_count > 9:
        raise HTTPException(status_code=400, detail="生成数量必须在1-9之间")
    
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("model_name", "gemini-3-pro-image-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="Missing API Configuration")
    
    try:
        log = ImageGenerationLog(user_id=user.id, count=gen_count)
        db.add(log)
        activity = UserActivity(
            user_id=user.id,
            action="simple_batch_start",
            details=f"开始简单批量生成 | {len(product_imgs)}张参考图 → 生成{gen_count}张结果"
        )
        db.add(activity)
        db.commit()
        await connection_manager.update_user_activity(user.id, f"正在生成 {gen_count} 张图片")
    except Exception as e:
        logger.error(f"Failed to log usage: {e}")
    
    final_prompt = prompt
    if scene_style_prompt:
        final_prompt = f"{prompt}. Visual style: {scene_style_prompt}"
    if aspect_ratio and aspect_ratio != "1:1":
        final_prompt = f"{final_prompt} --ar {aspect_ratio}"
    
    image_b64_list = []
    for product_file in product_imgs:
        b64 = await file_to_base64_compressed(product_file, max_size=800, quality=75)
        image_b64_list.append(b64)
    
    all_results = []
    sem = asyncio.Semaphore(1)
    
    async def generate_one_result(var_index: int):
        async with sem:
            async with httpx.AsyncClient() as client:
                result = await call_multi_image_gen(
                    client, api_url, api_key, image_b64_list,
                    final_prompt, var_index, model_name, aspect_ratio
                )
                result.angle_name = f"Result_{var_index + 1}"
                return result
    
    tasks = [generate_one_result(i) for i in range(gen_count)]
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    for idx, r in enumerate(results):
        if isinstance(r, Exception):
            logger.error(f"Failed to generate result {idx}: {r}")
            all_results.append({"error": str(r), "index": idx})
        else:
            all_results.append(r)
    
    gallery_dir = "/app/uploads/gallery"
    os.makedirs(gallery_dir, exist_ok=True)
    saved_count = 0
    final_results = []
    
    async with httpx.AsyncClient() as download_client:
        for idx, r in enumerate(all_results):
            if isinstance(r, dict) and "error" in r:
                final_results.append(r)
                continue
            
            if hasattr(r, 'dict'):
                r_dict = r.dict()
            else:
                r_dict = r if isinstance(r, dict) else {"error": "Invalid result"}
                continue
            
            has_image = getattr(r, 'image_base64', None) or getattr(r, 'image_url', None)
            if has_image:
                try:
                    img_data = None
                    if r.image_base64:
                        b64 = r.image_base64
                        if b64.startswith("data:"):
                            b64 = b64.split(",", 1)[1]
                        img_data = base64.b64decode(b64)
                    elif r.image_url:
                        resp = await download_client.get(r.image_url, timeout=30.0)
                        if resp.status_code == 200:
                            img_data = resp.content
                    
                    if img_data:
                        timestamp = int(time.time() * 1000)
                        filename = f"simple_batch_{user.id}_{timestamp}_{saved_count}.png"
                        file_path = os.path.join(gallery_dir, filename)
                        with open(file_path, "wb") as f:
                            f.write(img_data)
                        
                        img_width, img_height = 1024, 1024
                        try:
                            from PIL import Image
                            with Image.open(file_path) as img:
                                img_width, img_height = img.size
                        except:
                            pass
                        
                        new_image = SavedImage(
                            user_id=user.id,
                            filename=filename,
                            file_path=file_path,
                            url=f"/uploads/gallery/{filename}",
                            prompt=getattr(r, 'video_prompt', None) or prompt,
                            width=img_width,
                            height=img_height,
                            category=category,
                            is_shared=user.default_share if user.default_share is not None else True
                        )
                        db.add(new_image)
                        saved_count += 1
                        r_dict["saved_url"] = f"/uploads/gallery/{filename}"
                except Exception as e:
                    logger.error(f"Failed to save image: {e}")
            
            r_dict["result_index"] = idx
            final_results.append(r_dict)
    
    if saved_count > 0:
        db.commit()
    
    try:
        await connection_manager.update_user_activity(user.id, "空闲")
    except:
        pass
    
    return {
        "status": "completed",
        "input_images": len(product_imgs),
        "total_generated": len(final_results),
        "saved_to_gallery": saved_count,
        "results": final_results
    }


# --- Video Queue Models ---


# --- Pydantic Models for Queue ---
class QueueItemCreate(BaseModel):
    prompt: str

class QueueItemUpdate(BaseModel):
    status: Optional[str] = None
    result_url: Optional[str] = None
    error_msg: Optional[str] = None

class StoryShot(BaseModel):
    shot: int
    prompt: str
    duration: int
    description: str
    shotStory: str
    heroSubject: Optional[str] = None

class StoryAnalysisResponse(BaseModel):
    shots: List[StoryShot]

@app.post("/api/v1/story-analyze", response_model=StoryAnalysisResponse)
async def analyze_storyboard_endpoint(
    image: UploadFile = File(...),
    topic: str = Form("一个产品的故事"), # Default topic if missing
    shot_count: int = Form(5), # Default 5 shots
    category: str = Form("other"),  # Product category for tailored prompts
    api_url: str = Form(None),
    gemini_api_key: str = Form(None),
    model_name: str = Form(None),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    # Resolve Config (Same pattern as other endpoints)
    db_config = None
    if not api_url or not gemini_api_key or not model_name:
        db_config_list = db.query(SystemConfig).all()
        config_dict = {item.key: item.value for item in db_config_list}
        if not api_url: api_url = config_dict.get("api_url")
        if not gemini_api_key: gemini_api_key = config_dict.get("api_key")
        # Use analysis model for script generation if available, else default
        if not model_name: model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    # Read Image
    image_bytes = await file_to_base64(image)
    
    # Category-specific placement logic (matching batch scene generation)
    category_guidance = {
        "security": "Security/Surveillance - Wall-mounted scenes, professional spaces (control rooms, corridors, building exteriors), technical precision, night vision/IR effects implied, industrial-grade aesthetics",
        "daily": "Daily Essentials - Home living scenes, natural lighting, warm atmosphere, human daily use interactions, organization/storage contexts",
        "beauty": "Beauty/Cosmetics - Soft textured backgrounds, organic materials (petals, silk, water droplets), feminine aesthetics, delicate close-ups, skin texture implied",
        "electronics": "Electronics/Tech - Minimalist surfaces, tech atmosphere, floating product effects, LED lighting, screen displays, metallic reflections",
        "other": "General Product - Flexibly choose scenes based on product characteristics"
    }
    category_hint = category_guidance.get(category, category_guidance["other"])
    
    # Construct System Prompt with VideoGenerationPromptGuide integration
    system_prompt = f"""Role: You are a specialized assistant combining the skills of:
- A film storyboard artist creating continuous visual narratives
- A prompt engineer crafting clear, structured video generation prompts  
- A creative director ensuring coherent storytelling and strong visuals
- A safety reviewer ensuring strict policy compliance

=== PRODUCT CATEGORY GUIDANCE ===
**Category**: {category_hint}
Use this category to guide scene selection, lighting styles, and product placement.
Tailor your storyboard shots to match the expected environment and aesthetic for this product type.

=== SAFETY CONSTRAINTS (HIGHEST PRIORITY) ===
Prohibited content - NEVER include:
- Explicit sexual content, sexual acts, or strong innuendo
- Graphic violence, gore, blood, open wounds, or detailed injuries
- Hate speech, harassment, or demeaning content toward protected groups
- Praise of extremist organizations, symbols, or slogans
- Instructions for dangerous or illegal activities (weapons, drugs, crimes)
- Real-world personal data (PII), names with identifying information
- Highly realistic depictions of specific real celebrities or public figures
- Minors in any sexual, violent, exploitative, or unsafe context

Safe substitutes:
- "Sexy/seductive" → "stylish and confident", "elegant attire", "charismatic presence"
- Violence → "tense atmosphere", "distant flashes suggesting conflict", focus on emotional stakes
- Real celebrities → "a famous [role] archetype" (e.g., "a charismatic tech CEO archetype")

=== VIDEO PROMPT STRUCTURE ===
Each prompt must implicitly contain:
1. Subject: Who/what is at center (person, object, environment, abstract)
   - Use roles/archetypes over specific individuals
   - Describe relevant visual traits (clothing, posture, age group)

2. Action: What happens - use strong verbs
   - "explaining", "demonstrating", "pointing", "walking", "assembling", "observing", "testing", "exploring"
   - Avoid explicit, violent, or unsafe actions

3. Scene/Context: Where/when/atmosphere
   - Location (indoor/outdoor/virtual)
   - Time of day (morning, sunset, night, golden hour)
   - Environmental details (weather, textures, background elements)

4. Camera: Angle and framing
   - Eye-level medium shot, wide establishing shot, close-up, bird's-eye view
   - Keep simple: 1-2 main angles per shot

5. Camera Movement (optional):
   - Stable shot for clarity, slow pan for scenery, gentle push-in for emphasis
   - Limit to one primary movement per short clip

6. DYNAMIC ELEMENTS (REQUIRED - not just camera movement!):
   - Each shot MUST include real motion in the scene, not just camera panning
   - Examples: particles floating, steam rising, reflections changing, hands interacting with product
   - Environmental motion: leaves, wind effects, light rays, shadows shifting
   - Avoid: static scenes where only the camera moves while everything else is frozen
   
   *** MANDATORY DYNAMIC REQUIREMENTS ***
   - Minimum 2 distinct motion elements per shot
   - Motion type categories (use at least 2 per shot):
     * AMBIENT: dust particles, motes of light, gentle haze, floating elements
     * ENVIRONMENTAL: wind ripples, water reflections, foliage movement, fabric flutter
     * SUBJECT: product interaction, hand gestures, expression changes, posture shifts
     * LIGHT: ray movement, caustics, shadow rotation, highlight shimmer
   - Intensity progression through story arc:
     * Opening shots: subtle, establishing motion
     * Middle shots: increased dynamics, peak action
     * Closing shots: calming motion, resolution feel

7. Visual Style:
   - Lighting: natural/artificial, mood (soft, dramatic, moody)
   - Tone: "warm and welcoming", "calm and meditative", "modern and energetic"
   - Art style: "photorealistic cinematic", "2D animation", "hand-drawn illustration"
   - Atmosphere: haze, dust particles, rain reflections, light beams

=== STORYBOARD REQUIREMENTS ===
Goal: Create a continuous storyboard with EXACTLY {shot_count} shots for the story: "{topic}".

Global visual style: Filmic realism, natural lighting, soft bokeh, 35mm lens, muted colors, subtle grain.

*** CRITICAL NARRATIVE REQUIREMENTS ***
1. **Continuous Story Arc**:
   - Begin (Shot 1): Set scene, introduce subject.
   - Middle (Shot 2-{shot_count-1}): Action, movement, conflict.
   - End (Shot {shot_count}): Resolution.
2. **Visual Consistency**: 
   - Define a "Hero Subject" in Shot 1 based on the provided image.
3. **Seamless Flow**: 
   - Each prompt must pick up IMMEDIATELY where the previous left off.
4. **Causal Relationship**: 
   - Explain WHY transition happens.
5. **Temporal Continuity**:
   - Strict chronological order.

=== OUTPUT FORMAT ===
Output: ONLY a raw JSON array (no markdown code fences).

Required fields per shot:
- shot: integer (1..{shot_count})
- prompt: detailed safe video generation prompt (English), following the structure above
- duration: integer (always 15)
- description: Concise Chinese action summary
- shotStory: Narrative description in Chinese (2-3 sentences) showing connection to previous shot
- heroSubject: (Required in ALL shots) Detailed visual description of the main product/subject from the image - must be IDENTICAL across all shots for consistency

=== SELF-CHECK BEFORE OUTPUT ===
✔ No explicit sexual content or innuendo
✔ No graphic violence, gore, or injuries
✔ No hate speech or extremist content
✔ No real celebrities or PII
✔ No minors in unsafe scenarios
✔ Each prompt has: subject, action, scene, camera, style
✔ Story flows continuously between shots
"""

    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user", 
                "content": [
                    {"type": "text", "text": f"Generate storyboard for: {topic}"},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_bytes}"}}
                ]
            }
        ],
        "max_tokens": 8192
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {gemini_api_key}"
    }
    
    logger.info(f"Analyzing storyboard for topic: {topic}")
    
    async with httpx.AsyncClient() as client:
        try:
            target_url = api_url
            if not target_url.endswith("/chat/completions"):
                 target_url = f"{target_url.rstrip('/')}/chat/completions"
            
            resp = await client.post(target_url, json=payload, headers=headers, timeout=60.0)
            if resp.status_code == 200:
                data = resp.json()
                content = data.get("choices", [])[0].get("message", {}).get("content", "").strip()
                
                # Clean markdown if present
                if content.startswith("```"):
                    lines = content.splitlines()
                    # Remove first line if it starts with ```
                    if lines[0].startswith("```"):
                        lines = lines[1:]
                    # Remove last line if it starts with ```
                    if lines and lines[-1].strip().startswith("```"):
                        lines = lines[:-1]
                    content = "\n".join(lines).strip()
                
                # Fix trailing commas which cause json.loads to fail
                import re
                content = re.sub(r',\s*\}', '}', content)
                content = re.sub(r',\s*\]', ']', content)
                
                try:
                    shots_data = json.loads(content)
                    
                    # --- FIX: Ensure exact number of shots ---
                    if len(shots_data) < shot_count:
                        logger.warning(f"AI generated {len(shots_data)} shots, but user requested {shot_count}. Padding...")
                        
                        original_count = len(shots_data)
                        if original_count > 0:
                            import copy
                            needed = shot_count - original_count
                            for i in range(needed):
                                # Round-robin selection from original shots
                                source_shot = shots_data[i % original_count]
                                new_shot = copy.deepcopy(source_shot)
                                new_shot["shot"] = original_count + i + 1
                                new_shot["description"] = f"{source_shot.get('description', 'Scene')} (Variation {i+1})"
                                shots_data.append(new_shot)
                        else:
                            # Fallback if 0 shots returned (rare)
                            for i in range(shot_count):
                                shots_data.append({
                                    "shot": i + 1,
                                    "prompt": f"Product showcase shot {i+1}, professional lighting, clean composition.",
                                    "duration": 15,
                                    "description": f"Auto-generated shot {i+1}",
                                    "shotStory": "Auto-generated content",
                                    "heroSubject": "Product"
                                })
                    
                    # Truncate if too many (rare but possible)
                    if len(shots_data) > shot_count:
                        shots_data = shots_data[:shot_count]
                    # ---------------------------------------------
                    
                    return StoryAnalysisResponse(shots=[StoryShot(**s) for s in shots_data])
                except Exception as e:
                     logger.error(f"Failed to parse JSON: {content} - Error: {e}")
                     raise HTTPException(status_code=500, detail="Failed to parse storyboard JSON")
            else:
                 logger.error(f"Analysis API Error: {resp.text}")
                 raise HTTPException(status_code=resp.status_code, detail=f"API Error: {resp.text}")
        except Exception as e:
            logger.error(f"Storyboard Analysis Failed: {e}")
            raise HTTPException(status_code=500, detail=str(e))

class VideoPromptResponse(BaseModel):
    video_prompt: str

@app.post("/api/v1/generate-video-prompt", response_model=VideoPromptResponse)
async def generate_video_prompt_endpoint(
    image: UploadFile = File(...),
    api_url: str = Form(None),
    gemini_api_key: str = Form(None),
    model_name: str = Form(None),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    # Resolve Config
    db_config = None
    if not api_url or not gemini_api_key or not model_name:
        db_config = get_config(db)

    if not api_url: api_url = db_config.api_url
    if not gemini_api_key: gemini_api_key = db_config.api_key
    
    # Use analysis model for prompt generation
    if not model_name: 
        model_name = db_config.analysis_model_name or "gemini-3-pro-preview"
        
    logger.info(f"Generating video prompt using model: {model_name} (from request: {model_name != (db_config.analysis_model_name or 'gemini-3-pro-preview')})")

    if not api_url or not gemini_api_key:
        raise HTTPException(status_code=400, detail="Missing API Configuration")

    image_b64 = await file_to_base64(image)

    # Enhanced video prompt generation with structured framework
    system_instruction = """Role: You are a cinematic video director, motion graphics specialist, and visual effects artist.

=== VIDEO GENERATION FRAMEWORK ===
Analyze the product image and generate a detailed text-to-video prompt (60-80 words) that captures:

1. SUBJECT FOCUS
   - What is the visual center of attention?
   - Key visual characteristics that must be preserved

2. CAMERA MOTION (Primary)
   - Choose ONE primary movement: slow orbit, dolly in/out, crane sweep, push-in, pull-back, tracking shot, static with subtle sway
   - Specify direction, speed, and arc

3. DYNAMIC SCENE ELEMENTS (Required - Beyond Camera!)
   - Ambient motion: dust particles, light rays, subtle reflections
   - Environmental: gentle air flow, fabric movement, liquid ripples
   - Subject interaction: product rotation, gentle vibration, surface highlights shifting

4. LIGHTING EVOLUTION
   - How light behaves during the clip: stable, slowly shifting, pulsing, flickering
   - Highlight and shadow movement

5. ATMOSPHERE & DEPTH
   - Bokeh quality, haze, lens flares, depth layers
   - Color mood: warm, cool, neutral, dramatic

=== OUTPUT CONSTRAINTS ===
- 60-80 words, English only
- Cinematic 24fps aesthetic with subtle film grain
- Smooth, professional camera movement
- Rich depth of field with product in sharp focus
- NO markdown, NO formatting, NO quotes - just the prompt text"""
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_instruction},
            {
                "role": "user", 
                "content": [
                    {"type": "text", "text": "Analyze this product image and generate a cinematic video prompt with camera movement, dynamic elements, and atmospheric details."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            }
        ],
        "max_tokens": 200
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {gemini_api_key}"
    }
    
    async with httpx.AsyncClient() as client:
        try:
            target_url = api_url
            if not target_url.endswith("/chat/completions"):
                 target_url = f"{target_url.rstrip('/')}/chat/completions"
            
            resp = await client.post(target_url, json=payload, headers=headers, timeout=60.0)
            if resp.status_code == 200:
                data = resp.json()
                prompt = data.get("choices", [])[0].get("message", {}).get("content", "").strip()
                return VideoPromptResponse(video_prompt=prompt)
            else:
                raise HTTPException(status_code=resp.status_code, detail=f"API Error: {resp.text}")
        except Exception as e:
            logger.error(f"Video prompt gen failed: {e}")
            raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v1/story-generate")
async def generate_story_endpoint(
    image: UploadFile = File(...),
    shots_json: str = Form(...),
    api_url: str = Form(None),
    gemini_api_key: str = Form(None),
    model_name: str = Form(None),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    # Config
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    if not api_url: api_url = config_dict.get("api_url")
    if not gemini_api_key: gemini_api_key = config_dict.get("api_key")
    if not model_name: model_name = config_dict.get("model_name")
    
    # Parse Scenes
    try:
        shots_data = json.loads(shots_json)
        shots = [StoryShot(**s) for s in shots_data]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")

    # Read Original Product
    original_bytes = await image.read()
    original_b64 = base64.b64encode(original_bytes).decode('utf-8')
    
    generated_results = []
    
    current_ref_b64 = original_b64 
    
    # Extract Hero Subject from Shot 1 to prepend to all
    hero_description = ""
    for s in shots:
        if s.shot == 1 and s.heroSubject:
            hero_description = s.heroSubject
            break
            
    async with httpx.AsyncClient() as client:
        for shot in shots:
            logger.info(f"Generating Shot {shot.shot}...")
            
            final_prompt = shot.prompt
            
            # Prepare Prompt with Injection
            system_instruction = (
                "Role: Cinematic frame artist. \n"
                "Goal: Render a single storyboard frame matching the style. \n"
                f"CRITICAL: Main Character: {hero_description} \n"
                "Style: Filmic realism, natural lighting, soft bokeh, 35mm lens. \n"
            )
            
            if shot.shot > 1:
                system_instruction += "Continuity: Maintain exact style continuity with previous shot. \n"
                final_prompt = (
                     f"Reference image above shows result of previous shot. "
                     f"Generate new frame where SAME character performs: {shot.prompt}"
                )
            else:
                 system_instruction += "Continuity: Establish base look. \n"
            
            full_prompt = (
                f"{system_instruction}\n"
                f"Frame Description: {final_prompt}\n"
                "Constraints: no text, 16:9, high fidelity."
            )
            
            try:
                # We use 'product_b64' slot for the main input image
                result = await call_openai_compatible_api(
                    client, api_url, gemini_api_key, 
                    current_ref_b64, # Input Image (Product for Shot 1, Prev result for Shot 2+)
                    original_b64,    # Ref Image (Always Original Product)
                    f"Shot {shot.shot}", 
                    full_prompt, 
                    model_name
                )
                
                if result.image_base64:
                    # Success
                    generated_results.append({
                        "shot": shot.shot,
                        "image_base64": result.image_base64,
                        "prompt": result.video_prompt or shot.prompt,
                        "description": shot.description,
                        "shotStory": shot.shotStory
                    })
                    # Update ref for next shot
                    current_ref_b64 = result.image_base64
                else:
                    logger.error(f"Shot {shot.shot} generation failed: No image returned.")
                    generated_results.append({"shot": shot.shot, "error": "Generation Failed"})
                    
            except Exception as e:
                logger.error(f"Shot {shot.shot} error: {e}")
                generated_results.append({"shot": shot.shot, "error": str(e)})

    return {
        "status": "completed",
        "results": generated_results
    }

class QueueItemResponse(BaseModel):
    id: str
    filename: str
    prompt: str
    status: str
    result_url: Optional[str] = None
    error_msg: Optional[str] = None
    created_at: datetime
    preview_url: Optional[str] = None
    user_id: Optional[int] = None  # 任务所有者ID，用于前端权限判断
    
    class Config:
        from_attributes = True


# --- Gallery Endpoints ---

@app.get("/api/v1/gallery/images")
def get_gallery_images(
    limit: int = 50, 
    offset: int = 0,
    category: str = None,  # Filter by category
    view_mode: str = "own",  # "own" | "all" | "user" (admin only)
    user_id: int = None,     # Filter by specific user (admin only, requires view_mode="user")
    start_date: str = None,  # Filter by start date (ISO format: YYYY-MM-DD)
    end_date: str = None,    # Filter by end date (ISO format: YYYY-MM-DD)
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    """Get images based on user role and view mode.
    
    - Admin with view_mode='all': All images from all users
    - Admin with view_mode='own': Only admin's own images
    - Admin with view_mode='user' + user_id: Images from specified user
    - Regular user: Own images + shared images (is_shared=True)
    - start_date/end_date: Filter by creation date (ISO format: YYYY-MM-DD)
    """
    from sqlalchemy import or_
    
    if user.role == "admin":
        if view_mode == "user" and user_id is not None:
            # Admin viewing specific user's images
            query = db.query(SavedImage).filter(SavedImage.user_id == user_id)
        elif view_mode == "own":
            query = db.query(SavedImage).filter(SavedImage.user_id == user.id)
        else:  # "all" - admin can see everything
            query = db.query(SavedImage)
    else:
        # Regular user: own images + shared images
        query = db.query(SavedImage).filter(
            or_(SavedImage.user_id == user.id, SavedImage.is_shared == True)
        )
    
    # Apply category filter if provided
    if category and category != "all":
        query = query.filter(SavedImage.category == category)
    
    # Apply date filters if provided
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(SavedImage.created_at >= start_dt)
        except ValueError:
            pass  # Invalid date format, skip filter
    
    if end_date:
        try:
            # Add one day to include the end date fully
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            query = query.filter(SavedImage.created_at < end_dt)
        except ValueError:
            pass  # Invalid date format, skip filter
        
    total = query.count()
    images = query.order_by(SavedImage.created_at.desc()).limit(limit).offset(offset).all()
    
    # Add username and metadata to each image
    result_items = []
    for img in images:
        creator = db.query(User).filter(User.id == img.user_id).first()
        result_items.append({
            "id": img.id,
            "user_id": img.user_id,
            "username": creator.username if creator else "Unknown",
            "filename": img.filename,
            "file_path": img.file_path,
            "url": img.url,
            "prompt": img.prompt,
            "width": img.width,
            "height": img.height,
            "category": img.category or "other",
            "is_shared": img.is_shared,
            "created_at": img.created_at
        })
    
    return {"total": total, "items": result_items}

@app.delete("/api/v1/gallery/images/{image_id}")
def delete_gallery_image(
    image_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Delete a single image. Users can delete their own, admins can delete any."""
    if user.role == "admin":
        image = db.query(SavedImage).filter(SavedImage.id == image_id).first()
    else:
        image = db.query(SavedImage).filter(SavedImage.id == image_id, SavedImage.user_id == user.id).first()
    
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    
    # Delete file from disk
    if image.file_path and os.path.exists(image.file_path):
        try:
            os.remove(image.file_path)
        except Exception as e:
            logger.error(f"Failed to delete gallery image file: {e}")
            
    db.delete(image)
    db.commit()
    return {"status": "deleted"}

class BatchDeleteRequest(BaseModel):
    ids: List[int]

@app.post("/api/v1/gallery/images/batch-delete")
def batch_delete_images(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Batch delete images (admin only)."""
    deleted_count = 0
    for image_id in request.ids:
        image = db.query(SavedImage).filter(SavedImage.id == image_id).first()
        if image:
            if image.file_path and os.path.exists(image.file_path):
                try:
                    os.remove(image.file_path)
                except Exception as e:
                    logger.error(f"Failed to delete file: {e}")
            db.delete(image)
            deleted_count += 1
    db.commit()
    return {"deleted": deleted_count}

class BatchDeleteVideoRequest(BaseModel):
    ids: List[str]

@app.post("/api/v1/gallery/videos/batch-delete")
def batch_delete_videos(
    request: BatchDeleteVideoRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Batch delete videos (admin only)."""
    deleted_count = 0
    for video_id in request.ids:
        video = db.query(VideoQueueItem).filter(VideoQueueItem.id == video_id).first()
        if video:
            # Delete video file
            if video.result_url:
                video_path = video.result_url.replace("/uploads", "/app/uploads")
                if os.path.exists(video_path):
                    try:
                        os.remove(video_path)
                    except Exception as e:
                        logger.error(f"Failed to delete video file: {e}")
            # Delete source image
            if video.file_path and os.path.exists(video.file_path):
                try:
                    os.remove(video.file_path)
                except Exception as e:
                    logger.error(f"Failed to delete source file: {e}")
            db.delete(video)
            deleted_count += 1
    db.commit()
    return {"deleted": deleted_count}


@app.get("/api/v1/gallery/videos")
def get_gallery_videos(
    limit: int = 20,
    offset: int = 0,
    category: str = None,  # Filter by category
    view_mode: str = "own",  # "own" | "all" | "user" (admin only)
    user_id: int = None,     # Filter by specific user (admin only, requires view_mode="user")
    start_date: str = None,  # Filter by start date (ISO format: YYYY-MM-DD)
    end_date: str = None,    # Filter by end date (ISO format: YYYY-MM-DD)
    db: Session = Depends(get_db), 
    user: User = Depends(get_current_user)
):
    """Get videos based on user role and view mode.
    
    - Admin with view_mode='all': All completed videos from all users
    - Admin with view_mode='own': Only admin's own completed videos
    - Admin with view_mode='user' + user_id: Videos from specified user
    - Regular user: Own videos + shared videos (is_shared=True)
    - start_date/end_date: Filter by creation date (ISO format: YYYY-MM-DD)
    """
    from sqlalchemy import or_, and_
    
    # Base filter: only completed videos
    base_filter = VideoQueueItem.status.in_(["done", "archived"])
    
    if user.role == "admin":
        if view_mode == "user" and user_id is not None:
            # Admin viewing specific user's videos
            query = db.query(VideoQueueItem).filter(
                and_(base_filter, VideoQueueItem.user_id == user_id)
            )
        elif view_mode == "own":
            query = db.query(VideoQueueItem).filter(
                and_(base_filter, VideoQueueItem.user_id == user.id)
            )
        else:  # "all" - admin can see everything
            query = db.query(VideoQueueItem).filter(base_filter)
    else:
        # Regular user: own videos + shared videos
        query = db.query(VideoQueueItem).filter(
            and_(
                base_filter,
                or_(VideoQueueItem.user_id == user.id, VideoQueueItem.is_shared == True)
            )
        )
    
    # Apply category filter if provided
    if category and category != "all":
        query = query.filter(VideoQueueItem.category == category)
    
    # Apply date filters if provided
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(VideoQueueItem.created_at >= start_dt)
        except ValueError:
            pass  # Invalid date format, skip filter
    
    if end_date:
        try:
            # Add one day to include the end date fully
            end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
            query = query.filter(VideoQueueItem.created_at < end_dt)
        except ValueError:
            pass  # Invalid date format, skip filter
    
    total = query.count()
    videos = query.order_by(VideoQueueItem.created_at.desc()).limit(limit).offset(offset).all()
    
    # Build response with username and preview_url
    result_items = []
    for vid in videos:
        creator = db.query(User).filter(User.id == vid.user_id).first()
        result_items.append({
            "id": vid.id,
            "filename": vid.filename,
            "file_path": vid.file_path,
            "prompt": vid.prompt,
            "status": vid.status,
            "result_url": vid.result_url,
            "error_msg": vid.error_msg,
            "user_id": vid.user_id,
            "username": creator.username if creator else "Unknown",
            "preview_url": vid.preview_url,
            "category": vid.category or "other",
            "is_merged": vid.is_merged or False,
            "is_shared": vid.is_shared,
            "created_at": vid.created_at,
            # Review fields
            "review_score": vid.review_score,
            "review_status": vid.review_status,
            "reviewed_at": vid.reviewed_at
        })
    
    return {"total": total, "items": result_items}

# --- Video Review Endpoints ---

@app.get("/api/v1/gallery/videos/{video_id}/review")
def get_video_review(
    video_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Get detailed review result for a video."""
    video = db.query(VideoQueueItem).filter(VideoQueueItem.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    # Check access permission
    if user.role != "admin" and video.user_id != user.id and not video.is_shared:
        raise HTTPException(status_code=403, detail="Access denied")
    
    review_details = None
    if video.review_result:
        try:
            review_details = json.loads(video.review_result)
        except:
            review_details = {"raw": video.review_result}
    
    return {
        "video_id": video_id,
        "review_score": video.review_score,
        "review_status": video.review_status,
        "reviewed_at": video.reviewed_at,
        "details": review_details
    }

@app.post("/api/v1/gallery/videos/{video_id}/review")
async def trigger_video_review_api(
    video_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Manually trigger video review (admin only)."""
    video = db.query(VideoQueueItem).filter(VideoQueueItem.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    if video.status != "done":
        raise HTTPException(status_code=400, detail="Video must be completed before review")
    
    # Check if video file exists
    local_path = f"/app/uploads/queue/video_{video_id}.mp4"
    if not os.path.exists(local_path):
        raise HTTPException(status_code=400, detail="Video file not found locally")
    
    # Trigger review (queued for sequential execution)
    try:
        from review_queue import enqueue_video_review
        asyncio.create_task(
            enqueue_video_review(
                video_id=video_id,
                video_path=local_path,
                video_prompt=video.prompt,
                db_session=SessionLocal,
                VideoQueueItem_model=VideoQueueItem
            )
        )
        return {"message": "Review queued", "video_id": video_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to queue review: {str(e)}")

# --- Share Toggle Endpoints ---

@app.post("/api/v1/gallery/images/{image_id}/share")
def toggle_share_image(
    image_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Toggle share status of an image (admin only).
    When is_shared=True, regular users can view this image.
    """
    image = db.query(SavedImage).filter(SavedImage.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    
    image.is_shared = not image.is_shared
    db.commit()
    return {"id": image_id, "is_shared": image.is_shared}

@app.post("/api/v1/gallery/videos/{video_id}/share")
def toggle_share_video(
    video_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Toggle share status of a video (admin only).
    When is_shared=True, regular users can view this video.
    """
    video = db.query(VideoQueueItem).filter(VideoQueueItem.id == video_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    
    video.is_shared = not video.is_shared
    db.commit()
    return {"id": video_id, "is_shared": video.is_shared}

class BatchShareRequest(BaseModel):
    ids: List[int]
    is_shared: bool  # True = share, False = unshare

class BatchShareVideoRequest(BaseModel):
    ids: List[str]
    is_shared: bool

@app.post("/api/v1/gallery/images/batch-share")
def batch_share_images(
    request: BatchShareRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Batch share/unshare images (admin only)."""
    updated_count = 0
    for image_id in request.ids:
        image = db.query(SavedImage).filter(SavedImage.id == image_id).first()
        if image:
            image.is_shared = request.is_shared
            updated_count += 1
    db.commit()
    return {"updated": updated_count, "is_shared": request.is_shared}

@app.post("/api/v1/gallery/videos/batch-share")
def batch_share_videos(
    request: BatchShareVideoRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Batch share/unshare videos (admin only)."""
    updated_count = 0
    for video_id in request.ids:
        video = db.query(VideoQueueItem).filter(VideoQueueItem.id == video_id).first()
        if video:
            video.is_shared = request.is_shared
            updated_count += 1
    db.commit()
    return {"updated": updated_count, "is_shared": request.is_shared}

class ShareAllRequest(BaseModel):
    is_shared: bool  # True = share all, False = unshare all
    skip_first_page: bool = False  # Optional: skip the first N items
    skip_count: int = 0  # Number of items to skip (e.g., first 9 items = 1 page)

@app.post("/api/v1/gallery/images/share-all")
def share_all_images(
    request: ShareAllRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Share/unshare all images (admin only)."""
    query = db.query(SavedImage).order_by(SavedImage.created_at.desc())
    if request.skip_count > 0:
        query = query.offset(request.skip_count)
    images = query.all()
    for image in images:
        image.is_shared = request.is_shared
    db.commit()
    return {"updated": len(images), "is_shared": request.is_shared}

@app.post("/api/v1/gallery/videos/share-all")
def share_all_videos(
    request: ShareAllRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """Share/unshare all completed videos (admin only)."""
    query = db.query(VideoQueueItem).filter(
        VideoQueueItem.status.in_(["done", "archived"])
    ).order_by(VideoQueueItem.created_at.desc())
    if request.skip_count > 0:
        query = query.offset(request.skip_count)
    videos = query.all()
    for video in videos:
        video.is_shared = request.is_shared
    db.commit()
    return {"updated": len(videos), "is_shared": request.is_shared}

# Batch Download Models
class BatchDownloadRequest(BaseModel):
    ids: List[int]

class BatchDownloadVideoRequest(BaseModel):
    ids: List[str]

@app.post("/api/v1/gallery/images/batch-download")
async def batch_download_images(
    request: BatchDownloadRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """批量下载图片 - 打包成 ZIP"""
    ids = request.ids
    if not ids:
        raise HTTPException(status_code=400, detail="未选择任何图片")
    
    # 查询图片
    images = db.query(SavedImage).filter(SavedImage.id.in_(ids)).all()
    
    if not images:
        raise HTTPException(status_code=404, detail="未找到选中的图片")
    
    # 权限检查:仅admin或所有者可下载
    if user.role != "admin":
        for img in images:
            if img.user_id != user.id:
                raise HTTPException(status_code=403, detail="无权下载其他用户的图片")
    
    # 创建临时 ZIP 文件
    with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
        zip_path = tmp.name
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for img in images:
                file_path = Path(img.file_path)
                if file_path.exists():
                    # 使用原文件名,避免重名冲突
                    arcname = f"{img.id}_{img.filename}"
                    zipf.write(file_path, arcname=arcname)
        
        # 返回文件下载响应,下载后自动删除临时文件
        background_tasks.add_task(os.remove, zip_path)
        return FileResponse(
            path=zip_path,
            media_type="application/zip",
            filename=f"images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )

@app.post("/api/v1/gallery/videos/batch-download")
async def batch_download_videos(
    request: BatchDownloadVideoRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """批量下载视频 - 打包成 ZIP"""
    ids = request.ids
    logger.info(f"Batch download videos request: ids={ids}, type={type(ids[0]) if ids else None}")
    
    if not ids:
        raise HTTPException(status_code=400, detail="未选择任何视频")
    
    # 查询视频
    videos = db.query(VideoQueueItem).filter(
        VideoQueueItem.id.in_(ids),
        VideoQueueItem.status.in_(["done", "archived"])
    ).all()
    
    logger.info(f"Found {len(videos)} videos for download")
    
    if not videos:
        raise HTTPException(status_code=404, detail="未找到选中的视频")
    
    # 权限检查
    if user.role != "admin":
        for vid in videos:
            if vid.user_id != user.id:
                raise HTTPException(status_code=403, detail="无权下载其他用户的视频")
    
    # 创建临时 ZIP 文件
    with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
        zip_path = tmp.name
        files_added = 0
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for vid in videos:
                if not vid.result_url:
                    logger.warning(f"Video {vid.id} has no result_url")
                    continue
                
                # 构建视频文件的实际路径
                # result_url格式: /uploads/queue/video_xxx.mp4
                video_file_path = None
                
                # 方法1: 直接使用result_url构建路径
                if vid.result_url.startswith('/uploads/'):
                    video_file_path = Path('/app' + vid.result_url)
                
                # 方法2: 使用视频ID构建标准路径
                if not video_file_path or not video_file_path.exists():
                    video_file_path = Path(f'/app/uploads/queue/video_{vid.id}.mp4')
                
                # 检查文件是否存在
                if video_file_path and video_file_path.exists() and video_file_path.suffix == '.mp4':
                    # 使用.mp4扩展名,不使用filename字段(因为那是图片名)
                    base_name = vid.filename.rsplit('.', 1)[0] if vid.filename else f'video_{vid.id}'
                    arcname = f"{vid.id}_{base_name}.mp4"
                    
                    zipf.write(video_file_path, arcname=arcname)
                    files_added += 1
                    logger.info(f"Added video {vid.id} from {video_file_path} as {arcname}")
                else:
                    logger.warning(f"Video file not found for {vid.id}, tried path: {video_file_path}")
        
        logger.info(f"Created ZIP with {files_added} video files at {zip_path}")
        
        if files_added == 0:
            os.remove(zip_path)
            raise HTTPException(status_code=404, detail="未找到任何视频文件")
        
        # 返回文件下载响应,下载后自动删除临时文件
        background_tasks.add_task(os.remove, zip_path)
        return FileResponse(
            path=zip_path,
            media_type="application/zip",
            filename=f"videos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        )

# --- Queue APIs ---

@app.get("/api/v1/queue", response_model=List[QueueItemResponse])
def get_queue(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from sqlalchemy import case, func, extract
    
    # 管理员优先队列系统 - 使用公平调度算法
    # 管理员 (role='admin'): 基础权重 0 (最高优先级)
    # 普通用户: 基础权重 600
    # 
    # 公平调度规则：
    # 最终分数 = 基础权重 - 等待秒数
    # 等待时间越长，分数越低（越优先）
    # 这样即使是低优先级任务，等待足够久后也会被处理
    
    # 获取所有管理员的ID（基于角色判断，而非硬编码用户名）
    admin_ids = [u.id for u in db.query(User).filter(User.role == "admin").all()]
    
    # 基础优先级权重（数字越小越优先）
    # 管理员任务最高优先级，普通用户次之
    base_priority_weight = case(
        (VideoQueueItem.user_id.in_(admin_ids), 0),  # 管理员最高优先级
        else_=600                                      # 普通用户
    )
    
    # 计算等待时间（秒）
    wait_seconds = extract('epoch', func.now() - VideoQueueItem.created_at)
    
    # 最终调度分数 = 基础权重 - 等待秒数
    # 例如：
    # - P0任务刚创建: 0 - 0 = 0 (最优先)
    # - P2任务等待10分钟: 600 - 600 = 0 (与P0新任务同等优先)
    # - P2任务等待15分钟: 600 - 900 = -300 (比P1新任务还优先)
    fair_score = base_priority_weight - wait_seconds
    
    # 权限过滤：管理员可以看到所有任务，普通用户只能看到自己的任务
    base_query = db.query(VideoQueueItem).filter(VideoQueueItem.status != "archived")
    
    if user.role != "admin":
        # 普通用户只能看到自己的任务
        base_query = base_query.filter(VideoQueueItem.user_id == user.id)
    
    # 排序：按公平分数排序（分数越低越优先）
    return base_query.order_by(fair_score.asc()).all()

@app.post("/api/v1/queue")
async def add_to_queue(
    file: UploadFile = File(None),
    image_url: str = Form(None),
    prompt: str = Form(...),
    category: str = Form("other"),  # Product category
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not file and not image_url:
        raise HTTPException(status_code=400, detail="Either file or image_url must be provided")

    # Save file
    upload_dir = "/app/uploads/queue"
    os.makedirs(upload_dir, exist_ok=True)
    
    file_id = ""
    file_path = ""
    filename = ""

    if file:
        filename = file.filename
        file_id = f"{int(datetime.now().timestamp())}_{filename}"
        file_path = os.path.join(upload_dir, file_id)
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
    elif image_url:
        filename = "downloaded_image.png" # Default name or extract from URL?
        if "/" in image_url:
             filename = image_url.split("/")[-1].split("?")[0] # Simple extraction
             if not filename: filename = "image.png"
        
        file_id = f"{int(datetime.now().timestamp())}_{filename}"
        file_path = os.path.join(upload_dir, file_id)
        
        # Download image server-side (Proxy)
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.get(image_url, timeout=30.0)
                resp.raise_for_status()
                with open(file_path, "wb") as f:
                    f.write(resp.content)
        except Exception as e:
            logger.error(f"Failed to download image proxy: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to fetch image URL: {str(e)}")

    # Create DB record
    item = VideoQueueItem(
        id=str(int(datetime.now().timestamp() * 1000)), # Simple ID
        filename=filename,
        file_path=file_path,
        prompt=prompt,
        status="pending",
        user_id=user.id,
        category=category,  # Product category
        is_shared=user.default_share if user.default_share is not None else True
    )
    db.add(item)
    
    # Log activity for monitoring
    activity = UserActivity(
        user_id=user.id,
        action="video_gen_start",
        details=f"开始生成视频 | 类目: {category} | 提示词: {prompt[:50]}..."
    )
    db.add(activity)
    
    db.commit()
    db.refresh(item)
    
    # Update user status to show video generation in progress
    try:
        await connection_manager.update_user_activity(user.id, "正在生成视频")
    except Exception:
        pass  # WebSocket not required
    
    return item

@app.put("/api/v1/queue/{item_id}")
def update_queue_item(
    item_id: str,
    update: QueueItemUpdate,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    if update.status:
        item.status = update.status
    if update.result_url:
        item.result_url = update.result_url
    if update.error_msg:
        item.error_msg = update.error_msg
        
    db.commit()
    db.refresh(item)
    db.commit()
    db.refresh(item)
    return item

# --- Video Merge ---
class MergeRequest(BaseModel):
    video_ids: List[str]

@app.post("/api/v1/merge-videos")
async def merge_videos_endpoint(
    req: MergeRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    # 1. Validate and fetch videos
    videos = []
    for vid in req.video_ids:
        item = db.query(VideoQueueItem).filter(VideoQueueItem.id == vid).first()
        if not item:
            logger.warning(f"Video {vid} not found during merge")
            continue
        # Only allow successfully generated videos using their *result* path
        # Assuming result_url is http://.../uploads/queue/xxx.mp4
        # We need the local file path.
        if item.status != "done" or not item.result_url:
             logger.warning(f"Video {vid} not ready or missing result")
             continue
        videos.append(item)
    
    if not videos:
        raise HTTPException(status_code=400, detail="No valid videos selected for merging")

    # 2. Resolve local paths
    # Result URL: http://base/uploads/queue/video_123.mp4
    # File Path: /app/uploads/queue/video_123.mp4
    # We can infer filename from result_url
    local_paths = []
    for v in videos:
        # result_url likely ends with filename
        filename = v.result_url.split('/')[-1]
        local_path = os.path.join(QUEUE_DIR, filename)
        if os.path.exists(local_path):
            local_paths.append(local_path)
        else:
            logger.error(f"File missing on disk: {local_path}")

    if len(local_paths) < 2:
        raise HTTPException(status_code=400, detail="Need at least 2 valid video files to merge")

    # 3. Create ffmpeg list file
    import tempfile
    timestamp = int(datetime.now().timestamp())
    concat_list_path = os.path.join(QUEUE_DIR, f"concat_list_{timestamp}.txt")
    output_filename = f"merged_{timestamp}.mp4"
    output_path = os.path.join(QUEUE_DIR, output_filename)

    try:
        with open(concat_list_path, 'w') as f:
            for path in local_paths:
                f.write(f"file '{path}'\n")
        
        # 4. Run ffmpeg
        # ffmpeg -f concat -safe 0 -i list.txt -c copy output.mp4
        cmd = [
            "ffmpeg", "-f", "concat", "-safe", "0", 
            "-i", concat_list_path, 
            "-c", "copy", "-y", 
            output_path
        ]
        
        logger.info(f"Running merge command: {' '.join(cmd)}")
        process = await asyncio.create_subprocess_exec(
            *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE
        )
        stdout, stderr = await process.communicate()
        
        if process.returncode != 0:
            logger.error(f"FFmpeg merge failed: {stderr.decode()}")
            raise Exception("FFmpeg merge process failed")

        # 5. Create new Queue Item for the merged video
        new_item = VideoQueueItem(
            id=str(int(datetime.now().timestamp() * 1000)),
            filename=output_filename,
            file_path=output_path,
            prompt="Merged Video: " + ", ".join([v.filename for v in videos]),
            status="done", # Immediately done
            result_url=f"/uploads/queue/{output_filename}",
            user_id=user.id,
            is_shared=user.default_share if user.default_share is not None else True
        )
        db.add(new_item)
        db.commit()
        db.refresh(new_item)
        
        # Cleanup list file
        if os.path.exists(concat_list_path):
            os.remove(concat_list_path)
            
        return new_item

    except Exception as e:
        logger.error(f"Merge error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/v1/queue/{item_id}")
def delete_queue_item(
    item_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # 权限检查：只有管理员或任务所有者可以删除
    if user.role != "admin" and item.user_id != user.id:
        raise HTTPException(status_code=403, detail="您只能删除自己的任务")
    
    # Delete file
    if item.file_path and os.path.exists(item.file_path):
        try:
            os.remove(item.file_path)
        except Exception as e:
            logger.error(f"Failed to delete file {item.file_path}: {e}")
            
    db.delete(item)
    db.commit()
    return {"status": "deleted"}

@app.post("/api/v1/queue/{item_id}/retry")
async def retry_queue_item(
    item_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Retry a failed video generation task."""
    item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    if item.status != "error":
        raise HTTPException(status_code=400, detail="Only failed items can be retried")
    
    # Reset status, error message, and retry counter for fresh retry cycle
    item.status = "pending"
    item.error_msg = None
    item.retry_count = 0  # Reset retry counter for fresh retry cycle
    item.last_retry_at = None  # Clear cooldown timer
    db.commit()
    
    # Get config and trigger background task
    db_config_list = db.query(SystemConfig).all()
    config_dict = {conf.key: conf.value for conf in db_config_list}
    
    video_api_url = config_dict.get("video_api_url")
    video_api_key = config_dict.get("video_api_key")
    video_model_name = config_dict.get("video_model_name", "sora-video-portrait")
    
    if not video_api_url or not video_api_key:
        raise HTTPException(status_code=400, detail="Video API not configured")
    
    # Trigger background generation
    background_tasks.add_task(process_video_with_auto_retry, item_id, video_api_url, video_api_key, video_model_name)
    
    return {"status": "retrying", "item_id": item_id}

@app.delete("/api/v1/queue")
def clear_queue(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    query = db.query(VideoQueueItem)
    
    # 非管理员只能操作自己的任务
    if user.role != "admin":
        query = query.filter(VideoQueueItem.user_id == user.id)
    
    if status:
        query = query.filter(VideoQueueItem.status == status)
    
    items = query.all()
    count = len(items)
    
    for item in items:
        if (item.status in ["done", "archived"]) and item.result_url:
            # Archive completed/archived videos instead of deleting (preserve for gallery)
            if item.status == "done":
                item.status = "archived"
            # Skip already archived items (no action needed)
        else:
            # Delete non-completed items (pending, error, processing, etc.) and their source files
            if item.file_path and os.path.exists(item.file_path):
                try:
                    os.remove(item.file_path)
                except Exception:
                    pass
            db.delete(item)
    
    db.commit()
    return {"status": "cleared", "count": count}

# --- Background Task for Video Generation ---

def detect_api_error_cn(response_content: str) -> str:
    """智能检测API响应中的错误并返回对应的中文提示。
    
    Args:
        response_content: API返回的完整响应内容
        
    Returns:
        中文错误提示信息
    """
    content_lower = response_content.lower()
    
    # 内容策略违规相关
    if any(kw in content_lower for kw in ['content policy', 'policy violation', 'violat', 'inappropriate', 'nsfw', 'safety']):
        return "❌ 内容审核未通过：图片可能包含敏感、暴力或不适内容，请更换图片后重试"
    
    # 速率限制相关
    if any(kw in content_lower for kw in ['rate limit', 'too many requests', 'quota exceeded', '429']):
        return "⏳ API请求频率超限：系统将在稍后自动重试，请耐心等待"
    
    # 模型/服务不可用
    if any(kw in content_lower for kw in ['model not available', 'service unavailable', 'temporarily unavailable', '503']):
        return "🔧 视频生成服务暂时不可用，系统将自动重试"
    
    # 图片格式/尺寸问题
    if any(kw in content_lower for kw in ['invalid image', 'unsupported format', 'image too', 'resolution']):
        return "🖼️ 图片格式或尺寸不符合要求，请使用标准JPG/PNG格式（推荐9:16竖版）"
    
    # 认证/权限问题
    if any(kw in content_lower for kw in ['unauthorized', 'authentication', 'invalid key', '401', '403']):
        return "🔑 API认证失败，请联系管理员检查API密钥配置"
    
    # 请求参数问题
    if any(kw in content_lower for kw in ['invalid request', 'bad request', 'parameter', '400']):
        return "⚠️ 请求参数错误，请检查提示词或图片格式"
    
    # 服务器错误
    if any(kw in content_lower for kw in ['internal server error', '500', 'server error']):
        return "🔥 视频生成服务器内部错误，系统将自动重试"
    
    # 默认提示
    return f"❓ 视频URL解析失败：{response_content[:100]}..."

async def convert_sora_to_watermark_free(sora_url: str) -> str:
    """
    Watermark removal is handled by sora2api container with aitalk.works service.
    This function is a passthrough - sora2api should return watermark-free URLs directly.
    """
    # Watermark removal handled at sora2api level - return original URL
    return sora_url

async def process_video_background(item_id: str, video_api_url: str, video_api_key: str, video_model_name: str):
    logger.info(f"Background Task: Starting Video Generation for {item_id}")
    
    # Create a fresh DB session for the background task
    db = SessionLocal()
    try:
        item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
        if not item:
            logger.error(f"Background Task: Item {item_id} not found")
            return

        # Update status to processing immediately
        item.status = "processing"
        db.commit()

        # Read file and encode
        if not os.path.exists(item.file_path):
             item.status = "error"
             item.error_msg = "Source file not found"
             db.commit()
             return
             
        with open(item.file_path, "rb") as f:
            file_bytes = f.read()
            b64_img = base64.b64encode(file_bytes).decode('utf-8')

        target_url = video_api_url
        if not target_url.endswith("/chat/completions"):
             target_url = f"{target_url.rstrip('/')}/chat/completions"

        payload = {
             "model": video_model_name,
             "messages": [
                 {"role": "user", "content": [
                     {"type": "text", "text": f"Generate a video based on this image: {item.prompt}"},
                     {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_img}"}}
                 ]}
             ],
             "stream": True 
        }

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {video_api_key}"
        }
        
        # Enable stream mode as required by Sora2 API
        payload["stream"] = True 

        logger.info(f"Posting to {target_url} (Stream Mode)")

        # Use longer timeout for video generation (30 min total, 2 min connect)
        async with httpx.AsyncClient(timeout=httpx.Timeout(1800.0, connect=120.0)) as client:
            try:
                # Stream response - client timeout applies globally
                async with client.stream("POST", target_url, json=payload, headers=headers, timeout=900.0) as resp:
                    if resp.status_code != 200:
                        error_text = await resp.aread()
                        logger.error(f"Video API Error {resp.status_code}: {error_text.decode()}")
                        item.error_msg = f"API Error {resp.status_code}: {error_text.decode()[:200]}"
                        item.status = "error"
                    else:
                        # Process streaming response (SSE format)
                        full_content = ""
                        import json
                        
                        async for line in resp.aiter_lines():
                            if not line.strip():
                                continue
                            
                            # SSE format: "data: {json}"
                            if line.startswith("data: "):
                                data_str = line[6:]  # Remove "data: " prefix
                                
                                # Check for [DONE] signal
                                if data_str.strip() == "[DONE]":
                                    logger.info("Stream completed with [DONE] signal")
                                    break
                                
                                try:
                                    chunk_data = json.loads(data_str)
                                    # Extract delta content from streaming chunk
                                    choices = chunk_data.get("choices", [])
                                    if choices:
                                        delta = choices[0].get("delta", {})
                                        # Sora2 API uses "reasoning_content" instead of "content"
                                        # Try reasoning_content first, fallback to content
                                        content_chunk = delta.get("reasoning_content") or delta.get("content", "")
                                        if content_chunk:
                                            full_content += content_chunk
                                            logger.debug(f"Stream chunk received ({len(content_chunk)} chars): {content_chunk[:100]}...")
                                except json.JSONDecodeError as je:
                                    logger.warning(f"Failed to parse SSE chunk: {data_str[:100]}")
                                    continue
                        
                        logger.info(f"Stream complete, total content length: {len(full_content)}")
                        
                        # Parse final accumulated content
                        import re
                        # 1. HTML video tag: <video src="{url}" ...></video> (Grok API format)
                        video_tag_match = re.search(r'<video[^>]+src=["\']([^"\']+)["\'][^>]*>', full_content)
                        # 2. Markdown Image
                        img_match = re.search(r'!\[.*?\]\((.*?)\)', full_content)
                        # 3. Raw URL (http...) - Updated to exclude trailing ' and )
                        url_match = re.search(r'https?://[^\s<>"\'\\\)]+|data:image/[^\s<>"\'\\\)]+', full_content)
                        
                        found_url = None
                        if video_tag_match:
                            # HTML video tag format (Grok grok-imagine-0.9 API)
                            found_url = video_tag_match.group(1)
                            logger.info(f"Extracted video URL from HTML video tag: {found_url[:100]}...")
                        elif img_match:
                            found_url = img_match.group(1)
                        elif url_match:
                            found_url = url_match.group(0)
                            
                        if found_url:
                            # Cleanup trailing punctuation just in case
                            found_url = found_url.strip("'\".,)>")
                            # Paranoid cleanup for trailing quotes
                            found_url = found_url.split("'")[0]
                            found_url = found_url.split('"')[0]
                            
                            logger.info(f"Extracted video URL: {found_url[:100]}...")
                            
                            # Try to convert Sora URL to watermark-free version
                            found_url = await convert_sora_to_watermark_free(found_url)
                            
                            # Download video to local storage if it's an external URL
                            final_url = found_url
                            preview_url = None
                            
                            if found_url.startswith("http"):
                                local_filename = f"video_{item_id}.mp4"
                                local_path = f"/app/uploads/queue/{local_filename}"
                                download_success = False
                                
                                # Prepare headers - Grok/xAI videos need specific headers
                                download_headers = {
                                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                                    "Accept": "video/mp4,video/*,*/*",
                                    "Accept-Language": "en-US,en;q=0.9",
                                }
                                
                                # Add specific headers for Grok domains
                                if "grok.com" in found_url or "assets.grok.com" in found_url:
                                    download_headers["Referer"] = "https://grok.com/"
                                    download_headers["Origin"] = "https://grok.com"
                                    logger.info(f"Detected Grok assets URL, adding browser-like headers")
                                elif "grok.codeedu.de" in found_url or "codeedu.de" in found_url:
                                    download_headers["Referer"] = "https://grok.codeedu.de/"
                                    logger.info(f"Detected Grok cached URL, adding Referer header")
                                
                                # Try downloading with retries
                                max_download_retries = 3
                                for dl_attempt in range(max_download_retries):
                                    try:
                                        logger.info(f"Downloading video (attempt {dl_attempt + 1}/{max_download_retries}) from {found_url[:100]}...")
                                        async with httpx.AsyncClient() as download_client:
                                            video_resp = await download_client.get(
                                                found_url, 
                                                timeout=300.0,
                                                headers=download_headers,
                                                follow_redirects=True
                                            )
                                            if video_resp.status_code == 200:
                                                content_type = video_resp.headers.get("content-type", "")
                                                if "video" in content_type or len(video_resp.content) > 100000:
                                                    with open(local_path, "wb") as f:
                                                        f.write(video_resp.content)
                                                    final_url = f"/uploads/queue/{local_filename}"
                                                    download_success = True
                                                    logger.info(f"Video downloaded to local: {final_url} ({len(video_resp.content)} bytes)")
                                                    
                                                    # Generate thumbnail from first frame
                                                    thumb_filename = f"video_{item_id}_thumb.jpg"
                                                    thumb_path = f"/app/uploads/queue/{thumb_filename}"
                                                    try:
                                                        thumb_cmd = [
                                                            "ffmpeg", "-y",
                                                            "-i", local_path,
                                                            "-ss", "00:00:00.500",
                                                            "-vframes", "1",
                                                            "-q:v", "2",
                                                            thumb_path
                                                        ]
                                                        subprocess.run(thumb_cmd, check=True, capture_output=True)
                                                        preview_url = f"/uploads/queue/{thumb_filename}"
                                                    except Exception as thumb_err:
                                                        logger.warning(f"Failed to generate thumbnail: {thumb_err}")
                                                    break
                                                else:
                                                    logger.warning(f"Response doesn't look like video (content-type: {content_type}, size: {len(video_resp.content)})")
                                            elif video_resp.status_code == 403:
                                                logger.warning(f"Video download 403 Forbidden (attempt {dl_attempt + 1})")
                                                if dl_attempt < max_download_retries - 1:
                                                    await asyncio.sleep(2 * (dl_attempt + 1))
                                            else:
                                                logger.warning(f"Failed to download video: HTTP {video_resp.status_code}")
                                                break
                                    except Exception as dl_err:
                                        logger.warning(f"Video download attempt {dl_attempt + 1} failed: {dl_err}")
                                        if dl_attempt < max_download_retries - 1:
                                            await asyncio.sleep(2)
                                
                                if not download_success:
                                    # Keep remote URL as fallback, but log warning
                                    logger.warning(f"All download attempts failed, keeping remote URL (may expire): {found_url[:80]}...")
                            
                            item.result_url = final_url
                            if preview_url:
                                item.preview_url = preview_url
                            item.status = "done"
                            logger.info(f"Video Generated Successfully: {final_url}")
                            
                            # Log activity and update user status
                            try:
                                activity = UserActivity(
                                    user_id=item.user_id,
                                    action="video_gen_complete",
                                    details=f"视频生成完成 | 提示词: {item.prompt[:30]}..."
                                )
                                db.add(activity)
                                
                                # Update user status to idle and broadcast
                                await connection_manager.update_user_activity(item.user_id, "空闲")
                            except Exception as act_err:
                                logger.warning(f"Failed to log video completion: {act_err}")
                            
                            # Trigger video quality review (queued for sequential execution)
                            try:
                                from review_queue import enqueue_video_review
                                video_local_path = local_path if 'local_path' in dir() and os.path.exists(local_path) else None
                                if video_local_path:
                                    asyncio.create_task(
                                        enqueue_video_review(
                                            video_id=item_id,
                                            video_path=video_local_path,
                                            video_prompt=item.prompt,
                                            db_session=SessionLocal,
                                            VideoQueueItem_model=VideoQueueItem
                                        )
                                    )
                                    logger.info(f"Video review task queued for {item_id}")
                            except Exception as review_err:
                                logger.warning(f"Failed to trigger video review: {review_err}")
                        else:
                            logger.warning(f"No URL found in video response: {full_content[:200]}")
                            # 智能错误检测 - 将API返回的错误翻译为中文提示
                            error_msg_cn = detect_api_error_cn(full_content)
                            item.error_msg = error_msg_cn
                            item.status = "error"
                              
            except httpx.TimeoutException:
                logger.error("Video Generation Timeout (900s / 15 minutes)")
                item.error_msg = "Video Generation Timed Out (超过15分钟)"
                item.status = "error"
            except Exception as e:
                logger.error(f"Video Client Error: {e}")
                item.error_msg = f"Client Error: {str(e)}"
                item.status = "error"

        db.commit()
    except Exception as e:
        logger.error(f"Background Task Critical Error: {e}")
        db.rollback()
    finally:
        db.close()


async def process_video_with_auto_retry(item_id: str, video_api_url: str, video_api_key: str, video_model_name: str, skip_concurrency_check: bool = False):
    """Wrapper function that adds automatic retry logic to video generation.
    
    Retries up to 3 times for retryable errors. Timeout errors are NOT retried.
    Retry count is persisted to database to prevent loss on restart.
    
    IMPORTANT: This function now enforces global video concurrency limits.
    
    Args:
        skip_concurrency_check: If True, skip acquiring global slot (used when caller already holds a slot, e.g., Story Fission)
    """
    MAX_AUTO_RETRIES = 3
    RETRY_BASE_DELAY = 60  # 增加到 60 秒基础延迟
    COOLDOWN_SECONDS = 120  # 同一任务两次尝试之间的最小间隔
    SLOT_ACQUIRE_TIMEOUT = 600  # 获取槽位的最大等待时间（秒）
    SLOT_RETRY_INTERVAL = 5  # 等待槽位时的重试间隔（秒）
    
    # 从数据库获取当前重试计数
    db = SessionLocal()
    try:
        item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
        if not item:
            logger.error(f"Video {item_id}: Item not found")
            return
        
        # 检查冷却期：如果最近处理过，跳过
        if item.last_retry_at:
            seconds_since_last = (get_china_now() - item.last_retry_at).total_seconds()
            if seconds_since_last < COOLDOWN_SECONDS and item.retry_count > 0:
                logger.warning(f"Video {item_id}: Still in cooldown period ({seconds_since_last:.0f}s < {COOLDOWN_SECONDS}s), skipping")
                return
        
        start_attempt = item.retry_count + 1  # 从持久化的计数继续
        
        # 如果已经超过最大重试次数，直接标记为失败并返回
        if start_attempt > MAX_AUTO_RETRIES:
            logger.error(f"Video {item_id}: Already exceeded max retries ({item.retry_count}/{MAX_AUTO_RETRIES}), marking as failed")
            if item.status != "error":
                item.status = "error"
                item.error_msg = item.error_msg or f"重试次数已达上限 ({MAX_AUTO_RETRIES}次)"
                db.commit()
            return
    finally:
        db.close()
    
    # 是否需要获取并发槽位
    limiter = None
    slot_acquired = False
    
    if not skip_concurrency_check:
        # 获取全局并发限制器
        limiter = await get_concurrency_limiter(get_concurrency_config)
        
        # 尝试获取全局视频生成槽位
        wait_start = asyncio.get_event_loop().time()
        
        while not slot_acquired:
            elapsed = asyncio.get_event_loop().time() - wait_start
            if elapsed > SLOT_ACQUIRE_TIMEOUT:
                logger.error(f"Video {item_id}: Failed to acquire slot after {SLOT_ACQUIRE_TIMEOUT}s timeout")
                # 更新状态为等待中
                db = SessionLocal()
                try:
                    item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                    if item and item.status == "processing":
                        item.status = "pending"
                        item.error_msg = "队列繁忙，等待重试"
                        db.commit()
                finally:
                    db.close()
                return
            
            slot_acquired = await limiter.acquire_global("video_gen", timeout=30)
            if not slot_acquired:
                logger.info(f"Video {item_id}: Waiting for slot (elapsed: {elapsed:.0f}s)")
                await asyncio.sleep(SLOT_RETRY_INTERVAL)
        
        logger.info(f"Video {item_id}: Acquired global video slot")
    else:
        logger.info(f"Video {item_id}: Skipping concurrency check (caller holds slot)")
    
    try:
        for attempt in range(start_attempt, MAX_AUTO_RETRIES + 1):
            logger.info(f"Video {item_id}: Starting attempt {attempt}/{MAX_AUTO_RETRIES}")
            
            # 更新重试计数和时间戳
            db = SessionLocal()
            try:
                item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                if item:
                    item.retry_count = attempt
                    item.last_retry_at = get_china_now()
                    db.commit()
            finally:
                db.close()
            
            # Apply global throttling before each attempt
            await throttle_request()
            
            # Call the original function
            await process_video_background(item_id, video_api_url, video_api_key, video_model_name)
            
            # Check the result - use short-lived DB session to avoid connection exhaustion
            should_retry = False
            retry_delay = 0
            
            db = SessionLocal()
            try:
                item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                if not item:
                    logger.error(f"Video {item_id}: Item not found after processing")
                    return
                
                if item.status == "done":
                    logger.info(f"Video {item_id}: Completed successfully on attempt {attempt}")
                    # 成功后重置重试计数
                    item.retry_count = 0
                    db.commit()
                    return
                
                if item.status == "error":
                    error_msg = item.error_msg or ""
                    
                    # Timeout errors should NOT be auto-retried
                    if "Timed Out" in error_msg or "超时" in error_msg:
                        logger.warning(f"Video {item_id}: Timeout error, not retrying")
                        return
                    
                    # Other errors can be retried
                    if attempt < MAX_AUTO_RETRIES:
                        # 更保守的重试策略：基础延迟 + 指数退避 + 随机波动
                        retry_delay = RETRY_BASE_DELAY * (1.5 ** (attempt - 1)) + random.uniform(10, 30)
                        logger.info(f"Video {item_id}: Error '{error_msg[:50]}...', retrying in {retry_delay:.1f}s (attempt {attempt}/{MAX_AUTO_RETRIES})")
                        
                        # Reset status to pending for next attempt
                        item.status = "pending"
                        item.error_msg = None
                        db.commit()
                        should_retry = True
                    else:
                        # 最后一次重试也失败了，确保状态为 error 并保留错误信息
                        logger.error(f"Video {item_id}: All {MAX_AUTO_RETRIES} attempts failed")
                        item.status = "error"
                        item.error_msg = f"重试 {MAX_AUTO_RETRIES} 次后仍失败: {error_msg[:100]}"
                        db.commit()
                        return
            finally:
                db.close()  # CRITICAL: Close connection BEFORE sleep to avoid pool exhaustion
            
            # Sleep AFTER closing DB connection to free up pool resources
            if should_retry:
                await asyncio.sleep(retry_delay)
    finally:
        # 只有在实际获取了槽位时才释放
        if slot_acquired and limiter:
            await limiter.release_global("video_gen")
            logger.info(f"Video {item_id}: Released global video slot")


@app.post("/api/v1/queue/{item_id}/generate")
async def generate_queue_item_endpoint(
    item_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    # 1. Get queue item
    item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # 2. Prevent duplicate trigger - if already processing, reject the request
    if item.status == "processing":
        raise HTTPException(status_code=409, detail="Task is already being processed")
    
    # 2.5 冷却期检查：防止短时间内重复触发
    COOLDOWN_SECONDS = 60  # 1 分钟冷却期
    if item.last_retry_at and item.status == "pending":
        seconds_since_last = (get_china_now() - item.last_retry_at).total_seconds()
        if seconds_since_last < COOLDOWN_SECONDS:
            remaining = int(COOLDOWN_SECONDS - seconds_since_last)
            raise HTTPException(
                status_code=429, 
                detail=f"任务正在处理中，请等待 {remaining} 秒后再试"
            )
    
    # 3. Get Config
    db_config_list = db.query(SystemConfig).all()
    config_dict = {conf.key: conf.value for conf in db_config_list}
    
    video_api_url = config_dict.get("video_api_url")
    video_api_key = config_dict.get("video_api_key")
    video_model_name = config_dict.get("video_model_name", "sora-video-portrait")

    if not video_api_url or not video_api_key:
         item.status = "error"
         item.error_msg = "Missing Video API Config"
         db.commit()
         raise HTTPException(status_code=400, detail="Missing Video API Config")

    # 3. Start Processing (Set status and trigger background task)
    item.status = "processing"
    
    # 4. Log activity for monitoring
    activity = UserActivity(
        user_id=item.user_id,
        action="video_gen_processing",
        details=f"视频生成中 | 提示词: {item.prompt[:50] if item.prompt else '无'}..."
    )
    db.add(activity)
    db.commit()
    
    # 5. Update user status via WebSocket
    try:
        await connection_manager.update_user_activity(item.user_id, "正在生成视频")
    except Exception:
        pass  # WebSocket not required

    logger.info(f"Triggering Background Generation for {item_id}")
    background_tasks.add_task(process_video_with_auto_retry, item_id, video_api_url, video_api_key, video_model_name)

    return {"status": "processing", "message": "Video generation started in background"}


# --- Background Zombie Task Recovery ---
# Detects "zombie" tasks stuck in processing state and recovers them

ZOMBIE_DETECTION_INTERVAL = 60  # Check every 60 seconds
ZOMBIE_TIMEOUT_SECONDS = 300  # Task is considered zombie after 5 minutes in processing

async def zombie_task_recovery():
    """
    Background task that detects and recovers zombie tasks.
    A zombie task is one that has been in 'processing' status for too long
    without completing (e.g., due to container restart or crash).
    """
    while True:
        try:
            await asyncio.sleep(ZOMBIE_DETECTION_INTERVAL)
            
            db = SessionLocal()
            try:
                now = get_china_now()
                cutoff_time = now - timedelta(seconds=ZOMBIE_TIMEOUT_SECONDS)
                
                # Find processing tasks that haven't been updated recently
                zombie_tasks = db.query(VideoQueueItem).filter(
                    VideoQueueItem.status == "processing",
                    VideoQueueItem.last_retry_at < cutoff_time
                ).all()
                
                if not zombie_tasks:
                    continue
                
                recovered_count = 0
                failed_count = 0
                MAX_AUTO_RETRIES = 3
                
                for task in zombie_tasks:
                    task_age = (now - task.last_retry_at).total_seconds() if task.last_retry_at else 0
                    
                    # Check if task has exceeded max retries
                    if task.retry_count >= MAX_AUTO_RETRIES:
                        # Mark as error instead of retrying forever
                        task.status = "error"
                        task.error_msg = f"任务超时且已达最大重试次数 ({task.retry_count}/{MAX_AUTO_RETRIES})"
                        failed_count += 1
                        logger.warning(f"[ZombieRecovery] Task {task.id} exceeded max retries, marked as error")
                    else:
                        # Reset to pending for retry
                        task.status = "pending"
                        # Don't increment retry_count here - it will be incremented when task is picked up
                        # Don't set last_retry_at to avoid cooldown - the task was interrupted, not failed
                        recovered_count += 1
                        logger.info(f"[ZombieRecovery] Task {task.id} recovered to pending (age: {task_age:.0f}s, retries: {task.retry_count})")
                
                db.commit()
                
                if recovered_count > 0 or failed_count > 0:
                    logger.info(f"[ZombieRecovery] Cycle complete: {recovered_count} recovered, {failed_count} marked as error")
                    
            finally:
                db.close()
                
        except Exception as e:
            logger.error(f"[ZombieRecovery] Error in zombie detection: {e}")
            await asyncio.sleep(10)  # Brief pause on error before retrying


# --- Background Cleanup Task ---
# Legacy env var for backward compatibility, will be overridden by DB config
CLEANUP_HOURS = int(os.getenv("CLEANUP_HOURS", "168"))  # Default 7 days (168 hours)

async def cleanup_task():
    while True:
        try:
            db = SessionLocal()
            
            # Read retention days from database config (0 = permanent, skip cleanup)
            retention_config = db.query(SystemConfig).filter(SystemConfig.key == "cache_retention_days").first()
            retention_days = int(retention_config.value) if retention_config else 7
            
            if retention_days == 0:
                logger.info("Cleanup task: cache_retention_days=0 (permanent), skipping cleanup")
                db.close()
                await asyncio.sleep(3600)
                continue
            
            logger.info(f"Running cleanup task... (retention: {retention_days} days)")
            cutoff = datetime.now() - timedelta(days=retention_days)
            
            # Find old items, EXCLUDE completed/archived videos and merged story videos
            old_items = db.query(VideoQueueItem).filter(
                VideoQueueItem.created_at < cutoff,
                VideoQueueItem.status.notin_(["done", "archived"]),  # Keep completed videos
                ~VideoQueueItem.filename.like("story_chain%"),  # Keep merged chain videos
                ~VideoQueueItem.filename.like("story_fission%")  # Keep merged fission videos
            ).all()
            
            for item in old_items:
                logger.info(f"Cleaning up old item: {item.id} ({item.filename})")
                if item.file_path and os.path.exists(item.file_path):
                    try:
                        os.remove(item.file_path)
                    except Exception as e:
                        logger.error(f"Failed to delete file {item.file_path}: {e}")
                db.delete(item)
            
            db.commit()
            db.close()
            logger.info(f"Cleanup task completed. Removed {len(old_items)} old items.")
        except Exception as e:
            logger.error(f"Cleanup task failed: {e}")
            
        # Wait for 1 hour
        await asyncio.sleep(3600)

@app.on_event("startup")
async def start_background_tasks():
    asyncio.create_task(cleanup_task())
    asyncio.create_task(zombie_task_recovery())

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# --- Story Chain Implementation ---

class StoryChainRequest(BaseModel):
    initial_image_url: str # Base64 or URL
    shots: List[dict] # From analyze step
    api_url: Optional[str] = None
    api_key: Optional[str] = None
    model_name: Optional[str] = None
    visual_style: Optional[str] = None  # Visual style ID
    visual_style_prompt: Optional[str] = None  # Visual style prompt text
    camera_movement: Optional[str] = None  # Camera movement ID
    camera_movement_prompt: Optional[str] = None  # Camera movement prompt text
    category: str = "other"  # Product category for gallery/video classification


STORY_CHAIN_STATUS = {}

def extract_last_frame(video_path: str) -> str:
    """Extracts last frame from video and saves as temp image. Returns path."""
    try:
        # Resolve path
        if video_path.startswith("/uploads"):
            video_path = f"/app{video_path}"
            
        if not os.path.exists(video_path):
            raise Exception(f"Video file not found: {video_path}")
        
        output_path = video_path.replace(".mp4", "_last.jpg")
        # Ensure we overwrite
        if os.path.exists(output_path):
            os.remove(output_path)
            
        # ffmpeg command to extract last frame
        # Use -0.5 instead of -0.1 to avoid seeking errors at very end of file
        cmd = [
            "ffmpeg", "-y", 
            "-sseof", "-0.5", 
            "-i", video_path, 
            "-update", "1", 
            "-q:v", "2", 
            output_path
        ]
        # Allow stderr to show in logs for debugging
        subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL)
        
        if not os.path.exists(output_path):
             raise Exception("Frame extraction failed: Output not created")
             
        return output_path
    except Exception as e:
        logging.error(f"Extract Frame Error: {e}")
        return None

async def process_story_chain(chain_id: str, req: StoryChainRequest, user_id: int):
    status = {"status": "processing", "current_shot": 0, "total_shots": len(req.shots), "video_ids": [], "error": None, "merged_video_url": None}
    STORY_CHAIN_STATUS[chain_id] = status
    
    # Config resolution
    db = SessionLocal()
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    db.close()
    
    final_api_url = req.api_url or config_dict.get("video_api_url", "")
    final_api_key = req.api_key or config_dict.get("video_api_key", "")
    final_model = req.model_name or config_dict.get("video_model_name", "sora-video-portrait")
    
    # Get image generation API config for pre-processing
    image_api_url = config_dict.get("api_url", "")
    image_api_key = config_dict.get("api_key", "")
    image_model = config_dict.get("image_model_portrait", "gemini-2.5-flash-preview-05-20")  # Portrait image model
    analysis_model = config_dict.get("analysis_model_name", "gemini-3-pro-preview")  # 尾帧分析模型
    
    # 存储从尾帧分析生成的下一镜头 prompt（用于替换原始脚本）
    next_shot_prompt_override = None

    try:
        current_image_source = req.initial_image_url
        
        # Ensure queue dir exists
        os.makedirs("/app/uploads/queue", exist_ok=True)
        
        # === STEP 0: Pre-process original image with image generation model ===
        # This "washes" the original image to create a stylized version before video generation
        logging.info(f"Chain {chain_id}: Preprocessing check - image_api_url={bool(image_api_url)}, image_api_key={bool(image_api_key)}, shots={len(req.shots)}, visual_style={req.visual_style_prompt[:50] if req.visual_style_prompt else 'None'}...")
        
        if image_api_url and image_api_key and len(req.shots) > 0:
            logging.info(f"Chain {chain_id}: Pre-processing original image with {image_model}")
            status["current_shot"] = 0
            status["status"] = "preprocessing"
            
            try:
                # Get the first shot's prompt and heroSubject for context
                first_shot = req.shots[0]
                hero_subject = first_shot.get("heroSubject", "")
                first_prompt = first_shot.get("prompt", "")
                
                # Prepare the original image as base64
                if current_image_source.startswith("data:"):
                    header, encoded = current_image_source.split(",", 1)
                    original_b64 = encoded
                elif current_image_source.startswith("/"):
                    with open(current_image_source, "rb") as f:
                        original_b64 = base64.b64encode(f.read()).decode('utf-8')
                else:
                    async with httpx.AsyncClient() as client:
                        resp = await client.get(current_image_source, timeout=30)
                        original_b64 = base64.b64encode(resp.content).decode('utf-8')
                
                # Construct preprocessing prompt with visual style
                visual_style = req.visual_style_prompt or "Filmic realism, natural lighting, soft bokeh, 35mm lens, muted colors, subtle grain."
                
                preprocess_prompt = (
                    f"Role: Cinematic frame artist establishing the visual style.\n"
                    f"Goal: Transform this product image into a stylized opening frame.\n"
                    f"CRITICAL Main Subject: {hero_subject}\n"
                    f"Scene Setup: {first_prompt}\n"
                    f"Visual Style: {visual_style}\n"
                    f"Constraints: no text, 9:16 portrait aspect ratio, high fidelity, maintain product identity.\n"
                    f"Output: A stylized version of the product image with the specified visual style, ready for video generation."
                )
                
                # Call image generation API
                async with httpx.AsyncClient(timeout=120) as client:
                    result = await call_openai_compatible_api(
                        client, image_api_url, image_api_key,
                        original_b64,  # Input image
                        original_b64,  # Reference image (same for first frame)
                        "Preprocessing",
                        preprocess_prompt,
                        image_model
                    )
                    
                    if result and result.image_base64:
                        # Save preprocessed image
                        preprocessed_filename = f"chain_{chain_id}_preprocessed.jpg"
                        preprocessed_path = os.path.join("/app/uploads/queue", preprocessed_filename)
                        
                        preprocessed_data = base64.b64decode(result.image_base64)
                        with open(preprocessed_path, "wb") as f:
                            f.write(preprocessed_data)
                        
                        # Update image source to use preprocessed image
                        current_image_source = preprocessed_path
                        logging.info(f"Chain {chain_id}: Preprocessing complete, saved to {preprocessed_path}")
                    else:
                        logging.warning(f"Chain {chain_id}: Preprocessing returned no image, using original")
                        
            except Exception as preprocess_err:
                logging.warning(f"Chain {chain_id}: Preprocessing failed ({preprocess_err}), using original image")
            
            status["status"] = "processing"
            
            # === STEP 0.5: 分析预处理后的首帧，生成第一镜头的 prompt ===
            if image_api_url and image_api_key and len(req.shots) > 0:
                try:
                    logging.info(f"Chain {chain_id}: Analyzing preprocessed first frame...")
                    
                    # 读取预处理后的图片（或原始图片如果预处理失败）
                    if current_image_source.startswith("data:"):
                        header, encoded = current_image_source.split(",", 1)
                        first_frame_b64 = encoded
                    elif current_image_source.startswith("/"):
                        with open(current_image_source, "rb") as f:
                            first_frame_b64 = base64.b64encode(f.read()).decode('utf-8')
                    else:
                        async with httpx.AsyncClient() as client:
                            resp = await client.get(current_image_source, timeout=30)
                            first_frame_b64 = base64.b64encode(resp.content).decode('utf-8')
                    
                    # 获取第一镜头的场景信息
                    first_shot = req.shots[0]
                    first_description = first_shot.get("description", "")
                    first_story = first_shot.get("shotStory", "")
                    
                    # 运镜选项列表
                    camera_movements = [
                        "slow push-in (推进)",
                        "gentle pull-back (拉远)",
                        "smooth pan left/right (横摇)",
                        "subtle tilt up/down (俯仰)",
                        "orbit around subject (环绕)",
                        "static with subject motion (静态+主体运动)"
                    ]
                    
                    # 获取视觉风格
                    visual_style = req.visual_style_prompt or "Filmic realism, natural lighting, soft bokeh, 35mm lens, muted colors, subtle grain."
                    
                    # 获取运镜风格（用户指定或自动选择）
                    camera_instruction = ""
                    if req.camera_movement_prompt:
                        camera_instruction = f"   - CAMERA MOVEMENT: {req.camera_movement_prompt}\n"
                    else:
                        camera_instruction = f"   - CAMERA MOVEMENT: Choose one from: {', '.join(camera_movements)}\n"
                    
                    # 动态动作关键词列表
                    dynamic_actions = [
                        "rotating slowly", "floating gently", "rising/falling",
                        "rippling", "shimmering", "flowing", "drifting",
                        "particles floating in the air", "light beams shifting",
                        "steam/mist moving", "leaves drifting", "water flowing"
                    ]
                    
                    first_frame_analysis_prompt = (
                        f"You are a professional video director. Analyze this image and generate a DYNAMIC video prompt for the OPENING SHOT.\n\n"
                        f"SCENE CONTEXT:\n"
                        f"- Description: {first_description}\n"
                        f"- Story: {first_story}\n"
                        f"- VISUAL STYLE: {visual_style}\n\n"
                        f"⚠️ MANDATORY PRODUCT VISIBILITY RULES:\n"
                        f"1. The ENTIRE product must be FULLY VISIBLE in frame at ALL times\n"
                        f"2. NO extreme close-ups or macro shots that crop the product\n"
                        f"3. Use MEDIUM or WIDE shots only - keep product occupying 30-70% of frame\n"
                        f"4. Camera movement must be GENTLE - no aggressive push-ins\n"
                        f"5. Product must remain the HERO of every frame\n\n"
                        f"DYNAMIC REQUIREMENTS:\n"
                        f"1. Include REAL MOTION - NOT just static image with camera movement\n"
                        f"2. DYNAMIC ELEMENTS: {', '.join(dynamic_actions[:5])}\n"
                        f"3. Product can have subtle movements but must stay fully visible\n\n"
                        f"GENERATE A PROMPT WITH:\n"
                        f"   - Subject: Full product description, ALWAYS FULLY VISIBLE in frame\n"
                        f"   - Framing: Medium or wide shot, product centered, complete view\n"
                        f"{camera_instruction}"
                        f"   - Motion: Environmental dynamics (particles, light shifts) around the product\n"
                        f"   - Style: {visual_style}\n\n"
                        f"FORBIDDEN: extreme close-up, macro, tight crop, partial product view\n"
                        f"OUTPUT: A single paragraph prompt. Product must be FULLY VISIBLE throughout."
                    )
                    
                    async with httpx.AsyncClient(timeout=60) as client:
                        target_url = image_api_url
                        if not target_url.endswith("/chat/completions"):
                            target_url = f"{target_url.rstrip('/')}/chat/completions"
                        
                        payload = {
                            "model": analysis_model,
                            "messages": [
                                {"role": "user", "content": [
                                    {"type": "text", "text": first_frame_analysis_prompt},
                                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{first_frame_b64}"}}
                                ]}
                            ]
                        }
                        headers = {
                            "Content-Type": "application/json",
                            "Authorization": f"Bearer {image_api_key}"
                        }
                        
                        resp = await client.post(target_url, json=payload, headers=headers, timeout=60.0)
                        if resp.status_code == 200:
                            data = resp.json()
                            new_prompt = data.get("choices", [])[0].get("message", {}).get("content", "").strip()
                            if new_prompt:
                                # 设置 next_shot_prompt_override，这样第一镜头也会使用分析生成的 prompt
                                next_shot_prompt_override = new_prompt
                                logging.info(f"Chain {chain_id}: Generated first shot prompt: {new_prompt[:100]}...")
                        else:
                            logging.warning(f"Chain {chain_id}: First frame analysis API error: {resp.status_code}")
                except Exception as first_analysis_err:
                    logging.warning(f"Chain {chain_id}: First frame analysis failed: {first_analysis_err}, using original prompt")
        
        # Extract heroSubject from first shot for product consistency
        hero_subject = ""
        if len(req.shots) > 0:
            hero_subject = req.shots[0].get("heroSubject", "")
        
        # Log hero subject for debugging
        logging.info(f"Chain {chain_id}: heroSubject = '{hero_subject[:100]}...' " if hero_subject else f"Chain {chain_id}: heroSubject is empty - product consistency may be affected")
        
        for i, shot in enumerate(req.shots):
            shot_num = i + 1
            status["current_shot"] = shot_num
            logging.info(f"Chain {chain_id}: Starting Shot {shot_num}")
            
            # Create Queue Item
            db = SessionLocal()
            try:
                # Build prompt with strong product consistency enforcement
                base_prompt = shot.get("prompt", "")
                
                # Get product description - try current shot's heroSubject first, then global hero_subject, then description
                product_desc = shot.get("heroSubject", "") or hero_subject
                if not product_desc:
                    # Try to extract from shot description or use generic fallback
                    product_desc = shot.get("description", "the product shown in the image")
                
                # 如果上一镜头的尾帧分析生成了新的 prompt，则使用它替换原始脚本
                # 这确保产品一致性和视觉延续性
                if next_shot_prompt_override:
                    prompt_text = next_shot_prompt_override
                    logging.info(f"Chain {chain_id}: Using override prompt for shot {shot_num}")
                    # 重置，避免影响后续镜头
                    next_shot_prompt_override = None
                else:
                    # 第一个镜头或分析失败时使用原始 prompt
                    prompt_text = base_prompt
                
                queue_filename = f"chain_{chain_id}_shot_{shot_num}_input.jpg"
                queue_file_path = os.path.join("/app/uploads/queue", queue_filename)
                
                # Check if current_image_source is a local path (from extraction) or URL
                if current_image_source.startswith("/"):
                    # Local path
                    import shutil
                    shutil.copy(current_image_source, queue_file_path)
                elif current_image_source.startswith("data:"):
                     # Base64
                     header, encoded = current_image_source.split(",", 1)
                     data = base64.b64decode(encoded)
                     with open(queue_file_path, "wb") as f:
                         f.write(data)
                else:
                    # URL
                     async with httpx.AsyncClient() as client:
                         resp = await client.get(current_image_source, timeout=30)
                         with open(queue_file_path, "wb") as f:
                             f.write(resp.content)
                                 
                # 获取用户的分享设置
                share_user = db.query(User).filter(User.id == user_id).first()
                user_default_share = share_user.default_share if share_user and share_user.default_share is not None else True
                
                new_item = VideoQueueItem(
                    id=str(int(uuid.uuid4().int))[:18],
                    filename=queue_filename,
                    file_path=queue_file_path,
                    prompt=prompt_text,
                    status="pending",
                    user_id=user_id,
                    category=req.category,  # Use user-selected product category
                    is_shared=user_default_share
                )
                db.add(new_item)
                db.commit()
                db.refresh(new_item)
                item_id = new_item.id
                status["video_ids"].append(item_id)
            finally:
                db.close()
            
            # Trigger Generation with retry mechanism (max 3 attempts)
            # With progressive delay to avoid CF rate limiting
            max_retries = 3
            VIDEO_RETRY_BASE_DELAY = 8  # Base delay for video retries (seconds)
            
            for attempt in range(1, max_retries + 1):
                logging.info(f"Chain {chain_id}: Shot {shot_num} attempt {attempt}/{max_retries}")
                
                # Reset status for retry
                if attempt > 1:
                    # Progressive delay with jitter: 8s, 16s, 32s (+ random 2-8s)
                    retry_delay = VIDEO_RETRY_BASE_DELAY * (2 ** (attempt - 2)) + random.uniform(2, 8)
                    logging.info(f"Chain {chain_id}: Waiting {retry_delay:.1f}s before retry (anti-CF)")
                    await asyncio.sleep(retry_delay)
                    
                    db = SessionLocal()
                    retry_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                    if retry_item:
                        retry_item.status = "pending"
                        retry_item.error_msg = None
                        db.commit()
                    db.close()
                
                # Apply global throttling before video API call
                await throttle_request()
                
                await process_video_with_auto_retry(item_id, final_api_url, final_api_key, final_model)
                
                # Check result
                db = SessionLocal()
                completed_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                
                if completed_item.status == "done":
                    logging.info(f"Chain {chain_id}: Shot {shot_num} completed successfully on attempt {attempt}")
                    result_url = completed_item.result_url
                    db.close()
                    break
                else:
                    error_msg = completed_item.error_msg or "Unknown error"
                    logging.warning(f"Chain {chain_id}: Shot {shot_num} attempt {attempt} failed: {error_msg}")
                    db.close()
                    
                    if attempt == max_retries:
                        logging.error(f"Chain {chain_id}: Shot {shot_num} failed after {max_retries} attempts")
                        status["status"] = "failed"
                        status["error"] = f"Shot {shot_num} failed after {max_retries} attempts: {error_msg}"
                        return
            
            # Get result_url after successful generation
            db = SessionLocal()
            completed_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
            result_url = completed_item.result_url
            db.close()
            
            # Local path resolution and download if needed
            # Wrap in a retry loop for download failures
            download_success = False
            download_retries = 3
            
            for download_attempt in range(1, download_retries + 1):
                if result_url.startswith("/uploads"):
                    local_video_path = f"/app{result_url}"
                    download_success = True
                    break
                elif result_url.startswith("http"):
                    # Convert to watermark-free URL if it's a Sora video
                    try:
                        converted_url = await convert_sora_to_watermark_free(result_url)
                        if converted_url and converted_url != result_url:
                            logging.info(f"Chain {chain_id}: Converted to watermark-free URL for shot {shot_num}")
                            result_url = converted_url
                    except Exception as wm_err:
                        logging.warning(f"Chain {chain_id}: Watermark removal failed for shot {shot_num}: {wm_err}")
                    
                    # Download remote video
                    try:
                        async with httpx.AsyncClient() as client:
                            resp = await client.get(result_url, timeout=300)
                            if resp.status_code == 200:
                                remote_filename = f"chain_{chain_id}_shot_{shot_num}_result.mp4"
                                local_video_path = f"/app/uploads/queue/{remote_filename}"
                                with open(local_video_path, "wb") as f:
                                    f.write(resp.content)
                                download_success = True
                                break
                            else:
                                logging.warning(f"Chain {chain_id}: Download failed for shot {shot_num} (attempt {download_attempt}/{download_retries}): HTTP {resp.status_code}")
                    except Exception as dl_err:
                        logging.warning(f"Chain {chain_id}: Download error for shot {shot_num} (attempt {download_attempt}/{download_retries}): {dl_err}")
                    
                    # If download failed and we have retries left, regenerate the video
                    if download_attempt < download_retries:
                        logging.info(f"Chain {chain_id}: Regenerating shot {shot_num} due to download failure...")
                        
                        # Reset the queue item status for regeneration
                        db = SessionLocal()
                        retry_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                        if retry_item:
                            retry_item.status = "pending"
                            retry_item.result_url = None
                            retry_item.error_msg = None
                            db.commit()
                        db.close()
                        
                        # Progressive delay for download retry regeneration
                        regen_delay = 10 + random.uniform(3, 10) * download_attempt
                        logging.info(f"Chain {chain_id}: Waiting {regen_delay:.1f}s before regeneration (anti-CF)")
                        await asyncio.sleep(regen_delay)
                        
                        # Apply throttling before regeneration
                        await throttle_request()
                        
                        # Regenerate video
                        await process_video_with_auto_retry(item_id, final_api_url, final_api_key, final_model)
                        
                        # Get new result_url
                        db = SessionLocal()
                        completed_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
                        if completed_item and completed_item.status == "done":
                            result_url = completed_item.result_url
                        else:
                            logging.warning(f"Chain {chain_id}: Regeneration failed for shot {shot_num}")
                            result_url = ""
                        db.close()
                else:
                    local_video_path = result_url
                    download_success = True
                    break
            
            if not download_success:
                status["status"] = "failed"
                status["error"] = f"Failed to download video for shot {shot_num} after {download_retries} attempts"
                return
            
            if i < len(req.shots) - 1:
                # Extract last frame for next shot
                extracted_path = extract_last_frame(local_video_path)
                if not extracted_path:
                    status["status"] = "failed"
                    status["error"] = f"Failed to extract frame from Shot {shot_num}"
                    return
                current_image_source = extracted_path
                
                # === 尾帧分析：保持产品一致性和延续性 ===
                next_shot_prompt_override = None
                if image_api_url and image_api_key:
                    try:
                        logging.info(f"Chain {chain_id}: Analyzing last frame for shot {shot_num + 1} continuity...")
                        with open(extracted_path, "rb") as f:
                            frame_b64 = base64.b64encode(f.read()).decode('utf-8')
                        
                        # 获取下一镜头的原始场景描述作为参考
                        next_shot = req.shots[i + 1]
                        next_shot_description = next_shot.get("description", "")
                        next_shot_story = next_shot.get("shotStory", "")
                        
                        # 运镜选项列表
                        camera_movements = [
                            "slow push-in (推进)",
                            "gentle pull-back (拉远)",
                            "smooth pan left/right (横摇)",
                            "subtle tilt up/down (俯仰)",
                            "orbit around subject (环绕)",
                            "dolly tracking (跟踪)",
                            "static with subject motion (静态+主体运动)"
                        ]
                        
                        # 获取视觉风格
                        visual_style = req.visual_style_prompt or "Filmic realism, natural lighting, soft bokeh, 35mm lens, muted colors, subtle grain."
                        
                        # 获取运镜风格（用户指定或自动选择）
                        camera_instruction = ""
                        if req.camera_movement_prompt:
                            camera_instruction = f"   - CAMERA MOVEMENT: {req.camera_movement_prompt}\n"
                        else:
                            camera_instruction = f"   - CAMERA MOVEMENT: Choose one from: {', '.join(camera_movements)}\n"
                        
                        # 动态动作关键词
                        dynamic_actions = [
                            "rotating", "floating", "rising", "falling", "rippling",
                            "shimmering", "flowing", "drifting", "swaying", "pulsing"
                        ]
                        
                        analysis_prompt = (
                            f"You are a professional video director ensuring continuity. Analyze this frame and generate a DYNAMIC video prompt.\n\n"
                            f"NEXT SHOT CONTEXT:\n"
                            f"- Scene: {next_shot_description}\n"
                            f"- Story: {next_shot_story}\n"
                            f"- Style: {visual_style}\n\n"
                            f"⚠️ MANDATORY PRODUCT VISIBILITY RULES:\n"
                            f"1. The ENTIRE product must be FULLY VISIBLE in frame at ALL times\n"
                            f"2. NO extreme close-ups or macro shots - product must NOT be cropped\n"
                            f"3. Use MEDIUM or WIDE shots only - product occupies 30-70% of frame\n"
                            f"4. Camera movement must be GENTLE - no aggressive push-ins\n"
                            f"5. Product appearance must MATCH the previous frame exactly\n\n"
                            f"DYNAMIC REQUIREMENTS:\n"
                            f"1. Include REAL MOTION - NOT just static image with camera movement\n"
                            f"2. DYNAMIC ELEMENTS: {', '.join(dynamic_actions[:6])}\n"
                            f"3. Background can evolve but product stays fully visible\n\n"
                            f"GENERATE A PROMPT WITH:\n"
                            f"   - Subject: COMPLETE product view, matching previous frame exactly\n"
                            f"   - Framing: Medium/wide shot, product centered and FULLY VISIBLE\n"
                            f"{camera_instruction}"
                            f"   - Environment: Evolving background (light shifts, particles) around product\n"
                            f"   - Continuity: Seamless transition, same product appearance\n\n"
                            f"FORBIDDEN: extreme close-up, macro, tight crop, partial product view\n"
                            f"OUTPUT: A single paragraph prompt. Product must be FULLY VISIBLE throughout."
                        )
                        
                        async with httpx.AsyncClient(timeout=60) as client:
                            target_url = image_api_url
                            if not target_url.endswith("/chat/completions"):
                                target_url = f"{target_url.rstrip('/')}/chat/completions"
                            
                            payload = {
                                "model": analysis_model,
                                "messages": [
                                    {"role": "user", "content": [
                                        {"type": "text", "text": analysis_prompt},
                                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{frame_b64}"}}
                                    ]}
                                ]
                            }
                            headers = {
                                "Content-Type": "application/json",
                                "Authorization": f"Bearer {image_api_key}"
                            }
                            
                            resp = await client.post(target_url, json=payload, headers=headers, timeout=60.0)
                            if resp.status_code == 200:
                                data = resp.json()
                                new_prompt = data.get("choices", [])[0].get("message", {}).get("content", "").strip()
                                if new_prompt:
                                    next_shot_prompt_override = new_prompt
                                    logging.info(f"Chain {chain_id}: Generated continuity prompt for shot {shot_num + 1}: {new_prompt[:100]}...")
                            else:
                                logging.warning(f"Chain {chain_id}: Frame analysis API error: {resp.status_code}")
                    except Exception as analysis_err:
                        logging.warning(f"Chain {chain_id}: Frame analysis failed: {analysis_err}, using original prompt")
        
        # Merge Logic
        logging.info(f"Chain {chain_id}: All shots done. Merging...")
        status["status"] = "merging"
        
        inputs = []
        db = SessionLocal()
        for i, vid_id in enumerate(status["video_ids"]):
             # We can't rely just on DB result_url because we might have downloaded it locally above.
             # But we didn't store the local path in DB. 
             # Re-construct the expected local path logic.
             v = db.query(VideoQueueItem).filter(VideoQueueItem.id == vid_id).first()
             if v and v.result_url:
                 if v.result_url.startswith("/uploads"):
                     inputs.append(f"/app{v.result_url}")
                 elif v.result_url.startswith("http"):
                     # Re-construct downloaded filename
                     shot_n = i + 1
                     inputs.append(f"/app/uploads/queue/chain_{chain_id}_shot_{shot_n}_result.mp4")
                 else:
                     inputs.append(v.result_url)
        db.close()
        
        if len(inputs) > 0:
            concat_list_path = f"/app/uploads/queue/chain_{chain_id}_concat.txt"
            with open(concat_list_path, "w") as f:
                for path in inputs:
                    f.write(f"file '{path}'\n")
            
            output_filename = f"story_chain_{chain_id}.mp4"
            output_path = f"/app/uploads/queue/{output_filename}"
            
            cmd = [
                "ffmpeg", "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_list_path,
                "-c", "copy",
                output_path
            ]
            subprocess.run(cmd, check=True)
            
            final_result_url = f"/uploads/queue/{output_filename}"
            status["merged_video_url"] = final_result_url
            status["status"] = "completed"
            
            # Generate thumbnail from first frame
            thumbnail_filename = f"story_chain_{chain_id}_thumb.jpg"
            thumbnail_path = f"/app/uploads/queue/{thumbnail_filename}"
            try:
                thumb_cmd = [
                    "ffmpeg", "-y",
                    "-i", output_path,
                    "-ss", "00:00:00.500",  # 0.5 second into video
                    "-vframes", "1",
                    "-q:v", "2",  # High quality
                    thumbnail_path
                ]
                subprocess.run(thumb_cmd, check=True, capture_output=True)
                preview_url = f"/uploads/queue/{thumbnail_filename}"
                logging.info(f"Chain {chain_id}: Thumbnail generated at {preview_url}")
            except Exception as thumb_err:
                logging.warning(f"Chain {chain_id}: Failed to generate thumbnail: {thumb_err}")
                preview_url = None
            
            # Add merged result to queue with preview
            db = SessionLocal()
            share_user = db.query(User).filter(User.id == user_id).first()
            user_default_share = share_user.default_share if share_user and share_user.default_share is not None else True
            merged_item = VideoQueueItem(
                id=str(int(uuid.uuid4().int))[:18],
                filename=output_filename,
                file_path=output_path,
                prompt=f"Story Chain {chain_id} Complete",
                status="done",
                result_url=final_result_url,
                user_id=user_id,
                is_merged=True,  # Mark as merged/composite video
                is_shared=user_default_share
            )
            # Set preview_url if thumbnail was generated
            if preview_url:
                merged_item.preview_url = preview_url
            db.add(merged_item)
            db.commit()
            db.close()
            
        # Clear user activity status
        try:
            await connection_manager.update_user_activity(user_id, "在线")
        except Exception:
            pass
            
    except Exception as e:
        logging.error(f"Chain {chain_id} Critical Error: {e}")
        status["status"] = "failed"
        status["error"] = str(e)
        
        # Clear user activity on error too
        try:
            await connection_manager.update_user_activity(user_id, "在线")
        except Exception:
            pass

@app.post("/api/v1/story-chain")
async def create_story_chain(
    req: StoryChainRequest,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user)
):
    chain_id = str(uuid.uuid4())
    
    # Update user activity status
    try:
        await connection_manager.update_user_activity(user.id, f"故事链生成中 ({len(req.shots)}镜头)")
    except Exception:
        pass  # WebSocket not required
    
    background_tasks.add_task(process_story_chain, chain_id, req, user.id)
    return {"chain_id": chain_id, "status": "started"}

@app.get("/api/v1/story-chain/{chain_id}")
async def get_story_chain_status(chain_id: str):
    status = STORY_CHAIN_STATUS.get(chain_id)
    if not status:
        raise HTTPException(status_code=404, detail="Chain not found")
    return status


# =============================================================================
# Story Fission Implementation (裂变式故事生成)
# =============================================================================

class StoryFissionRequest(BaseModel):
    initial_image_url: str  # Base64 or URL
    topic: str = "产品多角度展示"
    branch_count: int = 3  # Number of branches to generate
    visual_style: Optional[str] = None
    visual_style_prompt: Optional[str] = None
    camera_movement: Optional[str] = None
    camera_movement_prompt: Optional[str] = None
    api_url: Optional[str] = None
    api_key: Optional[str] = None
    model_name: Optional[str] = None
    category: str = "other"  # Product category for gallery/video classification

class FissionBranch(BaseModel):
    branch_id: int
    scene_name: str
    theme: str
    product_focus: Optional[str] = None  # 该分支突出的产品特点
    image_prompt: str  # 图片生成 prompt
    video_prompt: str  # 视频生成 prompt
    camera_movement: Optional[str] = "static with motion"  # 默认运镜

STORY_FISSION_STATUS = {}


def repair_truncated_json(content: str) -> str:
    """
    Attempt to repair a truncated JSON array response from AI.
    
    Common truncation patterns:
    - String interrupted mid-token (Unterminated string)
    - Object/array not closed
    
    Returns repaired JSON string or None if repair fails.
    """
    if not content or not content.strip():
        return None
    
    content = content.strip()
    
    # If it doesn't look like JSON array, can't repair
    if not content.startswith('['):
        return None
    
    try:
        # Already valid JSON
        json.loads(content)
        return content
    except json.JSONDecodeError:
        pass
    
    # Track bracket depth
    depth_curly = 0
    depth_square = 0
    in_string = False
    escape_next = False
    last_valid_obj_end = -1
    
    for i, char in enumerate(content):
        if escape_next:
            escape_next = False
            continue
        
        if char == '\\' and in_string:
            escape_next = True
            continue
        
        if char == '"' and not escape_next:
            in_string = not in_string
            continue
        
        if in_string:
            continue
        
        if char == '{':
            depth_curly += 1
        elif char == '}':
            depth_curly -= 1
            # Mark when we close a complete object at array level
            if depth_curly == 0 and depth_square == 1:
                last_valid_obj_end = i
        elif char == '[':
            depth_square += 1
        elif char == ']':
            depth_square -= 1
    
    # If we found at least one complete object, truncate there and close
    if last_valid_obj_end > 0:
        repaired = content[:last_valid_obj_end + 1].rstrip(',').rstrip() + ']'
        try:
            parsed = json.loads(repaired)
            if isinstance(parsed, list) and len(parsed) > 0:
                logging.info(f"JSON repair succeeded: kept {len(parsed)} complete branches")
                return repaired
        except json.JSONDecodeError:
            pass
    
    # Try simply closing unclosed brackets/braces
    repair_suffix = ''
    if in_string:
        repair_suffix += '"'
    for _ in range(depth_curly):
        repair_suffix += '}'
    for _ in range(depth_square):
        repair_suffix += ']'
    
    if repair_suffix:
        try:
            parsed = json.loads(content + repair_suffix)
            if isinstance(parsed, list):
                logging.info(f"JSON repair by closing brackets: {len(parsed)} branches")
                return content + repair_suffix
        except json.JSONDecodeError:
            pass
    
    return None


async def analyze_fission_branches(
    image_b64: str,
    topic: str,
    branch_count: int,
    visual_style_prompt: str,
    api_url: str,
    api_key: str,
    model_name: str,
    category: str = "other"  # Product category for classification only (not used in prompts)
) -> List[dict]:
    """Analyze image and generate multiple story branches."""
    
    # NOTE: category is only used for gallery/video classification, NOT for prompt generation
    # This ensures the AI generates based on the actual product in the image, not assumed category
    
    fission_prompt = f"""Role: 你是一位资深创意总监 + 产品摄影专家，负责将一张产品图片裂变成多个灵动、有表现力的故事分支。

=== 🔴 最高优先级：产品外观一致性 ===
**绝对铁律**：生成的所有图片和视频中，产品必须与输入图片中的产品**100%一致**
- 保持产品的形状、颜色、材质、比例、品牌标识、所有细节不变
- 只改变场景、灯光、背景、氛围，**绝不能改变产品本身的任何外观特征**
- 如果输入图片中产品是白色，输出必须也是白色；如果有特定 logo，必须保留

=== 第一步：产品深度分析 ===
首先，仔细观察图片中的产品，识别并分析：
1. **产品类型与用途**：这是什么产品？主要功能是什么？解决什么问题？
2. **核心特写卖点**：材质质感、工艺细节、按钮/接口、品牌标识、独特设计元素
3. **使用场景联想**：谁会使用？在哪里使用？什么时间使用？
4. **情感价值**：传达什么情绪？科技感？温馨？专业？时尚？

=== 第二步：生成 {branch_count} 个裂变分支 ===
基于以上分析，创作 {branch_count} 个不同的剧情场景分支：

=== 分支创作要求 ===
1. 每个分支必须**突出产品的一个独特卖点或使用场景**
2. **产品外观一致性（最高优先级）**：
   - 产品必须与输入图片中的产品**完全一致**，不允许任何外观变化
   - 保持产品的形状、颜色、材质、比例、品牌标识不变
   - 只改变场景、灯光、氛围，**绝不改变产品本身**
3. **灵动镜头**：避免静态呆板，加入动态元素：
   - 手部交互（触摸、按压、滑动）
   - 环境动态（光线变化、烟雾、水流、粒子）
   - 产品自身动态（旋转、展开、发光、工作状态）
4. **特写与全景结合**：每个分支中既有产品全貌，也有局部特写细节
5. **感官描述**：描述材质触感、光泽反射、声音暗示（如按键声、流水声）
6. 产品始终作为**视觉焦点**，占画面30-70%

=== 视觉风格 ===
{visual_style_prompt or 'Cinematic realism, natural lighting, soft bokeh, 35mm lens, rich textures.'}

=== 运镜选项（选择最能展现该场景特点的运镜）===
- slow push-in (推进) - 适合强调细节
- gentle pull-back (拉远) - 适合展示使用场景
- smooth orbit (环绕) - 适合展示产品全貌
- dolly tracking (跟踪) - 适合跟随手部交互
- macro focus shift (焦点转移) - 适合特写局部
- static with motion (静态+元素运动) - 适合产品工作状态

=== 动态元素关键词库（每个分支至少用2-3个）===
floating particles, light rays shifting, steam rising, water droplets, 
gentle rotation, finger touching, hand interacting, unboxing reveal,
product powering on, LED glowing, screen displaying, texture close-up,
material reflection, smooth sliding, button pressing, liquid pouring

=== 输出格式 ===
输出 JSON 数组，每个分支包含：
[
  {{
    "branch_id": 1,
    "scene_name": "场景名称（中文，如：晨光中的唤醒仪式）",
    "theme": "主题描述（中文，2-3句话，描述这个场景要传达的产品卖点和情感）",
    "product_focus": "该分支突出的产品特点（中文，如：按键触感、材质光泽、使用便捷性）",
    "image_prompt": "CRITICAL: Preserve the EXACT product appearance from the input image. Do not modify product shape, color, or design. [然后是详细的图片生成 prompt（英文），包含：主体描述、场景环境、光影氛围、动态元素]",
    "video_prompt": "The product must remain FULLY VISIBLE and maintain its original appearance throughout the video. [然后是详细的视频生成 prompt（英文），包含：具体的动态动作、运动元素、光影变化、情绪氛围]",
    "camera_movement": "推荐运镜（从上述选项中选择）"
  }},
  ...
]

=== 故事主题 ===
{topic}

只输出 JSON 数组，不要包含 markdown 代码块或其他说明文字。
"""

    target_url = api_url
    if not target_url.endswith("/chat/completions"):
        target_url = f"{target_url.rstrip('/')}/chat/completions"
    
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": fission_prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            }
        ],
        "max_tokens": 8192
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    max_retries = 3
    retry_delay = 5
    
    for attempt in range(1, max_retries + 1):
        try:
            async with httpx.AsyncClient(timeout=180) as client:  # Increased timeout to 3min
                resp = await client.post(target_url, json=payload, headers=headers)
                if resp.status_code == 200:
                    data = resp.json()
                    content = data.get("choices", [])[0].get("message", {}).get("content", "").strip()
                    
                    # Clean markdown if present
                    if content.startswith("```"):
                        lines = content.splitlines()
                        if lines[0].startswith("```"):
                            lines = lines[1:]
                        if lines and lines[-1].strip().startswith("```"):
                            lines = lines[:-1]
                        content = "\n".join(lines).strip()
                    
                    import re
                    content = re.sub(r',\s*\}', '}', content)
                    content = re.sub(r',\s*\]', ']', content)
                    
                    # 🆕 Enhanced JSON repair for truncated responses
                    try:
                        branches = json.loads(content)
                    except json.JSONDecodeError as parse_err:
                        logging.warning(f"Fission analysis: Initial JSON parse failed ({parse_err}), attempting repair...")
                        
                        # Try to repair truncated JSON
                        repaired_content = repair_truncated_json(content)
                        if repaired_content:
                            branches = json.loads(repaired_content)
                            logging.info(f"Fission analysis: JSON repair successful")
                        else:
                            # Re-raise original error if repair failed
                            raise parse_err
                    
                    # --- FIX: Ensure exact number of branches ---
                    if len(branches) < branch_count:
                        logging.warning(f"AI generated {len(branches)} branches, but user requested {branch_count}. Padding...")
                        
                        original_count = len(branches)
                        if original_count > 0:
                            import copy
                            needed = branch_count - original_count
                            for i in range(needed):
                                # Round-robin selection from original branches
                                source_branch = branches[i % original_count]
                                new_branch = copy.deepcopy(source_branch)
                                new_branch["branch_id"] = original_count + i + 1
                                new_branch["scene_name"] = f"{source_branch.get('scene_name', 'Scene')} (Variation {i+1})"
                                branches.append(new_branch)
                        else:
                            # Fallback if 0 branches returned (rare)
                            for i in range(branch_count):
                                branches.append({
                                    "branch_id": i + 1,
                                    "scene_name": f"Auto-generated scene {i+1}",
                                    "theme": "Auto-generated content",
                                    "product_focus": "Product showcase",
                                    "image_prompt": f"Product showcase in professional lighting, scene {i+1}",
                                    "video_prompt": f"Product presentation with dynamic elements, scene {i+1}",
                                    "camera_movement": "slow push-in"
                                })
                    
                    # Truncate if too many (rare but possible)
                    if len(branches) > branch_count:
                        branches = branches[:branch_count]
                    # ---------------------------------------------
                    
                    return branches
                    
                elif resp.status_code in [502, 503, 504, 429]:
                    logging.warning(f"Fission analysis attempt {attempt}: HTTP {resp.status_code}, retrying...")
                    if attempt < max_retries:
                        await asyncio.sleep(retry_delay * attempt)
                        continue
                    raise Exception(f"Fission analysis API error after {max_retries} attempts: {resp.status_code}")
                else:
                    raise Exception(f"Fission analysis API error: {resp.status_code} - {resp.text[:200]}")
                    
        except httpx.TimeoutException:
            logging.warning(f"Fission analysis timeout on attempt {attempt}")
            if attempt < max_retries:
                await asyncio.sleep(retry_delay)
                continue
            raise Exception(f"Fission analysis timed out after {max_retries} attempts")
        except json.JSONDecodeError as e:
            logging.warning(f"Fission analysis JSON parse error on attempt {attempt}: {e}")
            if attempt < max_retries:
                await asyncio.sleep(retry_delay)
                continue
            raise


async def generate_branch_image(
    branch: dict,
    original_b64: str,
    fission_id: str,
    image_api_url: str,
    image_api_key: str,
    image_model: str,
    user_id: int = None,  # for gallery save
    category: str = "other"  # NEW: Product category for classification
) -> str:
    """Generate a stylized image for a branch with retry mechanism. Returns local file path."""
    branch_id = branch.get("branch_id", 0)
    image_prompt = branch.get("image_prompt", "")
    scene_name = branch.get("scene_name", f"分支{branch_id}")
    
    max_retries = 3
    retry_delay = 5  # seconds between retries
    
    for attempt in range(1, max_retries + 1):
        try:
            logging.info(f"Fission {fission_id}: Generating image for branch {branch_id} (attempt {attempt}/{max_retries})")
            
            async with httpx.AsyncClient(timeout=300) as client:  # Increased timeout to 5 min
                # 强化产品一致性约束 - 添加强制性前缀
                enhanced_prompt = (
                    "🔴 MANDATORY: Keep the product EXACTLY as shown in the input image - "
                    "same shape, color, material, brand logo, and all visual details MUST remain unchanged. "
                    "Only modify the scene/lighting/background around it. "
                    f"{image_prompt}"
                )
                
                result = await call_openai_compatible_api(
                    client, image_api_url, image_api_key,
                    original_b64,  # Input image
                    original_b64,  # Reference image
                    f"Branch_{branch_id}",
                    enhanced_prompt,  # 使用强化后的 prompt
                    image_model
                )
                
                if result and result.image_base64:
                    raw_content = result.image_base64.strip()
                    image_data = None
                    
                    # Check if API returned a URL (common for some image generation APIs)
                    if raw_content.startswith("http://") or raw_content.startswith("https://"):
                        logging.info(f"Fission {fission_id}: Branch {branch_id} - downloading image from URL")
                        try:
                            async with httpx.AsyncClient(timeout=120) as dl_client:
                                img_resp = await dl_client.get(raw_content)
                                if img_resp.status_code == 200:
                                    image_data = img_resp.content
                                else:
                                    raise Exception(f"Failed to download image: HTTP {img_resp.status_code}")
                        except Exception as dl_err:
                            logging.warning(f"Fission {fission_id}: Branch {branch_id} URL download failed: {dl_err}")
                            raise Exception(f"Branch {branch_id}: Failed to download image from URL")
                    
                    # Check for data:image prefix
                    elif raw_content.startswith("data:"):
                        # Extract base64 part: data:image/jpeg;base64,XXXXX
                        if ",base64," in raw_content:
                            b64_data = raw_content.split(",base64,", 1)[1]
                        elif "," in raw_content:
                            b64_data = raw_content.split(",", 1)[1]
                        else:
                            b64_data = raw_content
                        
                        # Fix padding if needed
                        missing_padding = len(b64_data) % 4
                        if missing_padding:
                            b64_data += '=' * (4 - missing_padding)
                        
                        try:
                            image_data = base64.b64decode(b64_data)
                        except Exception as decode_err:
                            logging.warning(f"Fission {fission_id}: Branch {branch_id} base64 decode failed: {decode_err}")
                            raise Exception(f"Branch {branch_id}: Invalid base64 image data")
                    
                    # Assume raw base64 content
                    else:
                        b64_data = raw_content
                        
                        # Check if content looks like base64 (not text description)
                        if len(b64_data) < 1000:
                            logging.warning(f"Fission {fission_id}: Branch {branch_id} returned short content: {b64_data[:200]}")
                            raise Exception(f"Branch {branch_id}: API returned text description instead of image")
                        
                        # Fix padding if needed
                        missing_padding = len(b64_data) % 4
                        if missing_padding:
                            b64_data += '=' * (4 - missing_padding)
                        
                        try:
                            image_data = base64.b64decode(b64_data)
                        except Exception as decode_err:
                            logging.warning(f"Fission {fission_id}: Branch {branch_id} base64 decode failed: {decode_err}")
                            raise Exception(f"Branch {branch_id}: Invalid base64 image data")
                    
                    if not image_data:
                        raise Exception(f"Branch {branch_id}: No image data obtained")
                    
                    # Convert to standard JPEG using PIL for video API compatibility
                    from PIL import Image
                    from io import BytesIO
                    
                    try:
                        img = Image.open(BytesIO(image_data))
                    except Exception as pil_err:
                        logging.warning(f"Fission {fission_id}: Branch {branch_id} PIL cannot open: {pil_err}")
                        raise Exception(f"Branch {branch_id}: Cannot parse image data")
                    
                    # Convert to RGB if necessary (handles RGBA, P mode etc)
                    if img.mode in ('RGBA', 'P', 'LA'):
                        img = img.convert('RGB')
                    
                    img_width, img_height = img.size
                    
                    # Save to queue folder (for video generation input)
                    queue_filename = f"fission_{fission_id}_branch_{branch_id}.jpg"
                    queue_filepath = f"/app/uploads/queue/{queue_filename}"
                    img.save(queue_filepath, 'JPEG', quality=95)
                    logging.info(f"Fission {fission_id}: Branch {branch_id} image saved to {queue_filepath}")
                    
                    # Also save to gallery
                    gallery_dir = "/app/uploads/gallery"
                    os.makedirs(gallery_dir, exist_ok=True)
                    gallery_filename = f"fission_{fission_id}_{branch_id}_{scene_name[:20]}.jpg"
                    # Sanitize filename
                    gallery_filename = "".join(c if c.isalnum() or c in '._-' else '_' for c in gallery_filename)
                    gallery_filepath = os.path.join(gallery_dir, gallery_filename)
                    img.save(gallery_filepath, 'JPEG', quality=95)
                    
                    # Create gallery database record
                    if user_id:
                        try:
                            db = SessionLocal()
                            share_user = db.query(User).filter(User.id == user_id).first()
                            user_default_share = share_user.default_share if share_user and share_user.default_share is not None else True
                            gallery_item = SavedImage(
                                user_id=user_id,
                                filename=gallery_filename,
                                file_path=gallery_filepath,
                                url=f"/uploads/gallery/{gallery_filename}",
                                prompt=f"[Fission] {scene_name}: {image_prompt[:200]}",
                                width=img_width,
                                height=img_height,
                                category=category,  # Use user-selected category instead of hardcoded 'fission'
                                is_shared=user_default_share
                            )
                            db.add(gallery_item)
                            db.commit()
                            db.close()
                            logging.info(f"Fission {fission_id}: Branch {branch_id} image saved to gallery")
                        except Exception as gallery_err:
                            logging.warning(f"Fission {fission_id}: Gallery save error: {gallery_err}")
                    
                    return queue_filepath
                else:
                    raise Exception(f"Branch {branch_id} image generation returned no result")
                    
        except httpx.TimeoutException as e:
            logging.warning(f"Fission {fission_id}: Branch {branch_id} timeout on attempt {attempt}: {e}")
            if attempt < max_retries:
                await asyncio.sleep(retry_delay)
                continue
            raise Exception(f"Branch {branch_id} timed out after {max_retries} attempts")
            
        except httpx.HTTPStatusError as e:
            logging.warning(f"Fission {fission_id}: Branch {branch_id} HTTP error on attempt {attempt}: {e.response.status_code}")
            if attempt < max_retries and e.response.status_code in [502, 503, 504, 429]:
                await asyncio.sleep(retry_delay * attempt)  # Exponential backoff
                continue
            raise
            
        except Exception as e:
            error_msg = str(e)
            logging.warning(f"Fission {fission_id}: Branch {branch_id} error on attempt {attempt}: {error_msg}")
            if attempt < max_retries and ("timeout" in error_msg.lower() or "504" in error_msg or "503" in error_msg):
                await asyncio.sleep(retry_delay)
                continue
            raise
    
    raise Exception(f"Branch {branch_id} failed after {max_retries} attempts")


async def generate_branch_video(
    branch: dict,
    image_path: str,
    fission_id: str,
    video_api_url: str,
    video_api_key: str,
    video_model: str,
    user_id: int,
    category: str = "other"  # NEW: Product category for classification
) -> dict:
    """Generate video for a branch. Returns result dict."""
    branch_id = branch.get("branch_id", 0)
    video_prompt = branch.get("video_prompt", "")
    camera = branch.get("camera_movement", "")
    
    full_prompt = f"{video_prompt} Camera: {camera}"
    
    logging.info(f"Fission {fission_id}: Generating video for branch {branch_id}")
    
    # Create queue item
    db = SessionLocal()
    try:
        queue_filename = f"fission_{fission_id}_branch_{branch_id}_input.jpg"
        queue_file_path = f"/app/uploads/queue/{queue_filename}"
        
        import shutil
        shutil.copy(image_path, queue_file_path)
        
        # 获取用户的分享设置
        share_user = db.query(User).filter(User.id == user_id).first()
        user_default_share = share_user.default_share if share_user and share_user.default_share is not None else True
        
        new_item = VideoQueueItem(
            id=str(int(uuid.uuid4().int))[:18],
            filename=queue_filename,
            file_path=queue_file_path,
            prompt=full_prompt,
            status="pending",
            user_id=user_id,
            category=category,  # Use user-selected category
            is_shared=user_default_share
        )
        db.add(new_item)
        db.commit()
        db.refresh(new_item)
        item_id = new_item.id
    finally:
        db.close()
    
    # Generate video with retry
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        if attempt > 1:
            db = SessionLocal()
            retry_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
            if retry_item:
                retry_item.status = "pending"
                retry_item.error_msg = None
                db.commit()
            db.close()
        
        await process_video_with_auto_retry(item_id, video_api_url, video_api_key, video_model, skip_concurrency_check=True)
        
        db = SessionLocal()
        completed_item = db.query(VideoQueueItem).filter(VideoQueueItem.id == item_id).first()
        
        if completed_item.status == "done":
            result_url = completed_item.result_url
            db.close()
            
            # Download if remote
            local_video_path = None
            if result_url.startswith("/uploads"):
                local_video_path = f"/app{result_url}"
            elif result_url.startswith("http"):
                try:
                    result_url = await convert_sora_to_watermark_free(result_url)
                    async with httpx.AsyncClient() as client:
                        resp = await client.get(result_url, timeout=300)
                        if resp.status_code == 200:
                            local_filename = f"fission_{fission_id}_branch_{branch_id}_result.mp4"
                            local_video_path = f"/app/uploads/queue/{local_filename}"
                            with open(local_video_path, "wb") as f:
                                f.write(resp.content)
                except Exception as dl_err:
                    logging.warning(f"Fission {fission_id}: Branch {branch_id} download error: {dl_err}")
            
            return {
                "branch_id": branch_id,
                "scene_name": branch.get("scene_name", ""),
                "theme": branch.get("theme", ""),
                "video_url": f"/uploads/queue/fission_{fission_id}_branch_{branch_id}_result.mp4" if (local_video_path and os.path.exists(local_video_path)) else result_url,
                "local_video_path": local_video_path if (local_video_path and os.path.exists(local_video_path)) else None,
                "image_url": f"/uploads/queue/fission_{fission_id}_branch_{branch_id}.jpg",
                "status": "done"
            }
        else:
            error_msg = completed_item.error_msg or "Unknown error"
            db.close()
            if attempt == max_retries:
                return {
                    "branch_id": branch_id,
                    "status": "error",
                    "error": error_msg
                }
            await asyncio.sleep(5)
    
    return {"branch_id": branch_id, "status": "error", "error": "Max retries exceeded"}


async def process_story_fission(fission_id: str, req: StoryFissionRequest, user_id: int):
    """Main fission processing function with 3-concurrent batching."""
    status = {
        "status": "processing",
        "phase": "analyzing",
        "total_branches": req.branch_count,
        "completed_branches": 0,
        "branches": [],
        "error": None
    }
    STORY_FISSION_STATUS[fission_id] = status
    
    # Config resolution
    db = SessionLocal()
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    db.close()
    
    image_api_url = req.api_url or config_dict.get("api_url", "")
    image_api_key = req.api_key or config_dict.get("api_key", "")
    image_model = config_dict.get("model_name", "gemini-3-pro-image-preview")  # Use same as settings
    analysis_model = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    video_api_url = config_dict.get("video_api_url", "")
    video_api_key = config_dict.get("video_api_key", "")
    video_model = req.model_name or config_dict.get("video_model_name", "sora-video-portrait")
    
    try:
        # Step 1: Decode original image
        logging.info(f"Fission {fission_id}: Starting with {req.branch_count} branches")
        
        if req.initial_image_url.startswith("data:"):
            header, encoded = req.initial_image_url.split(",", 1)
            original_b64 = encoded
        else:
            async with httpx.AsyncClient() as client:
                resp = await client.get(req.initial_image_url, timeout=30)
                original_b64 = base64.b64encode(resp.content).decode('utf-8')
        
        
        # Step 2: Analyze and generate branches
        # 🆕 Add global concurrency control for analysis phase to prevent API overload
        status["phase"] = "waiting_analysis"
        logging.info(f"Fission {fission_id}: Waiting for analysis slot...")
        
        # Get the global limiter
        limiter = await get_concurrency_limiter()
        
        # Wait for analysis slot (max 1-2 concurrent analysis to avoid API queue overflow)
        analysis_acquired = False
        max_wait_attempts = 60  # Max 5 minutes waiting (60 * 5s)
        for wait_attempt in range(max_wait_attempts):
            analysis_acquired = await limiter.acquire_global("analysis", timeout=600)
            if analysis_acquired:
                logging.info(f"Fission {fission_id}: Acquired analysis slot after {wait_attempt * 5}s")
                break
            else:
                # Update status to show waiting
                status["phase"] = f"waiting_analysis ({wait_attempt + 1})"
                await asyncio.sleep(5)  # Wait 5 seconds before retry
        
        if not analysis_acquired:
            raise Exception("Analysis queue timeout - too many concurrent tasks")
        
        try:
            status["phase"] = "analyzing"
            logging.info(f"Fission {fission_id}: Analyzing image to generate {req.branch_count} branches")
            
            branches = await analyze_fission_branches(
                original_b64,
                req.topic,
                req.branch_count,
                req.visual_style_prompt or "",
                image_api_url,
                image_api_key,
                analysis_model,
                req.category  # Pass category for tailored prompts
            )
        finally:
            # Release analysis slot regardless of success/failure
            await limiter.release_global("analysis")
        
        logging.info(f"Fission {fission_id}: Generated {len(branches)} branch scripts")
        status["branches"] = [{"branch_id": b.get("branch_id"), "scene_name": b.get("scene_name"), "status": "pending", "retry_count": 0} for b in branches]
        
        # Step 3: Generate images with batch-level retry mechanism
        status["phase"] = "generating_images"
        os.makedirs("/app/uploads/queue", exist_ok=True)
        
        # 🆕 First branch uses original image directly (no generation needed)
        image_paths = {}
        if len(branches) > 0:
            first_branch_id = branches[0].get("branch_id", 1)
            first_branch_path = f"/app/uploads/queue/fission_{fission_id}_branch_{first_branch_id}.jpg"
            
            # Decode and save original image for first branch
            from PIL import Image
            from io import BytesIO
            original_img_data = base64.b64decode(original_b64)
            original_img = Image.open(BytesIO(original_img_data))
            if original_img.mode in ('RGBA', 'P', 'LA'):
                original_img = original_img.convert('RGB')
            original_img.save(first_branch_path, 'JPEG', quality=95)
            
            # Record first branch as complete
            image_paths[first_branch_id] = first_branch_path
            status["branches"][0]["status"] = "image_done"
            status["branches"][0]["image_url"] = f"/uploads/queue/fission_{fission_id}_branch_{first_branch_id}.jpg"
            logging.info(f"Fission {fission_id}: Branch {first_branch_id} using original image directly")
        
        # Only generate images for remaining branches (skip first)
        branches_to_process_for_images = branches[1:] if len(branches) > 1 else []
        
        # Get concurrency settings
        concurrency_config = await get_concurrency_config()
        # Default to 5 if not set
        concurrent_limit = int(concurrency_config.get("max_concurrent_image", 5))
        logging.info(f"Fission {fission_id}: Using dynamic concurrency limit: {concurrent_limit}")
        
        # Get global limiter
        limiter = await get_concurrency_limiter()
        
        async def generate_with_limit(branch, batch_index=0):
            """Wrapper to enforce global concurrency limit with request spacing"""
            branch_id = branch.get("branch_id")
            
            # Add spacing within batch to prevent CF rate limiting (0.3-0.8s per request)
            if batch_index > 0:
                spacing_delay = random.uniform(0.3, 0.8) * batch_index
                await asyncio.sleep(spacing_delay)
            
            # Apply global throttling
            await throttle_request()
            
            # Acquire global slot
            acquired = await limiter.acquire_global("image_gen", timeout=600)
            if not acquired:
                logging.warning(f"Fission {fission_id}: Branch {branch_id} failed to acquire global slot (timeout)")
                return Exception("Global concurrency limit reached (timeout)")
            
            try:
                return await generate_branch_image(
                    branch, original_b64, fission_id, 
                    image_api_url, image_api_key, image_model, user_id, req.category
                )
            finally:
                await limiter.release_global("image_gen")

        # Batch-level retry mechanism with Sliding Window (Semaphore-based)
        # This allows completed slots to be immediately filled with new tasks
        MAX_RETRY_ROUNDS = 3
        RETRY_DELAY = 5  # seconds between retry rounds
        
        branches_to_process = branches_to_process_for_images.copy()  # Start with branches except first
        retry_round = 0
        
        # Create semaphore for sliding window concurrency control
        semaphore = asyncio.Semaphore(concurrent_limit)

        
        async def generate_with_semaphore(branch, branch_index_in_batch):
            """Wrapper using semaphore for true sliding window behavior"""
            async with semaphore:
                # Add spacing to prevent CF rate limiting
                if branch_index_in_batch > 0:
                    spacing_delay = random.uniform(0.3, 0.8) * min(branch_index_in_batch, 3)
                    await asyncio.sleep(spacing_delay)
                
                # Apply global throttling
                await throttle_request()
                
                branch_id = branch.get("branch_id")
                
                # Acquire global slot
                acquired = await limiter.acquire_global("image_gen", timeout=600)
                if not acquired:
                    logging.warning(f"Fission {fission_id}: Branch {branch_id} failed to acquire global slot (timeout)")
                    return {"branch_id": branch_id, "error": "Global concurrency limit reached (timeout)"}
                
                try:
                    result = await generate_branch_image(
                        branch, original_b64, fission_id, 
                        image_api_url, image_api_key, image_model, user_id, req.category
                    )
                    return {"branch_id": branch_id, "result": result}
                except Exception as e:
                    return {"branch_id": branch_id, "error": str(e)}
                finally:
                    await limiter.release_global("image_gen")
        
        while retry_round < MAX_RETRY_ROUNDS:
            failed_branches = []
            status["retry_round"] = retry_round + 1
            
            # Use sliding window: all tasks start concurrently, semaphore limits active count
            logging.info(f"Fission {fission_id}: Starting {len(branches_to_process)} image tasks with sliding window (limit={concurrent_limit})")
            
            # Launch all tasks at once - semaphore handles concurrency
            tasks = [generate_with_semaphore(b, idx) for idx, b in enumerate(branches_to_process)]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, Exception):
                    # Gather-level exception
                    logging.error(f"Fission {fission_id}: Task exception: {result}")
                    continue
                
                branch_id = result.get("branch_id")
                
                # Find branch index in original branches list
                branch_idx = next((idx for idx, b in enumerate(branches) if b.get("branch_id") == branch_id), None)
                
                if branch_idx is None:
                    logging.error(f"Fission {fission_id}: Cannot find branch {branch_id} in original list")
                    continue
                
                if "error" in result:
                    # Track failure
                    original_branch = next((b for b in branches if b.get("branch_id") == branch_id), None)
                    logging.warning(f"Fission {fission_id}: Branch {branch_id} failed (attempt {retry_round + 1}/{MAX_RETRY_ROUNDS}): {result['error']}")
                    status["branches"][branch_idx]["status"] = "image_error"
                    status["branches"][branch_idx]["error"] = result["error"]
                    status["branches"][branch_idx]["retry_count"] = retry_round + 1
                    if original_branch:
                        failed_branches.append(original_branch)
                else:
                    # Success
                    image_paths[branch_id] = result["result"]
                    status["branches"][branch_idx]["status"] = "image_done"
                    status["branches"][branch_idx]["image_url"] = f"/uploads/queue/fission_{fission_id}_branch_{branch_id}.jpg"
                    logging.info(f"Fission {fission_id}: Branch {branch_id} image generated successfully")
            
            # Check if all branches succeeded
            if not failed_branches:
                logging.info(f"Fission {fission_id}: All {len(branches)} images generated successfully")
                break
            
            # Check if we should retry
            retry_round += 1
            if retry_round < MAX_RETRY_ROUNDS:
                status["failed_count"] = len(failed_branches)
                logging.warning(f"Fission {fission_id}: {len(failed_branches)} branches failed, starting retry round {retry_round + 1} after {RETRY_DELAY}s...")
                await asyncio.sleep(RETRY_DELAY)
                branches_to_process = failed_branches  # Only retry failed branches
            else:
                # Max retries reached
                logging.error(f"Fission {fission_id}: {len(failed_branches)} branches still failed after {MAX_RETRY_ROUNDS} retry rounds")
                status["failed_count"] = len(failed_branches)
                status["error"] = f"{len(failed_branches)} 个分支图片生成失败，已达最大重试次数"
                status["status"] = "partial_failure"
                # Don't break - continue to video generation with available images
        
        # Step 4: Generate videos SEQUENTIALLY with tail-frame continuation
        # 🆕 Changed from parallel to sequential to enable visual continuity between shots
        status["phase"] = "generating_videos"
        logging.info(f"Fission {fission_id}: Starting SEQUENTIAL video generation with tail-frame continuation")
        
        # Get video concurrent limit (still used for global slot acquisition)
        video_limit = int(concurrency_config.get("max_concurrent_video", 3))
        
        import shutil
        
        # Track current input image for each branch (starts with branch's own image, then uses tail frame)
        current_tail_frame = None
        
        for idx, branch in enumerate(branches):
            branch_id = branch.get("branch_id", idx + 1)
            
            # Check if this branch has an image
            if branch_id not in image_paths:
                logging.warning(f"Fission {fission_id}: Skipping branch {branch_id} - no input image")
                continue
            
            # Determine input image for this branch
            if idx > 0 and current_tail_frame and os.path.exists(current_tail_frame):
                # Use tail frame from previous video as input
                input_img_path = f"/app/uploads/queue/fission_{fission_id}_branch_{branch_id}_input_from_tail.jpg"
                shutil.copy(current_tail_frame, input_img_path)
                logging.info(f"Fission {fission_id}: Branch {branch_id} using tail frame from previous branch")
            else:
                # Use branch's own generated/original image
                input_img_path = image_paths[branch_id]
            
            # Generate video (sequential execution with retry)
            logging.info(f"Fission {fission_id}: Generating video for branch {branch_id} ({idx+1}/{len(branches)})")
            
            max_retries = 5
            VIDEO_RETRY_BASE_DELAY = 10
            result = None
            
            for attempt in range(max_retries):
                if attempt > 0:
                    retry_delay = VIDEO_RETRY_BASE_DELAY * (1.5 ** (attempt - 1)) + random.uniform(3, 8)
                    logging.info(f"Fission {fission_id}: Branch {branch_id} waiting {retry_delay:.1f}s before retry (anti-CF)")
                    await asyncio.sleep(retry_delay)
                
                # Apply global throttling
                await throttle_request()
                
                # Try to acquire global slot
                acquired = await limiter.acquire_global("video_gen", timeout=60)
                
                if acquired:
                    try:
                        logging.info(f"Fission {fission_id}: Branch {branch_id} acquired slot (attempt {attempt + 1})")
                        result = await generate_branch_video(
                            branch, input_img_path, fission_id,
                            video_api_url, video_api_key, video_model, user_id, req.category
                        )
                        break  # Success, exit retry loop
                    finally:
                        await limiter.release_global("video_gen")
                else:
                    if attempt < max_retries - 1:
                        logging.info(f"Fission {fission_id}: Branch {branch_id} slot busy (attempt {attempt + 1}/{max_retries})")
                    else:
                        logging.error(f"Fission {fission_id}: Branch {branch_id} failed after {max_retries} attempts")
                        result = {"branch_id": branch_id, "status": "error", "error": f"Failed to acquire slot after {max_retries} retries"}
            
            if result is None:
                result = {"branch_id": branch_id, "status": "error", "error": "Unexpected retry loop exit"}
            
            # Update branch status
            for bi, bs in enumerate(status["branches"]):
                if bs.get("branch_id") == branch_id:
                    status["branches"][bi].update(result)
                    if result.get("status") == "done":
                        status["completed_branches"] += 1
                    break
            
            # 🆕 Extract tail frame for next branch (only if this video succeeded and not the last branch)
            if result.get("status") == "done" and idx < len(branches) - 1:
                video_path = result.get("local_video_path") or result.get("video_url", "")
                if video_path:
                    tail_frame_path = extract_last_frame(video_path)
                    if tail_frame_path and os.path.exists(tail_frame_path):
                        current_tail_frame = tail_frame_path
                        logging.info(f"Fission {fission_id}: Extracted tail frame from branch {branch_id} -> {tail_frame_path}")
                    else:
                        logging.warning(f"Fission {fission_id}: Failed to extract tail frame from branch {branch_id}, next branch will use its own image")
                        current_tail_frame = None
                else:
                    logging.warning(f"Fission {fission_id}: No video path available for branch {branch_id}")
                    current_tail_frame = None
        
        # Step 5: Merge all successful videos into one
        status["phase"] = "merging"
        logging.info(f"Fission {fission_id}: Merging {status['completed_branches']} videos...")
        
        # Collect successful video paths (validate existence)
        video_inputs = []
        for branch in status["branches"]:
            if branch.get("status") == "done":
                # Prefer local video path
                local_path = branch.get("local_video_path")
                if local_path and os.path.exists(local_path):
                    video_inputs.append(local_path)
                else:
                    # Try to construct path from video_url
                    video_url = branch.get("video_url", "")
                    if video_url.startswith("/uploads"):
                        full_path = f"/app{video_url}"
                        if os.path.exists(full_path):
                            video_inputs.append(full_path)
                        else:
                            logging.warning(f"Fission {fission_id}: Video file not found: {full_path}")
                    # Skip remote URLs for now (not downloaded)
        
        logging.info(f"Fission {fission_id}: Found {len(video_inputs)} local videos to merge")
        
        if len(video_inputs) > 0:
            # Create concat list file
            concat_list_path = f"/app/uploads/queue/fission_{fission_id}_concat.txt"
            with open(concat_list_path, "w") as f:
                for path in video_inputs:
                    f.write(f"file '{path}'\n")
            
            # Merge videos using ffmpeg
            output_filename = f"story_fission_{fission_id}.mp4"
            output_path = f"/app/uploads/queue/{output_filename}"
            
            try:
                # Try fast copy first
                merge_cmd = [
                    "ffmpeg", "-y",
                    "-f", "concat",
                    "-safe", "0",
                    "-i", concat_list_path,
                    "-c", "copy",
                    output_path
                ]
                result = subprocess.run(merge_cmd, capture_output=True, text=True)
                
                if result.returncode != 0:
                    logging.warning(f"Fission {fission_id}: Fast merge failed ({result.returncode}), trying re-encode...")
                    # Fallback: re-encode for compatibility
                    merge_cmd_reencode = [
                        "ffmpeg", "-y",
                        "-f", "concat",
                        "-safe", "0",
                        "-i", concat_list_path,
                        "-c:v", "libx264",
                        "-c:a", "aac",
                        "-preset", "fast",
                        output_path
                    ]
                    subprocess.run(merge_cmd_reencode, check=True, capture_output=True)
                
                if os.path.exists(output_path):
                    final_result_url = f"/uploads/queue/{output_filename}"
                    status["merged_video_url"] = final_result_url
                else:
                    raise Exception("Merged video file not created")
                    
            except subprocess.CalledProcessError as ffmpeg_err:
                logging.error(f"Fission {fission_id}: FFmpeg merge failed: {ffmpeg_err}")
                status["error"] = f"Video merge failed: {ffmpeg_err}"
                status["status"] = "failed"
                return
            
            # Generate thumbnail
            thumbnail_filename = f"story_fission_{fission_id}_thumb.jpg"
            thumbnail_path = f"/app/uploads/queue/{thumbnail_filename}"
            try:
                thumb_cmd = [
                    "ffmpeg", "-y",
                    "-i", output_path,
                    "-ss", "00:00:00.500",
                    "-vframes", "1",
                    "-q:v", "2",
                    thumbnail_path
                ]
                subprocess.run(thumb_cmd, check=True, capture_output=True)
                status["thumbnail_url"] = f"/uploads/queue/{thumbnail_filename}"
            except Exception as thumb_err:
                logging.warning(f"Fission {fission_id}: Thumbnail generation failed: {thumb_err}")
            
            # Add merged result to queue
            db = SessionLocal()
            share_user = db.query(User).filter(User.id == user_id).first()
            user_default_share = share_user.default_share if share_user and share_user.default_share is not None else True
            merged_item = VideoQueueItem(
                id=str(int(uuid.uuid4().int))[:18],
                filename=output_filename,
                file_path=output_path,
                prompt=f"Story Fission {fission_id} - {req.topic}",
                status="done",
                result_url=final_result_url,
                user_id=user_id,
                is_merged=True,  # Mark as merged/composite video
                is_shared=user_default_share
            )
            if status.get("thumbnail_url"):
                merged_item.preview_url = status["thumbnail_url"]
            db.add(merged_item)
            db.commit()
            db.close()
            
            logging.info(f"Fission {fission_id}: Merged video saved to {final_result_url}")
        
        status["status"] = "completed"
        status["phase"] = "done"
        logging.info(f"Fission {fission_id}: Completed with {status['completed_branches']}/{len(branches)} successful branches")
        
        # Clear user activity status
        try:
            await connection_manager.update_user_activity(user_id, "在线")
        except Exception:
            pass
        
    except Exception as e:
        logging.error(f"Fission {fission_id} Critical Error: {e}")
        import traceback
        traceback.print_exc()
        status["status"] = "failed"
        status["error"] = str(e)
        
        # Clear user activity on error too
        try:
            await connection_manager.update_user_activity(user_id, "在线")
        except Exception:
            pass


@app.post("/api/v1/story-fission")
async def create_story_fission(
    req: StoryFissionRequest,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user)
):
    """Start a fission-style story generation (parallel branches)."""
    fission_id = str(uuid.uuid4())
    
    # Update user activity status
    try:
        await connection_manager.update_user_activity(user.id, f"裂变生成中 ({req.branch_count}分支)")
    except Exception:
        pass  # WebSocket not required
    
    background_tasks.add_task(process_story_fission, fission_id, req, user.id)
    return {"fission_id": fission_id, "status": "started", "branch_count": req.branch_count}


@app.get("/api/v1/story-fission/{fission_id}")
async def get_story_fission_status(fission_id: str):
    """Get the status of a fission generation."""
    status = STORY_FISSION_STATUS.get(fission_id)
    if not status:
        raise HTTPException(status_code=404, detail="Fission not found")
    return status


class RetryBranchRequest(BaseModel):
    """Request body for retrying a single branch."""
    pass  # No additional fields needed for now


@app.post("/api/v1/story-fission/{fission_id}/branch/{branch_id}/retry")
async def retry_fission_branch(
    fission_id: str,
    branch_id: int,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user)
):
    """Retry video generation for a failed branch."""
    status = STORY_FISSION_STATUS.get(fission_id)
    if not status:
        raise HTTPException(status_code=404, detail="Fission not found")
    
    # Find the failed branch
    branch_data = None
    branch_idx = None
    for idx, b in enumerate(status.get("branches", [])):
        if b.get("branch_id") == branch_id:
            branch_data = b
            branch_idx = idx
            break
    
    if branch_data is None:
        raise HTTPException(status_code=404, detail=f"Branch {branch_id} not found")
    
    # Check if branch has an image
    image_url = branch_data.get("image_url", "")
    if not image_url:
        raise HTTPException(status_code=400, detail="Branch has no image, cannot retry video")
    
    # Get config
    db = SessionLocal()
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    db.close()
    
    video_api_url = config_dict.get("video_api_url", "")
    video_api_key = config_dict.get("video_api_key", "")
    video_model = config_dict.get("video_model_name", "sora-video-portrait")
    category = branch_data.get("category", "other")
    
    # Construct image path
    image_path = f"/app{image_url}" if image_url.startswith("/uploads") else image_url
    
    if not os.path.exists(image_path):
        raise HTTPException(status_code=400, detail=f"Image file not found: {image_path}")
    
    # Prepare branch dict for video generation
    branch = {
        "branch_id": branch_id,
        "video_prompt": branch_data.get("video_prompt", ""),
        "camera_movement": branch_data.get("camera_movement", ""),
    }
    
    # Update status to processing
    status["branches"][branch_idx]["status"] = "processing"
    status["branches"][branch_idx]["error"] = None
    
    async def retry_video_task():
        try:
            result = await generate_branch_video(
                branch, image_path, fission_id,
                video_api_url, video_api_key, video_model, user.id, category
            )
            
            if result.get("status") == "done":
                status["branches"][branch_idx].update(result)
                status["completed_branches"] = sum(1 for b in status["branches"] if b.get("status") == "done")
                logging.info(f"Fission {fission_id}: Branch {branch_id} retry succeeded")
            else:
                status["branches"][branch_idx]["status"] = "video_error"
                status["branches"][branch_idx]["error"] = result.get("error", "Unknown error")
                logging.warning(f"Fission {fission_id}: Branch {branch_id} retry failed: {result.get('error')}")
        except Exception as e:
            status["branches"][branch_idx]["status"] = "video_error"
            status["branches"][branch_idx]["error"] = str(e)
            logging.error(f"Fission {fission_id}: Branch {branch_id} retry exception: {e}")
    
    background_tasks.add_task(retry_video_task)
    
    return {"status": "retrying", "branch_id": branch_id}


@app.post("/api/v1/story-fission/{fission_id}/remerge")
async def remerge_fission_story(
    fission_id: str,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user)
):
    """Re-merge all successful branch videos into the final story video."""
    status = STORY_FISSION_STATUS.get(fission_id)
    if not status:
        raise HTTPException(status_code=404, detail="Fission not found")
    
    # Collect successful video paths
    video_inputs = []
    for branch in status.get("branches", []):
        if branch.get("status") == "done":
            local_path = branch.get("local_video_path")
            if local_path and os.path.exists(local_path):
                video_inputs.append(local_path)
            else:
                video_url = branch.get("video_url", "")
                if video_url.startswith("/uploads"):
                    full_path = f"/app{video_url}"
                    if os.path.exists(full_path):
                        video_inputs.append(full_path)
    
    if len(video_inputs) == 0:
        raise HTTPException(status_code=400, detail="No successful videos to merge")
    
    logging.info(f"Fission {fission_id}: Remerging {len(video_inputs)} videos")
    
    async def remerge_task():
        try:
            # Create concat list file
            concat_list_path = f"/app/uploads/queue/fission_{fission_id}_concat.txt"
            with open(concat_list_path, "w") as f:
                for path in video_inputs:
                    f.write(f"file '{path}'\n")
            
            output_filename = f"story_fission_{fission_id}.mp4"
            output_path = f"/app/uploads/queue/{output_filename}"
            
            # Try fast copy first
            merge_cmd = [
                "ffmpeg", "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_list_path,
                "-c", "copy",
                output_path
            ]
            result = subprocess.run(merge_cmd, capture_output=True, text=True)
            
            if result.returncode != 0:
                logging.warning(f"Fission {fission_id}: Fast merge failed, trying re-encode")
                merge_cmd_reencode = [
                    "ffmpeg", "-y",
                    "-f", "concat",
                    "-safe", "0",
                    "-i", concat_list_path,
                    "-c:v", "libx264",
                    "-c:a", "aac",
                    "-preset", "fast",
                    output_path
                ]
                subprocess.run(merge_cmd_reencode, check=True, capture_output=True)
            
            if os.path.exists(output_path):
                final_result_url = f"/uploads/queue/{output_filename}"
                status["merged_video_url"] = final_result_url
                
                # Update thumbnail
                thumbnail_filename = f"story_fission_{fission_id}_thumb.jpg"
                thumbnail_path = f"/app/uploads/queue/{thumbnail_filename}"
                try:
                    thumb_cmd = [
                        "ffmpeg", "-y",
                        "-i", output_path,
                        "-ss", "00:00:00.500",
                        "-vframes", "1",
                        "-q:v", "2",
                        thumbnail_path
                    ]
                    subprocess.run(thumb_cmd, check=True, capture_output=True)
                    status["thumbnail_url"] = f"/uploads/queue/{thumbnail_filename}"
                except Exception:
                    pass
                
                # Update or create queue item
                db = SessionLocal()
                try:
                    # Find existing merged item
                    existing = db.query(VideoQueueItem).filter(
                        VideoQueueItem.filename == output_filename
                    ).first()
                    
                    if existing:
                        existing.result_url = final_result_url
                        existing.status = "done"
                        if status.get("thumbnail_url"):
                            existing.preview_url = status["thumbnail_url"]
                    else:
                        user_default_share = user.default_share if user.default_share is not None else True
                        merged_item = VideoQueueItem(
                            id=str(int(uuid.uuid4().int))[:18],
                            filename=output_filename,
                            file_path=output_path,
                            prompt=f"Story Fission {fission_id} (Remerged)",
                            status="done",
                            result_url=final_result_url,
                            user_id=user.id,
                            is_merged=True,
                            is_shared=user_default_share
                        )
                        if status.get("thumbnail_url"):
                            merged_item.preview_url = status["thumbnail_url"]
                        db.add(merged_item)
                    
                    db.commit()
                finally:
                    db.close()
                
                logging.info(f"Fission {fission_id}: Remerge completed: {final_result_url}")
            else:
                logging.error(f"Fission {fission_id}: Remerge failed - output file not created")
                status["error"] = "Remerge failed: output file not created"
                
        except Exception as e:
            logging.error(f"Fission {fission_id}: Remerge exception: {e}")
            status["error"] = f"Remerge failed: {str(e)}"
    
    background_tasks.add_task(remerge_task)
    
    return {"status": "remerging", "video_count": len(video_inputs)}


# =============================================================================
# WebSocket Endpoint for Real-time Updates
# =============================================================================

def verify_ws_token(token: str, db: Session) -> Optional[User]:
    """Verify JWT token for WebSocket connection."""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            return None
        user = db.query(User).filter(User.username == username).first()
        return user
    except JWTError:
        return None


@app.websocket("/ws/{token}")
async def websocket_endpoint(websocket: WebSocket, token: str):
    """
    WebSocket endpoint for real-time updates.
    
    Clients connect with their JWT token and receive:
    - Task progress updates
    - Queue status changes
    - System notifications
    """
    db = SessionLocal()
    try:
        user = verify_ws_token(token, db)
        if not user:
            await websocket.close(code=4001, reason="Invalid token")
            return
        
        # Connect the WebSocket
        await connection_manager.connect(
            websocket=websocket,
            user_id=user.id,
            username=user.username,
            is_admin=(user.role == "admin")
        )
        
        try:
            # Keep connection alive and handle incoming messages
            while True:
                # Wait for messages (ping/pong or commands)
                data = await websocket.receive_text()
                
                # Handle ping
                if data == "ping":
                    await websocket.send_text("pong")
                
                # Handle activity updates from client
                elif data.startswith("activity:"):
                    activity = data[9:]
                    await connection_manager.update_user_activity(user.id, activity)
                
        except WebSocketDisconnect:
            await connection_manager.disconnect(websocket, user.id)
    finally:
        db.close()


# =============================================================================
# Admin Monitoring API
# =============================================================================

@app.get("/api/v1/admin/live-status")
async def get_admin_live_status(
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Get real-time system status for admin dashboard.
    
    Returns:
    - Online users
    - Active tasks
    - Queue statistics
    - Recent activities
    """
    # Get online users from WebSocket manager
    online_users = connection_manager.get_online_users()
    
    # Get queue stats from Redis
    try:
        task_queue = await get_task_queue()
        queue_stats = await task_queue.get_queue_stats()
    except Exception as e:
        logger.warning(f"Failed to get queue stats: {e}")
        queue_stats = {"error": str(e)}
    
    # Get recent activities from database (last 12 hours only)
    twelve_hours_ago = get_china_now() - timedelta(hours=12)
    recent_activities = db.query(UserActivity).filter(
        UserActivity.created_at >= twelve_hours_ago
    ).order_by(
        UserActivity.created_at.desc()
    ).limit(50).all()
    
    # Build activities list with username lookup
    activities_list = []
    for a in recent_activities:
        # Get username for this activity
        user = db.query(User).filter(User.id == a.user_id).first()
        username = (user.nickname or user.username) if user else f"用户 {a.user_id}"
        
        activities_list.append({
            "id": a.id,
            "user_id": a.user_id,
            "username": username,  # Add username field
            "action": a.action,
            "details": a.details,
            "created_at": a.created_at.isoformat() if a.created_at else None
        })
    
    # Get current processing counts
    video_processing = db.query(VideoQueueItem).filter(
        VideoQueueItem.status == "processing"
    ).count()
    
    video_pending = db.query(VideoQueueItem).filter(
        VideoQueueItem.status == "pending"
    ).count()
    
    # Count active fission tasks
    active_fission_tasks = sum(
        1 for s in STORY_FISSION_STATUS.values() 
        if s.get("status") == "processing"
    )
    
    # Count active story chain tasks
    active_chain_tasks = sum(
        1 for s in STORY_CHAIN_STATUS.values() 
        if s.get("status") in ["processing", "merging"]
    )
    
    # Total active tasks
    total_active_tasks = video_processing + active_fission_tasks + active_chain_tasks
    
    return {
        "online_users": online_users,
        "online_count": len(online_users),
        "queue_stats": {
            "video_processing": video_processing,
            "video_pending": video_pending,
            "fission_active": active_fission_tasks,
            "chain_active": active_chain_tasks,
            "total_active": total_active_tasks,
            **queue_stats
        },
        "recent_activities": activities_list,
        "timestamp": datetime.now().isoformat()
    }


@app.delete("/api/v1/admin/activities")
async def clear_activities(
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Clear all activity logs (admin only)."""
    try:
        count = db.query(UserActivity).delete()
        db.commit()
        return {"status": "success", "deleted_count": count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to clear activities: {str(e)}")


@app.get("/api/v1/admin/user/{user_id}/tasks")
async def get_user_tasks_admin(
    user_id: int,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all tasks for a specific user (admin only)."""
    # Get tasks from Redis queue
    try:
        task_queue = await get_task_queue()
        redis_tasks = await task_queue.get_user_tasks(user_id)
    except Exception:
        redis_tasks = []
    
    # Get video queue items from database
    video_tasks = db.query(VideoQueueItem).filter(
        VideoQueueItem.user_id == user_id
    ).order_by(VideoQueueItem.created_at.desc()).limit(50).all()
    
    return {
        "user_id": user_id,
        "redis_tasks": redis_tasks,
        "video_tasks": [
            {
                "id": v.id,
                "filename": v.filename,
                "status": v.status,
                "created_at": v.created_at.isoformat() if v.created_at else None
            }
            for v in video_tasks
        ]
    }


@app.get("/api/v1/admin/activities")
async def get_all_activities(
    limit: int = 50,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all recent activities across all users."""
    activities = db.query(UserActivity).order_by(
        UserActivity.created_at.desc()
    ).limit(limit).all()
    
    # Enrich with user info
    user_ids = set(a.user_id for a in activities if a.user_id)
    users = db.query(User).filter(User.id.in_(user_ids)).all()
    user_map = {u.id: u.username for u in users}
    
    return [
        {
            "id": a.id,
            "user_id": a.user_id,
            "username": user_map.get(a.user_id, "Unknown"),
            "action": a.action,
            "details": a.details,
            "created_at": a.created_at.isoformat() if a.created_at else None
        }
        for a in activities
    ]


# Helper function to log user activity
async def log_user_activity(
    db: Session,
    user_id: int,
    action: str,
    details: str = None
):
    """Log a user activity for monitoring."""
    activity = UserActivity(
        user_id=user_id,
        action=action,
        details=details
    )
    db.add(activity)
    db.commit()
    
    # Also notify admins via WebSocket
    await connection_manager.broadcast_to_admins({
        "type": "user_activity",
        "data": {
            "user_id": user_id,
            "action": action,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
    })


# ========== Character Video Generation (Proxy) ==========

class CharacterVideoRequest(BaseModel):
    """角色视频生成请求模型"""
    video_base64: str  # Base64 编码的视频
    prompt: str = ""   # 动作提示词（可选，仅保存角色时可为空）


@app.post("/api/v1/character/generate")
async def generate_character_video(
    request: CharacterVideoRequest,
    user: User = Depends(get_current_user)
):
    """
    角色视频生成代理端点
    通过后端代理调用 sora2api，避免前端 CORS 问题
    使用 SSE 流式返回结果
    """
    from starlette.responses import StreamingResponse
    
    # 获取配置
    db = SessionLocal()
    try:
        video_api_url = db.query(SystemConfig).filter(SystemConfig.key == "video_api_url").first()
        video_api_key = db.query(SystemConfig).filter(SystemConfig.key == "video_api_key").first()
        video_model_name = db.query(SystemConfig).filter(SystemConfig.key == "video_model_name").first()
        
        api_url = video_api_url.value if video_api_url else os.getenv("VIDEO_API_URL", "")
        api_key = video_api_key.value if video_api_key else os.getenv("VIDEO_API_KEY", "")
        model_name = video_model_name.value if video_model_name else os.getenv("VIDEO_MODEL_NAME", "sora2-portrait-15s")
    finally:
        db.close()
    
    # 构建目标 URL
    target_url = api_url if api_url.endswith("/chat/completions") else f"{api_url.rstrip('/')}/chat/completions"
    
    # 构建请求内容
    content = [{"type": "video_url", "video_url": {"url": request.video_base64}}]
    if request.prompt.strip():
        content.append({"type": "text", "text": request.prompt})
    
    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": content}],
        "stream": True
    }
    
    async def stream_response():
        """流式返回 API 响应"""
        async with httpx.AsyncClient(timeout=900.0) as client:
            try:
                async with client.stream(
                    "POST",
                    target_url,
                    json=payload,
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {api_key}"
                    }
                ) as response:
                    if response.status_code != 200:
                        error_text = await response.aread()
                        yield f"data: {json.dumps({'error': f'API Error: {response.status_code}', 'detail': error_text.decode()})}\n\n"
                        return
                    
                    async for line in response.aiter_lines():
                        if line.strip():
                            yield f"{line}\n\n"
                            
            except Exception as e:
                logger.error(f"Character video generation error: {e}")
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
    
    return StreamingResponse(
        stream_response(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


# ============================================
# Keyword Extraction API Endpoints
# ============================================

class KeywordAnalyzeRequest(BaseModel):
    title: str
    prompt: Optional[str] = None

class KeywordAnalyzeResponse(BaseModel):
    translation: str
    keywords: str

class KeywordHistorySaveRequest(BaseModel):
    titles: List[dict]

class KeywordBatchRequest(BaseModel):
    titles: List[str]
    batch_size: int = 5  # Process in batches to avoid token limits

# In-memory storage for keyword history (could be moved to database)
KEYWORD_HISTORY = []

@app.post("/api/v1/keywords/analyze-single", response_model=KeywordAnalyzeResponse)
async def analyze_single_keyword(
    request: KeywordAnalyzeRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Analyze a single title: translate to Chinese and extract root keywords."""
    
    # Get config from database
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set. Please configure in Settings.")
    
    # Build prompt
    system_prompt = """You are a professional e-commerce title analyst. Analyze the given product title and return a JSON response.

RULES:
1. Translate the title to fluent Chinese
2. Extract 4 root keywords in the ORIGINAL language of the title (not Chinese)
3. Keywords should be the most important product descriptors
4. Separate keywords with commas

Return ONLY valid JSON, no additional text:
{"translation": "中文翻译", "keywords": "Keyword1, Keyword2, Keyword3, Keyword4"}"""

    user_prompt = f"Analyze this product title:\n\n{request.title}"
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.3,
        "max_tokens": 1024
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    target_url = api_url
    if not target_url.endswith("/chat/completions"):
        target_url = f"{target_url.rstrip('/')}/chat/completions"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(target_url, json=payload, headers=headers)
            
            if response.status_code != 200:
                error_text = response.text
                logger.error(f"Keyword analysis API error: {response.status_code} - {error_text[:200]}")
                raise HTTPException(status_code=500, detail=f"API error: {response.status_code}")
            
            data = response.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            # Parse JSON response
            try:
                # Clean potential markdown code blocks
                content = content.strip()
                if content.startswith("```json"):
                    content = content[7:]
                if content.startswith("```"):
                    content = content[3:]
                if content.endswith("```"):
                    content = content[:-3]
                content = content.strip()
                
                result = json.loads(content)
                translation = result.get("translation", "")
                keywords = result.get("keywords", "")
                
                await auto_sync_to_feishu(db, request.title, translation, keywords)
                
                return KeywordAnalyzeResponse(
                    translation=translation,
                    keywords=keywords
                )
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse keyword response: {content[:200]}")
                # Try to extract manually
                return KeywordAnalyzeResponse(
                    translation=content[:200] if content else "解析失败",
                    keywords=""
                )
                
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="API request timeout")
    except Exception as e:
        logger.error(f"Keyword analysis error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v1/keywords/history")
async def save_keyword_history(
    request: KeywordHistorySaveRequest,
    token: str = Depends(verify_token)
):
    """Save keyword extraction history."""
    record = {
        "created_at": datetime.now().isoformat(),
        "count": len(request.titles),
        "titles": request.titles
    }
    KEYWORD_HISTORY.insert(0, record)
    # Keep only last 50 records
    if len(KEYWORD_HISTORY) > 50:
        KEYWORD_HISTORY.pop()
    return {"success": True}


@app.get("/api/v1/keywords/history")
async def get_keyword_history(token: str = Depends(verify_token)):
    return {"records": KEYWORD_HISTORY}


@app.delete("/api/v1/keywords/history/{index}")
async def delete_keyword_history(index: int, token: str = Depends(verify_token)):
    if 0 <= index < len(KEYWORD_HISTORY):
        KEYWORD_HISTORY.pop(index)
        return {"success": True}
    raise HTTPException(status_code=404, detail="History record not found")


@app.delete("/api/v1/keywords/history")
async def clear_keyword_history(token: str = Depends(verify_token)):
    KEYWORD_HISTORY.clear()
    return {"success": True}


@app.post("/api/v1/keywords/export-excel")
async def export_keywords_excel(
    request: KeywordHistorySaveRequest,
    token: str = Depends(verify_token)
):
    """Export keyword results to Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "核心大词分析"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["标题", "标题_中文翻译", "标题_核心大词(Root Keywords)", "处理状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, title_data in enumerate(request.titles, 2):
            ws.cell(row=row_idx, column=1, value=title_data.get("original", ""))
            ws.cell(row=row_idx, column=2, value=title_data.get("translation", ""))
            ws.cell(row=row_idx, column=3, value=title_data.get("keywords", ""))
            
            status = title_data.get("status", "pending")
            status_text = "已完成" if status == "completed" else "失败" if status == "failed" else "未处理"
            ws.cell(row=row_idx, column=4, value=status_text)
            
            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"keywords_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename_cn = f"核心词提取_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from urllib.parse import quote
        encoded_filename = quote(filename_cn)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_filename}"
            }
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Please install: pip install openpyxl")
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class FeishuSyncRequest(BaseModel):
    titles: List[dict]

class FeishuSyncResponse(BaseModel):
    success: bool
    synced_count: int
    message: str

async def get_feishu_tenant_token(app_id: str, app_secret: str) -> str:
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(url, json={"app_id": app_id, "app_secret": app_secret})
        data = response.json()
        if data.get("code") == 0:
            return data["tenant_access_token"]
        raise Exception(f"Feishu auth failed: {data.get('msg', 'Unknown error')}")

async def auto_sync_to_feishu(db: Session, original: str, translation: str, keywords: str):
    import time
    try:
        db_config_list = db.query(SystemConfig).all()
        config_dict = {item.key: item.value for item in db_config_list}
        
        app_id = config_dict.get("feishu_app_id")
        app_secret = config_dict.get("feishu_app_secret")
        app_token = config_dict.get("feishu_app_token")
        table_id = config_dict.get("feishu_table_id")
        
        if not all([app_id, app_secret, app_token, table_id]):
            return
        
        feishu_token = await get_feishu_tenant_token(app_id, app_secret)
        beijing_timestamp = int(time.time() * 1000)
        
        record = {
            "fields": {
                "标题": original,
                "中文翻译": translation,
                "核心大词": keywords,
                "同步时间": beijing_timestamp
            }
        }
        
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create"
        headers = {
            "Authorization": f"Bearer {feishu_token}",
            "Content-Type": "application/json"
        }
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            await client.post(url, headers=headers, json={"records": [record]})
    except Exception as e:
        logger.warning(f"Auto sync to Feishu failed: {str(e)}")

@app.post("/api/v1/keywords/sync-feishu", response_model=FeishuSyncResponse)
async def sync_keywords_to_feishu(
    request: FeishuSyncRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    db_config_list = db.query(SystemConfig).all()
    config_dict = {item.key: item.value for item in db_config_list}
    
    app_id = config_dict.get("feishu_app_id")
    app_secret = config_dict.get("feishu_app_secret")
    app_token = config_dict.get("feishu_app_token")
    table_id = config_dict.get("feishu_table_id")
    
    if not all([app_id, app_secret, app_token, table_id]):
        raise HTTPException(status_code=400, detail="飞书配置不完整，请在系统设置中配置飞书 App ID、App Secret、App Token 和 Table ID")
    
    try:
        feishu_token = await get_feishu_tenant_token(app_id, app_secret)
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"飞书认证失败: {str(e)}")
    
    completed_titles = [t for t in request.titles if t.get("status") == "completed"]
    if not completed_titles:
        raise HTTPException(status_code=400, detail="没有已完成的记录可以同步")
    
    import time
    beijing_timestamp = int(time.time() * 1000)
    
    records = []
    for title in completed_titles:
        records.append({
            "fields": {
                "标题": title.get("original", ""),
                "中文翻译": title.get("translation", ""),
                "核心大词": title.get("keywords", ""),
                "同步时间": beijing_timestamp
            }
        })
    
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create"
    headers = {
        "Authorization": f"Bearer {feishu_token}",
        "Content-Type": "application/json"
    }
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(url, headers=headers, json={"records": records})
            data = response.json()
            
            if data.get("code") == 0:
                return FeishuSyncResponse(
                    success=True,
                    synced_count=len(completed_titles),
                    message=f"成功同步 {len(completed_titles)} 条记录到飞书多维表格"
                )
            else:
                error_msg = data.get("msg", "Unknown error")
                logger.error(f"Feishu sync failed: {data}")
                raise HTTPException(status_code=500, detail=f"飞书同步失败: {error_msg}")
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="飞书 API 请求超时")
    except Exception as e:
        logger.error(f"Feishu sync error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Mexico Beauty Station ---
# Helper function to load Mexico Beauty prompts
def load_mexico_beauty_prompt(module_name: str) -> str:
    """Load system prompt for Mexico Beauty module from text file."""
    import os
    prompt_file = os.path.join(os.path.dirname(__file__), 'prompts', f'mexico_beauty_{module_name}.txt')
    try:
        with open(prompt_file, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        logger.error(f"Mexico Beauty prompt file not found: {prompt_file}")
        return ""

async def call_chat_completion_api(api_url: str, api_key: str, payload: dict, max_retries: int = 3) -> str:
    """Generic chat completion API call for Mexico Beauty endpoints."""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    target_url = api_url
    if not target_url.endswith("/chat/completions"):
        target_url = f"{target_url.rstrip('/')}/chat/completions"
    
    last_error = None
    for attempt in range(max_retries):
        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                response = await client.post(target_url, json=payload, headers=headers)
                
                if response.status_code == 503 or response.status_code == 429:
                    if attempt < max_retries - 1:
                        wait_time = 2 * (attempt + 1)
                        logger.warning(f"API returned {response.status_code}, retrying in {wait_time}s... (attempt {attempt+1}/{max_retries})")
                        await asyncio.sleep(wait_time)
                        continue
                
                if response.status_code != 200:
                    error_text = response.text
                    logger.error(f"Chat completion API error: {response.status_code} - {error_text[:200]}")
                    raise HTTPException(status_code=500, detail=f"API error: {response.status_code}")
                
                data = response.json()
                content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
                return content
        except HTTPException:
            raise
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                logger.warning(f"API call failed: {e}, retrying... (attempt {attempt+1}/{max_retries})")
                await asyncio.sleep(2)
                continue
            raise HTTPException(status_code=500, detail=f"API call failed after {max_retries} attempts: {str(e)}")

async def call_content_review_api(prompt_text: str, db: Session) -> dict:
    """
    调用内容审核 API 检查图片生成提示词是否包含违规内容。
    
    Args:
        prompt_text: 待审核的提示词文本
        db: 数据库会话
        
    Returns:
        dict: {"passed": bool, "reviewed_prompt": str, "reason": str, "is_modified": bool}
    """
    # 默认返回结果（通过，无修改）
    default_result = {
        "passed": True,
        "reviewed_prompt": prompt_text,
        "reason": "审核未启用",
        "is_modified": False
    }
    
    # 读取审核配置
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    
    # 检查是否启用审核
    review_enabled = config_dict.get("content_review_enabled", "false")
    if review_enabled.lower() != "true":
        logger.debug("Content review is disabled, skipping review")
        return default_result
    
    # 获取 API 配置
    api_url = config_dict.get("content_review_api_url", "")
    api_key = config_dict.get("content_review_api_key", "")
    model_name = config_dict.get("content_review_model", "content-review")
    
    if not api_key:
        logger.warning("Content review API key not configured, skipping review")
        default_result["reason"] = "审核 API 密钥未配置"
        return default_result
    
    # 构建请求
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    target_url = api_url
    if not target_url.endswith("/chat/completions"):
        target_url = f"{target_url.rstrip('/')}/chat/completions"
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": "你是一个内容审核助手。检查以下图片生成提示词是否包含违规内容（色情、暴力、政治敏感等）。如果安全，返回JSON: {\"passed\": true, \"reason\": \"内容安全\"}。如果不安全，返回JSON: {\"passed\": false, \"reason\": \"违规原因\", \"safe_version\": \"修改后的安全版本\"}"},
            {"role": "user", "content": f"请审核以下提示词:\n\n{prompt_text}"}
        ],
        "temperature": 0.3,
        "max_tokens": 1024
    }
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(target_url, json=payload, headers=headers)
            
            if response.status_code != 200:
                error_text = response.text
                logger.warning(f"Content review API error: {response.status_code} - {error_text[:200]}, allowing content by default")
                default_result["reason"] = "审核 API 调用失败，默认通过"
                return default_result
            
            data = response.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            # 尝试解析 JSON 响应
            try:
                clean_content = content.strip()
                if clean_content.startswith("```"):
                    lines = clean_content.split("\n")
                    clean_content = "\n".join(lines[1:-1]) if len(lines) > 2 else clean_content
                if clean_content.endswith("```"):
                    clean_content = clean_content[:-3].strip()
                
                review_result = json.loads(clean_content)
                
                passed = review_result.get("passed", review_result.get("pass", True))
                reason = review_result.get("reason", "未知原因")
                safe_version = review_result.get("safe_version", "")
                
                if passed:
                    return {
                        "passed": True,
                        "reviewed_prompt": prompt_text,
                        "reason": reason,
                        "is_modified": False
                    }
                else:
                    if safe_version:
                        logger.info(f"Content review modified prompt: {reason}")
                        return {
                            "passed": True,
                            "reviewed_prompt": safe_version,
                            "reason": f"原内容违规({reason})，已自动修改",
                            "is_modified": True
                        }
                    else:
                        logger.warning(f"Content review rejected: {reason}")
                        return {
                            "passed": False,
                            "reviewed_prompt": prompt_text,
                            "reason": reason,
                            "is_modified": False
                        }
                        
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse content review response: {e}, content: {content[:200]}")
                default_result["reason"] = "审核响应解析失败，默认通过"
                return default_result
                
    except Exception as e:
        logger.warning(f"Content review API call failed: {e}, allowing content by default")
        default_result["reason"] = f"审核 API 调用异常: {str(e)}"
        return default_result

# Mexico Beauty request/response models
class MexicoBeautyKeywordRequest(BaseModel):
    title: str

class MexicoBeautyTitleRequest(BaseModel):
    title: str
    image_base64: Optional[str] = None

class MexicoBeautyImageRequest(BaseModel):
    image_base64: str

class MexicoBeautyDescriptionRequest(BaseModel):
    title: str
    image_base64: str


# Mexico Beauty API Endpoints
@app.post("/api/v1/mexico-beauty/keyword-analysis-single")
async def mexico_keyword_analysis_single(
    request: MexicoBeautyKeywordRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Analyze competitor title to extract root keywords and attributes."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    system_prompt = load_mexico_beauty_prompt("keyword")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": request.title}
        ],
        "temperature": 0.7,
        "max_tokens": 2048
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        return {"result": result_text}
    except Exception as e:
        logger.error(f"Mexico keyword analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/mexico-beauty/title-optimization-single")
async def mexico_title_optimization_single(
    title: str = Form(...),
    image: UploadFile = File(None),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Optimize product title for TikTok Shop Mexico SEO."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    system_prompt = load_mexico_beauty_prompt("title")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    messages = [{"role": "system", "content": system_prompt}]
    
    if image:
        image_b64 = await file_to_base64_compressed(image, max_size=800, quality=75)
        messages.append({
            "role": "user",
            "content": [
                {"type": "text", "text": f"Competitor Title: {title}"},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
            ]
        })
    else:
        messages.append({"role": "user", "content": f"Competitor Title: {title}"})
    
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": 0.7,
        "max_tokens": 1024
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        return {"result": result_text}
    except Exception as e:
        logger.error(f"Mexico title optimization failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/mexico-beauty/image-prompt-single")
async def mexico_image_prompt_single(
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Generate image prompts and marketing copy from reference image."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    system_prompt = load_mexico_beauty_prompt("image")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    image_b64 = await file_to_base64_compressed(image, max_size=800, quality=75)
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Please analyze this product image and generate visual prompts and marketing copy."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            }
        ],
        "temperature": 0.8,
        "max_tokens": 2048
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        return {"result": result_text}
    except Exception as e:
        logger.error(f"Mexico image prompt generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/mexico-beauty/description-single")
async def mexico_description_single(
    title: str = Form(...),
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Generate product description (Modo de Uso) for TikTok Shop."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    system_prompt = load_mexico_beauty_prompt("description")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    image_b64 = await file_to_base64_compressed(image, max_size=800, quality=75)
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": f"Product Title: {title}\n\nGenerate usage instructions (Modo de Uso)."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            }
        ],
        "temperature": 0.6,
        "max_tokens": 2048
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        return {"result": result_text}
    except Exception as e:
        logger.error(f"Mexico description generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Mexico Beauty: Image Prompts Generation (10 prompts) ---
class ImagePromptItem(BaseModel):
    id: int
    type: str  # "Main Image" or "Detail/Scenario"
    title: str
    promptText: str
    rationale: str
    review_status: Optional[str] = None  # "passed", "modified", "failed", or None

class ImagePromptsRequest(BaseModel):
    title: Optional[str] = None
    keywords: Optional[str] = None
    description: Optional[str] = None

class ImagePromptsResponse(BaseModel):
    prompts: List[ImagePromptItem]

@app.post("/api/v1/mexico-beauty/image-prompts-batch", response_model=ImagePromptsResponse)
async def mexico_image_prompts_batch(
    title: str = Form(None),
    keywords: str = Form(None),
    description: str = Form(None),
    aspect_ratio: str = Form("1:1"),
    target_language: str = Form("es-MX"),
    image: UploadFile = File(...),
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Generate 10 image prompts (2 Main + 8 Detail) based on product info."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    # Language/Region configuration mapping
    LANGUAGE_CONFIG = {
        "es-MX": {"region": "Mexico", "language": "Mexican Spanish", "locale_examples": '"Súper práctico," "No gasta luz," "Material resistente"'},
        "th-TH": {"region": "Thailand", "language": "Thai", "locale_examples": '"สะดวกสุดๆ," "ประหยัดไฟ," "วัสดุทนทาน"'},
        "zh-CN": {"region": "China", "language": "Simplified Chinese", "locale_examples": '"超级实用," "省电," "材质耐用"'},
        "en-US": {"region": "United States", "language": "American English", "locale_examples": '"Super practical," "Energy-saving," "Durable material"'},
        "id-ID": {"region": "Indonesia", "language": "Indonesian", "locale_examples": '"Sangat praktis," "Hemat listrik," "Bahan tahan lama"'},
        "vi-VN": {"region": "Vietnam", "language": "Vietnamese", "locale_examples": '"Siêu tiện lợi," "Tiết kiệm điện," "Chất liệu bền"'},
        "ms-MY": {"region": "Malaysia", "language": "Malay", "locale_examples": '"Sangat praktikal," "Jimat elektrik," "Bahan tahan lasak"'},
        "tl-PH": {"region": "Philippines", "language": "Filipino/Tagalog", "locale_examples": '"Sobrang praktiko," "Tipid sa kuryente," "Matibay na materyales"'},
    }
    
    lang_config = LANGUAGE_CONFIG.get(target_language, LANGUAGE_CONFIG["es-MX"])
    
    system_prompt = load_mexico_beauty_prompt("image_prompts")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    # Dynamic replacement of region/language in the prompt
    system_prompt = system_prompt.replace("Mexico", lang_config["region"])
    system_prompt = system_prompt.replace("Mexican Spanish", lang_config["language"])
    system_prompt = system_prompt.replace("TikTok Mexico", f"TikTok {lang_config['region']}")
    system_prompt = system_prompt.replace(
        '"Súper práctico," "No gasta luz," "Material resistente"',
        lang_config["locale_examples"]
    )
    
    image_b64 = await file_to_base64_compressed(image, max_size=800, quality=75)
    
    user_message = f"""
Product Title: {title or "Not provided"}
Keywords: {keywords or "Not provided"}
Context/Description: {description or "Not provided"}

*** CRITICAL REQUIREMENT - TARGET LANGUAGE/REGION: {lang_config["language"]} ({lang_config["region"]}) ***
All text overlays, headlines, feature points, and CTAs MUST be written in {lang_config["language"]}.
Use localized e-commerce copy appropriate for {lang_config["region"]} market.

*** CRITICAL REQUIREMENT - ASPECT RATIO: {aspect_ratio} ***
You MUST use "{aspect_ratio}" as the aspect ratio in EVERY prompt's [LAYOUT & COMPOSITION] section.
- If aspect ratio is "1:1": Use "Square 1:1 format" in all prompts
- If aspect ratio is "9:16": Use "Vertical 9:16 portrait" in all prompts  
- If aspect ratio is "16:9": Use "Horizontal 16:9 landscape" in all prompts
DO NOT use any other aspect ratio. This is mandatory.

The user has provided an image of the product. Analyze the image to understand the product's features.

Generate 10 prompts in JSON format following the strict structure defined above.
"""
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_message},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}}
                ]
            }
        ],
        "temperature": 0.8,
        "max_tokens": 8192
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        
        json_match = re.search(r'\[[\s\S]*\]', result_text)
        if json_match:
            prompts_data = json.loads(json_match.group())
        else:
            prompts_data = json.loads(result_text)
        
        prompts = []
        for i, p in enumerate(prompts_data[:10]):
            original_prompt_text = p.get("promptText", "")
            
            prompts.append(ImagePromptItem(
                id=p.get("id", i + 1),
                type=p.get("type", "Main Image" if i < 2 else "Detail/Scenario"),
                title=p.get("title", f"Prompt {i + 1}"),
                promptText=original_prompt_text,
                rationale=p.get("rationale", ""),
                review_status=None
            ))
        
        return ImagePromptsResponse(prompts=prompts)
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse JSON from AI response: {str(e)}")
        logger.error(f"Raw response: {result_text[:500]}")
        raise HTTPException(status_code=500, detail="Failed to parse AI response as JSON")
    except Exception as e:
        logger.error(f"Mexico image prompts generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Mexico Beauty: Refine Prompt ---
class RefinePromptRequest(BaseModel):
    original_prompt: ImagePromptItem
    feedback: str
    product_title: Optional[str] = None
    product_description: Optional[str] = None

@app.post("/api/v1/mexico-beauty/refine-prompt", response_model=ImagePromptItem)
async def mexico_refine_prompt(
    request: RefinePromptRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Refine a specific prompt based on user feedback."""
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("analysis_model_name", "gemini-3-pro-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    system_prompt = load_mexico_beauty_prompt("refine_prompt")
    if not system_prompt:
        raise HTTPException(status_code=500, detail="System prompt not loaded")
    
    user_message = f"""
PRODUCT INFO:
Title: {request.product_title or "Not provided"}
Context: {request.product_description or "Not provided"}

ORIGINAL PROMPT DATA:
ID: {request.original_prompt.id}
Type: {request.original_prompt.type}
Current Title: {request.original_prompt.title}
Current Prompt Text: {request.original_prompt.promptText}

USER FEEDBACK (PROBLEM TO FIX):
"{request.feedback}"

INSTRUCTION:
Rewrite the 'promptText', 'title', and 'rationale' to strictly address the User Feedback.
- If they complain about text, fix the [TEXT OVERLAYS].
- If they complain about background, fix the [VISUAL CORE].
- Ensure NO MINORS are mentioned.
- Keep the output strict JSON.
"""
    
    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message}
        ],
        "temperature": 0.7,
        "max_tokens": 2048
    }
    
    try:
        result_text = await call_chat_completion_api(api_url, api_key, payload)
        
        json_match = re.search(r'\{[\s\S]*\}', result_text)
        if json_match:
            refined_data = json.loads(json_match.group())
        else:
            refined_data = json.loads(result_text)
        
        return ImagePromptItem(
            id=refined_data.get("id", request.original_prompt.id),
            type=refined_data.get("type", request.original_prompt.type),
            title=refined_data.get("title", request.original_prompt.title),
            promptText=refined_data.get("promptText", request.original_prompt.promptText),
            rationale=refined_data.get("rationale", "Refined based on user feedback")
        )
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse JSON from AI response: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to parse AI response as JSON")
    except Exception as e:
        logger.error(f"Mexico refine prompt failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Mexico Beauty: Generate Image from Prompt ---
class MexicoGenerateImageRequest(BaseModel):
    prompt_text: str
    aspect_ratio: str = "1:1"

@app.post("/api/v1/mexico-beauty/generate-image")
async def mexico_generate_image(
    prompt_text: str = Form(...),
    aspect_ratio: str = Form("1:1"),
    reference_image: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    await throttle_request()
    
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    api_url = config_dict.get("api_url")
    api_key = config_dict.get("api_key")
    model_name = config_dict.get("model_name", "gemini-3-pro-image-preview")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=400, detail="API configuration not set")
    
    # 调用内容审核（生图前二次审核）
    review_result = await call_content_review_api(prompt_text, db)
    if not review_result["passed"] and not review_result["is_modified"]:
        raise HTTPException(status_code=400, detail=f"内容审核未通过: {review_result['reason']}")
    if review_result["is_modified"]:
        prompt_text = review_result["reviewed_prompt"]
        logger.info(f"Mexico generate-image: Prompt modified by content review")
    
    image_b64 = await file_to_base64_compressed(reference_image, max_size=800, quality=75)
    
    final_prompt = f"{prompt_text}\n\nFINAL RENDER INSTRUCTIONS:\n1. Ensure the reference product is the star.\n2. STRICTLY follow the text rendering instructions for Spanish overlays.\n3. Use {aspect_ratio} aspect ratio.\n4. DO NOT include realistic people/faces to ensure safety compliance."
    
    # Map aspect ratio to size
    size_map = {
        "1:1": "1024x1024",
        "9:16": "768x1344",
        "16:9": "1344x768",
        "4:5": "896x1120",
        "3:4": "896x1152"
    }
    image_size = size_map.get(aspect_ratio, "1024x1024")
    
    try:
        async with httpx.AsyncClient() as client:
            result = None
            
            # Check if model is imagen-type (use /images/generations endpoint)
            if "imagen" in model_name.lower():
                logger.info(f"Using /images/generations endpoint for model: {model_name}")
                target_url = api_url.rstrip("/")
                if target_url.endswith("/v1"):
                    target_url = f"{target_url}/images/generations"
                elif not target_url.endswith("/images/generations"):
                    target_url = f"{target_url}/images/generations"
                
                headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
                timeout = httpx.Timeout(connect=15.0, read=300.0, write=30.0, pool=30.0)
                
                payload = {
                    "model": model_name,
                    "prompt": final_prompt,
                    "n": 1,
                    "size": image_size
                }
                
                response = await client.post(target_url, json=payload, headers=headers, timeout=timeout)
                
                if response.status_code != 200:
                    error_msg = response.text[:500]
                    logger.error(f"Imagen API error: {response.status_code} - {error_msg}")
                    raise HTTPException(status_code=500, detail=f"Image generation failed: {error_msg}")
                
                data = response.json()
                if data.get("data") and len(data["data"]) > 0:
                    item = data["data"][0]
                    if "b64_json" in item:
                        result = ImageResult(angle_name="Generated", image_base64=item["b64_json"])
                    elif "url" in item:
                        result = ImageResult(angle_name="Generated", image_url=item["url"])
                    else:
                        raise HTTPException(status_code=500, detail="Unknown image format in response")
                else:
                    raise HTTPException(status_code=500, detail="No image data returned from API")
            else:
                result = await call_multi_image_gen(
                    client, api_url, api_key,
                    [image_b64],
                    final_prompt,
                    0,
                    model_name,
                    aspect_ratio
                )
            
            if result is None:
                raise HTTPException(status_code=500, detail="No result from image generation")
            
            if result.error:
                raise HTTPException(status_code=500, detail=result.error)
            
            image_url = None
            saved_url = None
            
            if result.image_base64:
                image_url = f"data:image/png;base64,{result.image_base64}"
                
                try:
                    gallery_dir = "/app/uploads/gallery"
                    os.makedirs(gallery_dir, exist_ok=True)
                    
                    b64 = result.image_base64
                    if b64.startswith("data:"):
                        b64 = b64.split(",", 1)[1]
                    img_data = base64.b64decode(b64)
                    
                    timestamp = int(time.time() * 1000)
                    filename = f"mexico_product_{user.id}_{timestamp}.png"
                    file_path = os.path.join(gallery_dir, filename)
                    
                    with open(file_path, "wb") as f:
                        f.write(img_data)
                    
                    img_width, img_height = 1024, 1024
                    try:
                        from PIL import Image
                        with Image.open(file_path) as img:
                            img_width, img_height = img.size
                    except:
                        pass
                    
                    new_image = SavedImage(
                        user_id=user.id,
                        filename=filename,
                        file_path=file_path,
                        url=f"/uploads/gallery/{filename}",
                        prompt=prompt_text[:500],
                        width=img_width,
                        height=img_height,
                        category="mexico_product",
                        is_shared=user.default_share if user.default_share is not None else True
                    )
                    db.add(new_image)
                    db.commit()
                    saved_url = f"/uploads/gallery/{filename}"
                    logger.info(f"Mexico product image saved to gallery: {filename}")
                except Exception as save_err:
                    logger.error(f"Failed to save mexico product image to gallery: {save_err}")
                
            elif result.image_url:
                image_url = result.image_url
                
                try:
                    gallery_dir = "/app/uploads/gallery"
                    os.makedirs(gallery_dir, exist_ok=True)
                    
                    async with httpx.AsyncClient(timeout=60) as dl_client:
                        img_response = await dl_client.get(result.image_url)
                        if img_response.status_code == 200:
                            img_data = img_response.content
                            
                            timestamp = int(time.time() * 1000)
                            filename = f"mexico_product_{user.id}_{timestamp}.png"
                            file_path = os.path.join(gallery_dir, filename)
                            
                            with open(file_path, "wb") as f:
                                f.write(img_data)
                            
                            img_width, img_height = 1024, 1024
                            try:
                                from PIL import Image
                                with Image.open(file_path) as img:
                                    img_width, img_height = img.size
                            except:
                                pass
                            
                            new_image = SavedImage(
                                user_id=user.id,
                                filename=filename,
                                file_path=file_path,
                                url=f"/uploads/gallery/{filename}",
                                prompt=prompt_text[:500],
                                width=img_width,
                                height=img_height,
                                category="mexico_product",
                                is_shared=user.default_share if user.default_share is not None else True
                            )
                            db.add(new_image)
                            db.commit()
                            saved_url = f"/uploads/gallery/{filename}"
                            image_url = f"data:image/png;base64,{base64.b64encode(img_data).decode()}"
                            logger.info(f"Mexico product image downloaded and saved: {filename}")
                except Exception as dl_err:
                    logger.error(f"Failed to download and save mexico product image: {dl_err}")
            else:
                raise HTTPException(status_code=500, detail="No image generated")
            
            return {
                "image_url": image_url,
                "saved_url": saved_url
            }
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Mexico generate image failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Mexico Beauty Feishu Sync Models
class MexicoBeautyFeishuSyncRequest(BaseModel):
    module: str
    results: List[dict]

class MexicoBeautyFeishuSyncResponse(BaseModel):
    success: bool
    synced_count: int
    failed_count: int = 0
    message: str

@app.post("/api/v1/mexico-beauty/sync-feishu", response_model=MexicoBeautyFeishuSyncResponse)
async def mexico_beauty_sync_feishu(
    request: MexicoBeautyFeishuSyncRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    """Sync Mexico Beauty results to Feishu spreadsheet with chunking."""
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    
    feishu_app_id = config_dict.get("feishu_app_id")
    feishu_app_secret = config_dict.get("feishu_app_secret")
    feishu_app_token = config_dict.get("feishu_app_token")
    feishu_table_id = config_dict.get("feishu_table_id")
    
    if not all([feishu_app_id, feishu_app_secret, feishu_app_token, feishu_table_id]):
        raise HTTPException(status_code=400, detail="飞书配置未设置，请在系统设置中配置")
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            token_response = await client.post(
                "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
                json={"app_id": feishu_app_id, "app_secret": feishu_app_secret}
            )
            token_data = token_response.json()
            
            if token_data.get("code") != 0:
                raise HTTPException(status_code=500, detail=f"飞书Token获取失败: {token_data.get('msg')}")
            
            tenant_token = token_data["tenant_access_token"]
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="飞书Token请求超时")
    except Exception as e:
        logger.error(f"Feishu token error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    module_headers = {
        "keyword": ["标题", "中文翻译", "核心词", "同步时间"],
        "core_keyword": ["标题", "中文翻译", "核心词", "同步时间"],
        "title": ["原标题", "优化标题1", "优化标题2", "优化标题3", "同步时间"],
        "image": ["图片名称", "Visual Prompt", "Marketing Copy", "同步时间"],
        "description": ["标题", "产品描述", "使用说明", "同步时间"]
    }
    
    headers = module_headers.get(request.module, ["输入", "输出", "同步时间"])
    
    records = []
    for item in request.results:
        timestamp = datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S")
        
        if request.module == "keyword" or request.module == "core_keyword":
            records.append([
                item.get("input", ""),
                item.get("translation", ""),
                item.get("keywords", ""),
                timestamp
            ])
        elif request.module == "title":
            output_lines = item.get("output", "").split("\n")
            titles = [line for line in output_lines if line.strip()][:3]
            records.append([
                item.get("input", ""),
                titles[0] if len(titles) > 0 else "",
                titles[1] if len(titles) > 1 else "",
                titles[2] if len(titles) > 2 else "",
                timestamp
            ])
        else:
            records.append([
                item.get("input", ""),
                item.get("output", "")[:1000],
                "",
                timestamp
            ])
    
    CHUNK_SIZE = 300
    synced_count = 0
    failed_count = 0
    
    for i in range(0, len(records), CHUNK_SIZE):
        chunk = records[i:i + CHUNK_SIZE]
        
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{feishu_app_token}/tables/{feishu_table_id}/records/batch_create"
        headers_req = {"Authorization": f"Bearer {tenant_token}"}
        
        feishu_records = []
        for record in chunk:
            fields = {}
            for idx, value in enumerate(record):
                if idx < len(headers):
                    fields[headers[idx]] = value
            feishu_records.append({"fields": fields})
        
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            for chunk in chunks:
                response = await client.post(
                    url,
                    headers=headers_req,
                    json={"records": feishu_records}
                )
                data = response.json()
                
                if data.get("code") == 0:
                    synced_count += len(chunk)
                else:
                    failed_count += len(chunk)
                    logger.error(f"Feishu sync chunk failed: {data}")
                
                await asyncio.sleep(0.5)
    except Exception as e:
        failed_count += len(chunk)
        logger.error(f"Feishu sync chunk error: {str(e)}")
    
    if failed_count > 0:
        return MexicoBeautyFeishuSyncResponse(
            success=False,
            synced_count=synced_count,
            failed_count=failed_count,
            message=f"部分同步失败：成功 {synced_count} 条，失败 {failed_count} 条"
        )
    else:
        return MexicoBeautyFeishuSyncResponse(
            success=True,
            synced_count=synced_count,
            failed_count=0,
            message=f"成功同步 {synced_count} 条记录到飞书多维表格"
        )

class ProductDescriptionFeishuSyncRequest(BaseModel):
    product_title: str
    prompts: List[dict]

@app.post("/api/v1/mexico-beauty/sync-description-feishu", response_model=MexicoBeautyFeishuSyncResponse)
async def sync_description_to_feishu(
    request: ProductDescriptionFeishuSyncRequest,
    db: Session = Depends(get_db),
    token: str = Depends(verify_token)
):
    config_dict = {item.key: item.value for item in db.query(SystemConfig).all()}
    
    feishu_app_id = config_dict.get("feishu_app_id")
    feishu_app_secret = config_dict.get("feishu_app_secret")
    desc_app_token = config_dict.get("feishu_description_app_token")
    desc_table_id = config_dict.get("feishu_description_table_id")
    
    if not all([feishu_app_id, feishu_app_secret, desc_app_token, desc_table_id]):
        raise HTTPException(status_code=400, detail="产品描述飞书配置未设置，请在系统设置中配置")
    
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            token_response = await client.post(
                "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
                json={"app_id": feishu_app_id, "app_secret": feishu_app_secret}
            )
            token_data = token_response.json()
            
            if token_data.get("code") != 0:
                raise HTTPException(status_code=500, detail=f"飞书Token获取失败: {token_data.get('msg')}")
            
            tenant_token = token_data["tenant_access_token"]
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="飞书Token请求超时")
    except Exception as e:
        logger.error(f"Feishu token error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    timestamp = datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M:%S")
    
    feishu_records = []
    for prompt in request.prompts:
        feishu_records.append({
            "fields": {
                "产品标题": request.product_title or "未命名",
                "序号": prompt.get("id", 0),
                "类型": prompt.get("type", ""),
                "策略标题": prompt.get("title", ""),
                "策略思路": prompt.get("rationale", ""),
                "提示词": prompt.get("promptText", "")[:2000],
                "时间": timestamp
            }
        })
    
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{desc_app_token}/tables/{desc_table_id}/records/batch_create"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                url,
                headers={"Authorization": f"Bearer {tenant_token}"},
                json={"records": feishu_records}
            )
            data = response.json()
            
            if data.get("code") == 0:
                return MexicoBeautyFeishuSyncResponse(
                    success=True,
                    synced_count=len(feishu_records),
                    message=f"成功同步 {len(feishu_records)} 条记录"
                )
            else:
                logger.error(f"Feishu sync failed: {data}")
                raise HTTPException(status_code=500, detail=f"飞书同步失败: {data.get('msg', 'Unknown error')}")
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="飞书同步请求超时")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Feishu sync error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ========== Voice Clone (音色模仿) API ==========

class VoiceCloneAnalyzeRequest(BaseModel):
    target_lang: str = "th-TH"  # th-TH, es-ES, en-US, ja-JP, ko-KR
    video_duration: float = 0

class VoiceCloneScriptSegment(BaseModel):
    id: int
    timeRange: str
    targetContent: str
    chinese: str

class VoiceCloneAnalyzeResponse(BaseModel):
    segments: List[VoiceCloneScriptSegment]
    flaggedWords: List[dict]
    detectedSourceLanguage: Optional[str] = None

class VoiceCloneSynthesizeRequest(BaseModel):
    segments: List[VoiceCloneScriptSegment]
    voice_name: str = "Kore"
    target_lang: str = "th-TH"

class VoiceCloneSynthesizeResponse(BaseModel):
    audio_base64: str
    segment_durations: List[float]

# Supported languages for Voice Clone
VOICE_CLONE_LANGUAGES = {
    "th-TH": {"name": "泰语", "flag": "🇹🇭"},
    "es-ES": {"name": "西班牙语", "flag": "🇪🇸"},
    "en-US": {"name": "英语", "flag": "🇺🇸"},
    "ja-JP": {"name": "日语", "flag": "🇯🇵"},
    "ko-KR": {"name": "韩语", "flag": "🇰🇷"},
}

@app.post("/api/v1/voice-clone/analyze-video", response_model=VoiceCloneAnalyzeResponse)
async def voice_clone_analyze_video(
    video: UploadFile = File(...),
    target_lang: str = Form("th-TH"),
    video_duration: float = Form(0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    分析视频并生成多语种脱敏脚本。
    使用 Gemini Flash 模型分析视频内容，生成目标语言的配音脚本。
    """
    logger.info(f"Voice Clone: Analyzing video for user {user.username}, target_lang={target_lang}")
    
    # 获取配置
    config_dict = {}
    for item in db.query(SystemConfig).all():
        config_dict[item.key] = item.value
    
    api_url = config_dict.get("voice_clone_api_url") or config_dict.get("api_url") or os.getenv("DEFAULT_API_URL", "")
    api_key = config_dict.get("voice_clone_api_key") or config_dict.get("api_key") or os.getenv("DEFAULT_API_KEY", "")
    model_name = config_dict.get("voice_clone_analysis_model") or config_dict.get("analysis_model_name") or os.getenv("DEFAULT_ANALYSIS_MODEL_NAME", "")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=500, detail="API配置缺失，请在系统设置中配置API密钥")
    
    # 读取视频文件
    video_content = await video.read()
    video_b64 = base64.b64encode(video_content).decode("utf-8")
    video_mime = video.content_type or "video/mp4"
    
    lang_info = VOICE_CLONE_LANGUAGES.get(target_lang, {"name": "泰语"})
    lang_name = lang_info["name"]
    
    # 构建分析提示词
    prompt = f"""
You are a high-level Compliance AI for a global e-commerce platform.
Analyze this video ({video_duration:.1f}s) and generate a voiceover script in {lang_name} ({target_lang}) with Chinese translation.

### MANDATORY COMPLIANCE RULES:

1. **ABSOLUTE BRAND MASKING**:
   - NEVER include specific brand names (e.g., "SUCHCUTE", "TikTok", "Apple", "Lazada") in the script.
   - Replace brands with generic terms like "this brand", "this product", or "the platform" in the target language.
   - For Chinese translation, use "某品牌", "这款产品", "该平台".

2. **LANGUAGE PURITY**:
   - The 'targetContent' field MUST ONLY contain characters of the target language ({lang_name}). 
   - If Target is Thai, NO English letters. If Target is Spanish, use Spanish conventions.
   - Ensure the tone is natural for social media marketing in that locale.

3. **OUTPUT JSON**:
   - 'segments': Array of script blocks (~10s each).
   - 'flaggedWords': List original brand/risk terms found for user review.
   - 'detectedSourceLanguage': What language is being spoken in the video?

### CATEGORIES:
- **brand**: Shop/Platform/Product names.
- **price**: Monetary values.
- **risk**: Extreme claims (e.g., "Best", "Guaranteed").

Return valid JSON only with the following structure:
{{
  "segments": [
    {{"id": 1, "timeRange": "0:00-0:10", "targetContent": "...", "chinese": "..."}}
  ],
  "flaggedWords": [
    {{"word": "...", "category": "brand", "reason": "..."}}
  ],
  "detectedSourceLanguage": "..."
}}
"""
    
    # 调用 Gemini API
    api_endpoint = api_url.rstrip("/")
    if not api_endpoint.endswith("/chat/completions"):
        api_endpoint = api_endpoint + "/chat/completions"
    
    try:
        async with httpx.AsyncClient(timeout=180.0) as client:
            response = await client.post(
                api_endpoint,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": model_name,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "video_url",
                                    "video_url": {
                                        "url": f"data:{video_mime};base64,{video_b64}"
                                    }
                                },
                                {
                                    "type": "text",
                                    "text": prompt
                                }
                            ]
                        }
                    ],
                    "max_tokens": 4096,
                    "response_format": {"type": "json_object"}
                }
            )
            
            if response.status_code != 200:
                logger.error(f"Voice Clone API error: {response.status_code} - {response.text}")
                raise HTTPException(status_code=response.status_code, detail=f"视频分析失败: {response.text[:200]}")
            
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            # 解析JSON响应
            try:
                parsed = json.loads(content)
            except json.JSONDecodeError:
                json_match = re.search(r'\{[\s\S]*\}', content)
                if json_match:
                    parsed = json.loads(json_match.group())
                else:
                    raise HTTPException(status_code=500, detail="无法解析AI响应")
            
            segments = parsed.get("segments", [])
            flagged_words = parsed.get("flaggedWords", [])
            detected_lang = parsed.get("detectedSourceLanguage", None)
            
            # 品牌词脱敏过滤
            for seg in segments:
                content_text = seg.get("targetContent", "")
                chinese_text = seg.get("chinese", "")
                for fw in flagged_words:
                    if fw.get("category") == "brand":
                        word = fw.get("word", "")
                        if word:
                            content_text = re.sub(re.escape(word), "---", content_text, flags=re.IGNORECASE)
                            chinese_text = re.sub(re.escape(word), "某品牌", chinese_text, flags=re.IGNORECASE)
                seg["targetContent"] = content_text
                seg["chinese"] = chinese_text
            
            logger.info(f"Voice Clone: Analysis complete, {len(segments)} segments generated")
            
            return VoiceCloneAnalyzeResponse(
                segments=[VoiceCloneScriptSegment(**seg) for seg in segments],
                flaggedWords=flagged_words,
                detectedSourceLanguage=detected_lang
            )
            
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="视频分析请求超时，请尝试较短的视频")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Voice Clone analyze error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"视频分析失败: {str(e)}")


@app.post("/api/v1/voice-clone/synthesize-speech", response_model=VoiceCloneSynthesizeResponse)
async def voice_clone_synthesize_speech(
    request: VoiceCloneSynthesizeRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    使用 Gemini TTS 模型合成语音。
    将脚本段落转换为音频数据。
    """
    logger.info(f"Voice Clone: Synthesizing speech for user {user.username}, voice={request.voice_name}")
    
    # 获取配置
    config_dict = {}
    for item in db.query(SystemConfig).all():
        config_dict[item.key] = item.value
    
    api_url = config_dict.get("voice_clone_api_url") or config_dict.get("api_url") or os.getenv("DEFAULT_API_URL", "")
    api_key = config_dict.get("voice_clone_api_key") or config_dict.get("api_key") or os.getenv("DEFAULT_API_KEY", "")
    tts_model = config_dict.get("voice_clone_tts_model") or os.getenv("VOICE_CLONE_TTS_MODEL", "")
    
    if not api_url or not api_key:
        raise HTTPException(status_code=500, detail="API配置缺失，请在系统设置中配置API密钥")
    
    api_endpoint = api_url.rstrip("/")
    if not api_endpoint.endswith("/chat/completions"):
        api_endpoint = api_endpoint + "/chat/completions"
    
    all_audio_chunks = []
    segment_durations = []
    
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            for segment in request.segments:
                text = segment.targetContent.strip()
                if not text:
                    all_audio_chunks.append(b"")
                    segment_durations.append(0.0)
                    continue
                
                # 调用 TTS API
                tts_response = await client.post(
                    api_endpoint,
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "model": tts_model,
                        "messages": [
                            {
                                "role": "user",
                                "content": text
                            }
                        ],
                        "modalities": ["audio"],
                        "audio": {
                            "voice": request.voice_name.lower(),
                            "format": "pcm16"
                        }
                    }
                )
                
                if tts_response.status_code != 200:
                    logger.error(f"TTS API error: {tts_response.status_code} - {tts_response.text[:200]}")
                    all_audio_chunks.append(b"")
                    segment_durations.append(0.0)
                    continue
                
                tts_result = tts_response.json()
                logger.info(f"TTS API response keys: {tts_result.keys()}")
                
                audio_data = None
                choices = tts_result.get("choices", [])
                if choices:
                    message = choices[0].get("message", {})
                    logger.info(f"TTS message keys: {message.keys()}")
                    
                    audio_info = message.get("audio", {})
                    if audio_info and audio_info.get("data"):
                        audio_data = audio_info.get("data")
                        logger.info(f"Found audio in message.audio.data, length: {len(audio_data)}")
                    
                    if not audio_data:
                        content = message.get("content", [])
                        if isinstance(content, list):
                            for part in content:
                                if isinstance(part, dict):
                                    if part.get("inlineData", {}).get("data"):
                                        audio_data = part["inlineData"]["data"]
                                        logger.info(f"Found audio in content.inlineData.data")
                                        break
                                    if part.get("audio", {}).get("data"):
                                        audio_data = part["audio"]["data"]
                                        logger.info(f"Found audio in content.audio.data")
                                        break
                        elif isinstance(content, str) and len(content) > 100:
                            try:
                                import base64 as b64_test
                                b64_test.b64decode(content[:100])
                                audio_data = content
                                logger.info(f"Found audio as raw base64 string in content")
                            except:
                                pass
                    
                    if not audio_data and message.get("audio_content"):
                        audio_data = message.get("audio_content")
                        logger.info(f"Found audio in message.audio_content")
                
                if audio_data:
                    audio_bytes = base64.b64decode(audio_data)
                    all_audio_chunks.append(audio_bytes)
                    # 计算时长: PCM 16-bit, 24000 Hz, mono
                    duration = len(audio_bytes) / (24000 * 2)  # 2 bytes per sample
                    segment_durations.append(duration)
                else:
                    logger.warning(f"No audio data in TTS response for segment {segment.id}")
                    all_audio_chunks.append(b"")
                    segment_durations.append(0.0)
        
        # 合并所有音频块
        combined_audio = b"".join(all_audio_chunks)
        audio_base64 = base64.b64encode(combined_audio).decode("utf-8")
        
        logger.info(f"Voice Clone: Speech synthesis complete, total duration: {sum(segment_durations):.1f}s")
        
        return VoiceCloneSynthesizeResponse(
            audio_base64=audio_base64,
            segment_durations=segment_durations
        )
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="语音合成请求超时")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Voice Clone synthesize error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"语音合成失败: {str(e)}")
